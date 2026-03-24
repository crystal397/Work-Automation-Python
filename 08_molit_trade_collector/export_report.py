#!/usr/bin/env python3
"""
실거래가 DB → 연도별 Excel 보고서
=====================================
사용법:
  python export_report.py                        # 기본 (10년, 매매 전체)
  python export_report.py --region 강남구        # 특정 지역
  python export_report.py --years 5              # 최근 5년
  python export_report.py --min-price 50000      # 최소 거래금액 (만원)
  python export_report.py --max-price 200000     # 최대 거래금액 (만원)
"""

import argparse
import json
import sqlite3
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)

# =============================================================
#  설정
# =============================================================
DB_PATH = Path(__file__).parent / "bulk_data.db"

API_TYPE_LABELS = {
    "apt_trade":  "아파트 매매",
    "rh_trade":   "연립다세대 매매",
    "sh_trade":   "단독다가구 매매",
    "offi_trade": "오피스텔 매매",
}

REGION_NAMES = {
    "11110": "서울 종로구",      "11140": "서울 중구",
    "11170": "서울 용산구",      "11200": "서울 성동구",
    "11215": "서울 광진구",      "11230": "서울 동대문구",
    "11260": "서울 중랑구",      "11290": "서울 성북구",
    "11305": "서울 강북구",      "11320": "서울 도봉구",
    "11350": "서울 노원구",      "11380": "서울 은평구",
    "11410": "서울 서대문구",    "11440": "서울 마포구",
    "11470": "서울 양천구",      "11500": "서울 강서구",
    "11530": "서울 구로구",      "11545": "서울 금천구",
    "11560": "서울 영등포구",    "11590": "서울 동작구",
    "11620": "서울 관악구",      "11650": "서울 서초구",
    "11680": "서울 강남구",      "11710": "서울 송파구",
    "11740": "서울 강동구",
    "41111": "경기 수원시장안구", "41113": "경기 수원시권선구",
    "41115": "경기 수원시팔달구", "41117": "경기 수원시영통구",
    "41131": "경기 성남시수정구", "41133": "경기 성남시중원구",
    "41135": "경기 성남시분당구", "41150": "경기 의정부시",
    "41171": "경기 안양시만안구", "41173": "경기 안양시동안구",
    "41192": "경기 부천시원미구", "41194": "경기 부천시소사구",
    "41196": "경기 부천시오정구", "41210": "경기 광명시",
    "41220": "경기 평택시",      "41250": "경기 동두천시",
    "41271": "경기 안산시상록구", "41273": "경기 안산시단원구",
    "41281": "경기 고양시덕양구", "41285": "경기 고양시일산동구",
    "41287": "경기 고양시일산서구","41290": "경기 과천시",
    "41310": "경기 구리시",      "41360": "경기 남양주시",
    "41370": "경기 오산시",      "41390": "경기 시흥시",
    "41410": "경기 군포시",      "41430": "경기 의왕시",
    "41450": "경기 하남시",      "41461": "경기 용인시처인구",
    "41463": "경기 용인시기흥구", "41465": "경기 용인시수지구",
    "41480": "경기 파주시",      "41500": "경기 이천시",
    "41550": "경기 안성시",      "41570": "경기 김포시",
    "41591": "경기 화성시(향남)", "41593": "경기 화성시(봉담)",
    "41595": "경기 화성시(병점)", "41610": "경기 광주시",
    "41630": "경기 양주시",      "41650": "경기 포천시",
    "41670": "경기 여주시",      "41800": "경기 연천군",
    "41820": "경기 가평군",      "41830": "경기 양평군",
}

REGION_CODES = {
    "종로구": "11110", "중구": "11140", "용산구": "11170",
    "성동구": "11200", "광진구": "11215", "동대문구": "11230",
    "중랑구": "11260", "성북구": "11290", "강북구": "11305",
    "도봉구": "11320", "노원구": "11350", "은평구": "11380",
    "서대문구": "11410", "마포구": "11440", "양천구": "11470",
    "강서구": "11500", "구로구": "11530", "금천구": "11545",
    "영등포구": "11560", "동작구": "11590", "관악구": "11620",
    "서초구": "11650", "강남구": "11680", "송파구": "11710",
    "강동구": "11740",
    "수원시": ["41111", "41113", "41115", "41117"],
    "성남시": ["41131", "41133", "41135"],
    "의정부시": "41150",
    "안양시": ["41171", "41173"],
    "부천시": ["41192", "41194", "41196"],
    "광명시": "41210", "평택시": "41220", "동두천시": "41250",
    "안산시": ["41271", "41273"],
    "고양시": ["41281", "41285", "41287"],
    "과천시": "41290", "구리시": "41310", "남양주시": "41360",
    "오산시": "41370", "시흥시": "41390", "군포시": "41410",
    "의왕시": "41430", "하남시": "41450",
    "용인시": ["41461", "41463", "41465"],
    "파주시": "41480", "이천시": "41500", "안성시": "41550",
    "김포시": "41570",
    "화성시": ["41591", "41593", "41595"],
    "광주시": "41610", "양주시": "41630", "포천시": "41650",
    "여주시": "41670", "연천군": "41800", "가평군": "41820", "양평군": "41830",
}

# 연도 시트 컬럼 정의
YEAR_SHEET_COLS = [
    ("_type_label",  "거래유형"),
    ("_region_name", "지역"),
    ("umdNm",        "법정동"),
    ("roadNm",       "도로명"),
    ("aptNm",        "건물명"),
    ("excluUseAr",   "전용면적(㎡)"),
    ("floor",        "층"),
    ("buildYear",    "건축연도"),
    ("dealAmount",   "거래금액(만원)"),
    ("dealMonth",    "거래월"),
    ("dealDay",      "거래일"),
    ("dealingGbn",   "거래유형상세"),
    ("estateAgentSggNm", "중개사소재지"),
]

# =============================================================
#  스타일
# =============================================================
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
SUMMARY_FILL = PatternFill("solid", fgColor="D6E4F0")
STRIPE_A     = PatternFill("solid", fgColor="EBF3FB")  # 짝수 지역
STRIPE_B     = PatternFill("solid", fgColor="FFFFFF")  # 홀수 지역
GROUP_FILL   = PatternFill("solid", fgColor="FFF2CC")  # 지역 구분 행

THIN_SIDE   = Side(style="thin",   color="BFBFBF")
MED_SIDE    = Side(style="medium", color="2E75B6")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE,
                     top=THIN_SIDE,  bottom=THIN_SIDE)
MED_BORDER  = Border(left=MED_SIDE,  right=MED_SIDE,
                     top=MED_SIDE,   bottom=MED_SIDE)

def h1(size=13, color="FFFFFF"):
    return Font(name="Arial", bold=True, size=size, color=color)

def h2(size=10, color="FFFFFF"):
    return Font(name="Arial", bold=True, size=size, color=color)

def body(size=9, color="000000", bold=False):
    return Font(name="Arial", size=size, color=color, bold=bold)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_mid():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# =============================================================
#  데이터 로드
# =============================================================
def load_data(api_types, region_codes, years, min_price, max_price):
    if not DB_PATH.exists():
        print(f"DB 파일을 찾을 수 없습니다: {DB_PATH}")
        sys.exit(1)

    today    = date.today()
    start_yr = today.year - years + 1

    conn = sqlite3.connect(DB_PATH)

    placeholders_type   = ",".join("?" * len(api_types))
    placeholders_region = ",".join("?" * len(region_codes))

    query = f"""
        SELECT api_type, region_code, year_month, data
        FROM transactions
        WHERE api_type    IN ({placeholders_type})
          AND region_code IN ({placeholders_region})
          AND CAST(SUBSTR(year_month, 1, 4) AS INTEGER) >= ?
        ORDER BY year_month ASC
    """
    params = api_types + region_codes + [start_yr]
    rows = conn.execute(query, params).fetchall()
    conn.close()

    records = []
    for api_type, region_code, year_month, data_str in rows:
        rec = json.loads(data_str)
        rec["_api_type"]     = api_type
        rec["_region_code"]  = region_code
        rec["_year_month"]   = year_month
        rec["_year"]         = year_month[:4]
        rec["_type_label"]   = API_TYPE_LABELS.get(api_type, api_type)
        rec["_region_name"]  = REGION_NAMES.get(region_code, region_code)

        # 가격 필터
        price_str = rec.get("dealAmount", "").replace(",", "").strip()
        if price_str:
            try:
                price = int(price_str)
                if min_price and price < min_price:
                    continue
                if max_price and price > max_price:
                    continue
            except ValueError:
                pass

        records.append(rec)

    return records

# =============================================================
#  시트 1 : 요약 (연도 × 거래유형 교차표)
# =============================================================
def write_summary_sheet(wb, records, years, region_filter):
    ws = wb.active
    ws.title = "요약"
    ws.sheet_view.showGridLines = False

    today_str = date.today().strftime("%Y년 %m월 %d일")

    # 타이틀
    col_count = len(API_TYPE_LABELS) + 2  # 연도 + 유형별 + 합계
    ws.merge_cells(f"A1:{get_column_letter(col_count)}1")
    title = ws["A1"]
    title.value     = f"부동산 매매 실거래가 현황  |  기준일: {today_str}"
    title.font      = h1()
    title.fill      = HEADER_FILL
    title.alignment = center()
    title.border    = MED_BORDER
    ws.row_dimensions[1].height = 34

    # 조건
    ws.merge_cells(f"A2:{get_column_letter(col_count)}2")
    cond = ws["A2"]
    cond.value     = f"조회기간: 최근 {years}년  /  지역: {region_filter}  /  거래유형: 매매"
    cond.font      = body(size=9, color="404040")
    cond.fill      = PatternFill("solid", fgColor="D9E2F3")
    cond.alignment = left_mid()

    ws.row_dimensions[3].height = 8  # 여백

    # 연도별 × 거래유형 교차표
    ws.merge_cells(f"A4:{get_column_letter(col_count)}4")
    ws["A4"].value     = "▌ 연도별 거래 현황 (단위: 건)"
    ws["A4"].font      = h2(size=10, color="1F4E79")
    ws["A4"].fill      = SUMMARY_FILL
    ws["A4"].alignment = left_mid()

    type_keys = list(API_TYPE_LABELS.keys())
    headers   = ["연도"] + [API_TYPE_LABELS[t] for t in type_keys] + ["합계"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=ci)
        cell.value     = h
        cell.font      = h2()
        cell.fill      = SUBHEAD_FILL
        cell.alignment = center()
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.column_dimensions["A"].width = 10
    ws.row_dimensions[5].height = 20

    # 집계
    year_type_count: dict = defaultdict(lambda: defaultdict(int))
    for rec in records:
        year_type_count[rec["_year"]][rec["_api_type"]] += 1

    today_yr  = date.today().year
    start_yr  = today_yr - years + 1
    all_years = [str(y) for y in range(start_yr, today_yr + 1)]

    row = 6
    total_by_type: dict = defaultdict(int)
    for yr in all_years:
        counts = year_type_count.get(yr, {})
        row_total = sum(counts.get(t, 0) for t in type_keys)
        fill = STRIPE_A if row % 2 == 0 else STRIPE_B

        ws.cell(row=row, column=1).value = yr
        for ci, t in enumerate(type_keys, 2):
            v = counts.get(t, 0)
            ws.cell(row=row, column=ci).value = v
            total_by_type[t] += v
        ws.cell(row=row, column=len(type_keys) + 2).value = row_total

        for ci in range(1, col_count + 1):
            cell = ws.cell(row=row, column=ci)
            cell.font      = body()
            cell.fill      = fill
            cell.alignment = center()
            cell.border    = THIN_BORDER
        row += 1

    # 합계 행
    grand = sum(total_by_type.values())
    ws.cell(row=row, column=1).value = "합 계"
    for ci, t in enumerate(type_keys, 2):
        ws.cell(row=row, column=ci).value = total_by_type[t]
    ws.cell(row=row, column=len(type_keys) + 2).value = grand
    for ci in range(1, col_count + 1):
        cell = ws.cell(row=row, column=ci)
        cell.font      = body(bold=True)
        cell.fill      = SUMMARY_FILL
        cell.alignment = center()
        cell.border    = THIN_BORDER
    row += 2

    # 거래유형별 평균 금액 섹션
    ws.merge_cells(f"A{row}:{get_column_letter(col_count)}{row}")
    ws[f"A{row}"].value     = "▌ 거래유형별 평균 거래금액 (만원)"
    ws[f"A{row}"].font      = h2(size=10, color="1F4E79")
    ws[f"A{row}"].fill      = SUMMARY_FILL
    ws[f"A{row}"].alignment = left_mid()
    row += 1

    price_headers = ["거래유형", "총 건수", "평균금액", "최고금액", "최저금액"]
    for ci, h in enumerate(price_headers, 1):
        cell = ws.cell(row=row, column=ci)
        cell.value     = h
        cell.font      = h2()
        cell.fill      = SUBHEAD_FILL
        cell.alignment = center()
        cell.border    = THIN_BORDER
    row += 1

    for t in type_keys:
        prices = []
        for rec in records:
            if rec["_api_type"] != t:
                continue
            try:
                prices.append(int(rec.get("dealAmount", "0").replace(",", "")))
            except (ValueError, TypeError):
                pass

        fill = STRIPE_A if row % 2 == 0 else STRIPE_B
        avg  = f"{int(sum(prices)/len(prices)):,}" if prices else "-"
        mx   = f"{max(prices):,}" if prices else "-"
        mn   = f"{min(prices):,}" if prices else "-"
        vals = [API_TYPE_LABELS[t], len(prices), avg, mx, mn]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci)
            cell.value     = v
            cell.font      = body()
            cell.fill      = fill
            cell.alignment = center() if ci > 1 else left_mid()
            cell.border    = THIN_BORDER
        row += 1

# =============================================================
#  시트 2~ : 연도별 상세 데이터
# =============================================================
def write_year_sheet(wb, records, year):
    year_records = [r for r in records if r["_year"] == year]
    if not year_records:
        return 0

    ws = wb.create_sheet(title=year)
    ws.sheet_view.showGridLines = False

    # 지역코드 → 거래유형 순으로 정렬
    type_order = list(API_TYPE_LABELS.keys())
    year_records.sort(key=lambda r: (
        r["_region_code"],
        type_order.index(r["_api_type"]) if r["_api_type"] in type_order else 99,
        r["_year_month"],
    ))

    # 헤더
    for ci, (_, label) in enumerate(YEAR_SHEET_COLS, 1):
        cell = ws.cell(row=1, column=ci)
        cell.value     = label
        cell.font      = h2()
        cell.fill      = SUBHEAD_FILL
        cell.alignment = center()
        cell.border    = THIN_BORDER
    ws.row_dimensions[1].height = 22

    # 컬럼 너비
    col_widths = [14, 18, 12, 20, 20, 12, 6, 10, 14, 6, 6, 12, 16]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 데이터 (지역이 바뀔 때마다 배경색 교체)
    prev_region = None
    region_toggle = 0

    for ri, rec in enumerate(year_records, 2):
        if rec["_region_code"] != prev_region:
            prev_region   = rec["_region_code"]
            region_toggle = 1 - region_toggle
        fill = STRIPE_A if region_toggle == 0 else STRIPE_B

        for ci, (field, _) in enumerate(YEAR_SHEET_COLS, 1):
            val  = rec.get(field, "")
            cell = ws.cell(row=ri, column=ci)

            if field in ("excluUseAr",):
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass
            elif field in ("floor", "buildYear", "dealMonth", "dealDay"):
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    pass
            elif field == "dealAmount":
                try:
                    val = int(str(val).replace(",", ""))
                except (ValueError, TypeError):
                    pass

            cell.value     = val
            cell.font      = body()
            cell.fill      = fill
            cell.alignment = center() if ci > 2 else left_mid()
            cell.border    = THIN_BORDER

    ws.auto_filter.ref = f"A1:{get_column_letter(len(YEAR_SHEET_COLS))}1"
    ws.freeze_panes    = "C2"

    return len(year_records)

# =============================================================
#  메인
# =============================================================
def main():
    parser = argparse.ArgumentParser(description="실거래가 연도별 Excel 보고서 생성")
    parser.add_argument("--region",    nargs="+", default=[],  help="지역 이름 (예: 강남구 서초구)")
    parser.add_argument("--years",     type=int,  default=10,  help="최근 N년 (기본 10년)")
    parser.add_argument("--min-price", type=int,  default=None, help="최소 거래금액 (만원)")
    parser.add_argument("--max-price", type=int,  default=None, help="최대 거래금액 (만원)")
    parser.add_argument("--output",    default=None, help="출력 파일명")
    args = parser.parse_args()

    api_types = list(API_TYPE_LABELS.keys())

    # 지역코드 변환
    if args.region:
        codes = []
        for r in args.region:
            v = REGION_CODES.get(r)
            if v is None:
                print(f"알 수 없는 지역: {r}")
                sys.exit(1)
            codes.extend(v if isinstance(v, list) else [v])
        region_filter = " / ".join(args.region)
    else:
        codes = list(REGION_NAMES.keys())
        region_filter = "서울 + 수도권 전체"

    print(f"▶ 데이터 조회 중... (최근 {args.years}년, 매매 4종)")
    records = load_data(api_types, codes, args.years, args.min_price, args.max_price)
    print(f"  총 {len(records):,}건 로드됨")

    if not records:
        print("조건에 맞는 데이터가 없습니다.")
        sys.exit(0)

    # 연도 목록
    today_yr  = date.today().year
    start_yr  = today_yr - args.years + 1
    all_years = [str(y) for y in range(start_yr, today_yr + 1)]

    # Excel 생성
    wb = openpyxl.Workbook()
    write_summary_sheet(wb, records, args.years, region_filter)
    print(f"  └ 요약 시트 작성 완료")

    for yr in all_years:
        cnt = write_year_sheet(wb, records, yr)
        if cnt:
            print(f"  └ {yr}년: {cnt:,}건")

    # 저장
    today_str = date.today().strftime("%Y%m%d")
    out_path  = args.output or f"realestate_report_{today_str}.xlsx"
    wb.save(out_path)
    print(f"\n저장 완료: {out_path}")

if __name__ == "__main__":
    main()

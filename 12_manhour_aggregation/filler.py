"""
filler.py — 취합된 출역 데이터를 산출내역서 노임 시트에 기입
-------------------------------------------------------------
출력 시트 컬럼 구조 (formula_writer.py 기준):
  A(1) = 직종   B(2) = 성  명   C(3) = (주민번호 등, 비워둠)
  D(4)~R(18)  = 1~15일 (홀수행)
  D(4)~S(19)  = 16~31일 (짝수행)
  V(22) = 연도   X(24) = 월

사용법:
  - template/ 폴더에 산출내역서 원본(틀)을 넣어두면
    해당 파일을 복사해서 데이터를 채운 뒤 output/ 에 저장
  - template/ 에 파일이 없으면 빈 워크북을 새로 생성
"""

import shutil
import calendar
import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── 출력 시트 컬럼 상수 ──────────────────────────────────────────────────────
COL_JIKJONG  = 1   # A: 직종
COL_NAME     = 2   # B: 성  명
COL_DATE_START = 4 # D: 날짜 시작 (1일/16일)
COL_YEAR     = 22  # V: 연도
COL_MONTH    = 24  # X: 월

# 스타일
THIN = Side(style='thin', color='AAAAAA')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT_NORMAL = Font(name='맑은 고딕', size=9)
FONT_HEADER = Font(name='맑은 고딕', size=9, bold=True)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT   = Alignment(horizontal='left',   vertical='center')


# ── 기존 노임 시트 탐색 ───────────────────────────────────────────────────────
def find_noim_sheet(wb, year: int, month: int):
    """
    V1=year, X1=month 가 일치하는 노임 시트 반환.
    없으면 None.
    """
    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws['V1'].value == year and ws['X1'].value == month:
            return ws
    return None


def find_header_row(ws) -> int | None:
    """A='직종', B='성  명' 인 행 번호 반환"""
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == '직종' and ws.cell(r, 2).value == '성  명':
            return r
    return None


def find_insert_row(ws, header_row: int) -> int:
    """
    헤더행 이후에서 데이터를 삽입할 위치 반환.
    '소 계' / '합 계' 행 직전, 또는 마지막 데이터행 다음.
    """
    skip = {'소 계', '소계', '합 계', '합계', '총  계', '총계'}
    for r in range(header_row + 1, ws.max_row + 2):
        v = ws.cell(r, COL_JIKJONG).value or ws.cell(r, COL_NAME).value
        if v in skip:
            return r
    return ws.max_row + 1


# ── AD~BK 헤더 구조 복사 ─────────────────────────────────────────────────────
def _copy_adbk_headers(src_ws, dst_ws, col_ad=30, col_bk=63):
    """
    기존 노임 시트(src_ws)의 row 1~3 중 AD~BK 범위 헤더를 복사.
    formula_writer 가 row 2의 '평일'/'토요일'/'일요일' 레이블을 참조하므로 필수.
    """
    for r in range(1, 4):
        for c in range(col_ad, col_bk + 1):
            v = src_ws.cell(r, c).value
            if v is not None:
                dst_ws.cell(r, c).value = v


def _find_reference_noim_sheet(wb):
    """워크북에서 AD~BK 레이블이 있는 기존 노임 시트 반환 (없으면 None)"""
    for sname in wb.sheetnames:
        if '노임' not in sname:
            continue
        ws = wb[sname]
        for c in range(30, 64):
            if ws.cell(2, c).value in ('평일', '토요일', '일요일'):
                return ws
    return None


# ── 새 노임 시트 생성 ─────────────────────────────────────────────────────────
def create_noim_sheet(wb, year: int, month: int) -> object:
    """새 노임 시트를 생성하고 기본 헤더를 작성"""
    sheet_name = f"노임_{year % 100:02d}{month:02d}"
    ws = wb.create_sheet(title=sheet_name)

    # Row 1: 연도·월 기록 (formula_writer 참조용)
    ws.cell(1, COL_YEAR).value  = year
    ws.cell(1, COL_MONTH).value = month

    # Row 2: 날짜 헤더 레이블 (1~15 / 16~31)
    ws.cell(2, COL_NAME).value = '성  명'
    max_day = calendar.monthrange(year, month)[1]
    for day in range(1, 16):
        col = COL_DATE_START + (day - 1)
        ws.cell(2, col).value = day
        ws.cell(2, col).font = FONT_HEADER
        ws.cell(2, col).alignment = CENTER

    # Row 3: 16~31일 레이블
    for day in range(16, max_day + 1):
        col = COL_DATE_START + (day - 16)
        ws.cell(3, col).value = day
        ws.cell(3, col).font = FONT_HEADER
        ws.cell(3, col).alignment = CENTER

    # Row 4: 컬럼 헤더 (formula_writer 탐지용)
    ws.cell(4, COL_JIKJONG).value = '직종'
    ws.cell(4, COL_NAME).value    = '성  명'
    ws.cell(4, COL_JIKJONG).font  = FONT_HEADER
    ws.cell(4, COL_NAME).font     = FONT_HEADER

    # AD~BK 헤더: 기존 노임 시트에서 복사 (없으면 기본값 작성)
    ref_ws = _find_reference_noim_sheet(wb)
    if ref_ws:
        _copy_adbk_headers(ref_ws, ws)
    else:
        # 기본 레이블 (평일/토요일/일요일 각 구간 시작에 레이블)
        # 실제 산출내역서 AD~BK 구조에 맞게 조정 필요
        ws.cell(2, 30).value = '평일'    # AD
        ws.cell(2, 40).value = '토요일'  # AN
        ws.cell(2, 50).value = '일요일'  # AX

    return ws


# ── 인원 행 기입 ──────────────────────────────────────────────────────────────
def write_person_rows(ws, insert_row: int, people: list, year: int, month: int):
    """
    insert_row 위치부터 1인당 2행씩 출역 데이터 기입.
    기존 행이 있으면 밀어내지 않고 덮어씀 (템플릿이 비어있다고 가정).
    """
    max_day = calendar.monthrange(year, month)[1]
    current_row = insert_row

    for p in people:
        name = p['name']
        att  = p['attendance']

        # 홀수행 (1~15일)
        ws.cell(current_row, COL_NAME).value = name
        ws.cell(current_row, COL_NAME).font  = FONT_NORMAL
        for day in range(1, 16):
            val = att.get(day)
            if val:
                col = COL_DATE_START + (day - 1)
                c = ws.cell(current_row, col)
                c.value     = val
                c.font      = FONT_NORMAL
                c.alignment = CENTER

        # 짝수행 (16~31일)
        next_row = current_row + 1
        for day in range(16, max_day + 1):
            val = att.get(day)
            if val:
                col = COL_DATE_START + (day - 16)
                c = ws.cell(next_row, col)
                c.value     = val
                c.font      = FONT_NORMAL
                c.alignment = CENTER

        current_row += 2


# ── 메인: 템플릿에 데이터 채우기 ─────────────────────────────────────────────
def fill_template(template_dir: str | Path,
                  output_dir: str | Path,
                  data: dict) -> Path:
    """
    template/ 의 xlsx 를 복사하여 data 를 채운 뒤 output/ 에 저장.

    Args:
        template_dir : 템플릿 폴더 (xlsx 1개 있어야 함)
        output_dir   : 결과물 저장 폴더
        data         : {(year, month): [person_dict, ...]}

    Returns:
        저장된 출력 파일 경로
    """
    template_dir = Path(template_dir)
    output_dir   = Path(output_dir)
    output_dir.mkdir(exist_ok=True)

    # 템플릿 파일 결정
    templates = list(template_dir.glob("*.xlsx"))
    if templates:
        src = templates[0]
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        dst = output_dir / f"{src.stem}_완성_{stamp}.xlsx"
        shutil.copy2(src, dst)
        wb = openpyxl.load_workbook(str(dst))
        print(f"[OK] 템플릿 복사: {src.name} → {dst.name}")
    else:
        # 템플릿 없으면 빈 워크북 생성
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        dst = output_dir / f"산출내역서_노임_{stamp}.xlsx"
        wb = openpyxl.Workbook()
        # 기본 시트 제거
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        print(f"[INFO] 템플릿 없음 → 새 파일 생성: {dst.name}")

    # 연월별 데이터 기입
    for (year, month), people in sorted(data.items()):
        print(f"\n  {year}년 {month}월 ({len(people)}명) 기입 중...")

        # 기존 시트 탐색 또는 새 시트 생성
        ws = find_noim_sheet(wb, year, month)
        if ws:
            print(f"    → 기존 시트 '{ws.title}' 에 추가")
        else:
            ws = create_noim_sheet(wb, year, month)
            print(f"    → 새 시트 '{ws.title}' 생성")

        # 삽입 위치 결정
        header_row = find_header_row(ws) or (4 if ws.max_row < 5 else ws.max_row)
        insert_row = find_insert_row(ws, header_row)

        write_person_rows(ws, insert_row, people, year, month)
        print(f"    → {len(people)}명 {len(people)*2}행 기입 완료")

    wb.save(str(dst))
    print(f"\n[OK] 저장 완료: {dst}")
    return dst

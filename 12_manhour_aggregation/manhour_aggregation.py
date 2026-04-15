"""
산출내역서(돌관공사비) 노임 시트 자동화
- 각 노임 시트의 AD~BK 열에 공수별 카운트 수식을 자동으로 입력
- 날짜-요일 기반으로 평일/토요일/일요일 COUNTIF 수식 생성
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime
import shutil

# ─────────────────────────────────────────
# 설정
# ─────────────────────────────────────────
INPUT_FILE  = "1__산출내역서_돌관공사비_.xlsx"
OUTPUT_FILE = "1__산출내역서_돌관공사비_완성.xlsx"

# D~S열 상수
COL_D = column_index_from_string("D")   # 4
COL_S = column_index_from_string("S")   # 19

# AD~BK열 상수
COL_AD = column_index_from_string("AD") # 30
COL_BK = column_index_from_string("BK") # 63


# ─────────────────────────────────────────
# 날짜 → 열/줄 매핑
# ─────────────────────────────────────────
def get_date_col_info(year: int, month: int) -> dict:
    """
    각 날짜에 대해 (열번호, 줄오프셋, 요일종류) 반환
    요일종류: 0=평일, 1=토요일, 2=일요일
    줄오프셋: 0=이름행(1번줄), 1=다음행(2번줄)
    """
    result = {}
    for day in range(1, 32):
        try:
            d = datetime.date(year, month, day)
        except ValueError:
            break
        if day <= 15:
            col = COL_D + (day - 1)   # D=1일, E=2일 ...
            offset = 0
        else:
            col = COL_D + (day - 16)  # D=16일, E=17일 ...
            offset = 1
        wd = d.weekday()
        dtype = 0 if wd < 5 else (1 if wd == 5 else 2)
        result[day] = (col, offset, dtype)
    return result


# ─────────────────────────────────────────
# 연속 열 구간 묶기
# ─────────────────────────────────────────
def group_consecutive(cols: list) -> list:
    """[4,5,6,9,10,14] → [(4,6),(9,10),(14,14)]"""
    if not cols:
        return []
    cols = sorted(set(cols))
    groups = []
    start = prev = cols[0]
    for c in cols[1:]:
        if c == prev + 1:
            prev = c
        else:
            groups.append((start, prev))
            start = prev = c
    groups.append((start, prev))
    return groups


# ─────────────────────────────────────────
# COUNTIF 파트 문자열 생성
# ─────────────────────────────────────────
def make_countif_parts(col_groups: list, data_row: int, target_col: int) -> list:
    """
    [(start_col, end_col)] + 데이터행번호 + 비교대상열번호
    → ['COUNTIF($D5:$H5,AD$1)', ...] 형태 리스트
    """
    parts = []
    for (sc, ec) in col_groups:
        sc_ltr = get_column_letter(sc)
        ec_ltr = get_column_letter(ec)
        tgt_ltr = get_column_letter(target_col)
        if sc == ec:
            range_str = f"${sc_ltr}{data_row}"
        else:
            range_str = f"${sc_ltr}{data_row}:${ec_ltr}{data_row}"
        parts.append(f"COUNTIF({range_str},{tgt_ltr}$1)")
    return parts


# ─────────────────────────────────────────
# 한 사람(2행)의 AD~BK 수식 생성
# ─────────────────────────────────────────
def build_formulas(year: int, month: int,
                   row1: int,   # 이름행(1번줄)
                   ws) -> dict:
    """
    {열번호: '=수식문자열'} 반환 (AD~BK)
    """
    row2 = row1 + 1  # 2번줄

    date_info = get_date_col_info(year, month)

    # 요일별 열 분류
    평일_r1, 평일_r2 = [], []
    토_r1,   토_r2   = [], []
    일_r1,   일_r2   = [], []

    for day, (col, offset, dtype) in date_info.items():
        if dtype == 0:
            (평일_r1 if offset == 0 else 평일_r2).append(col)
        elif dtype == 1:
            (토_r1   if offset == 0 else 토_r2  ).append(col)
        else:
            (일_r1   if offset == 0 else 일_r2  ).append(col)

    # 연속 그룹
    평일_grp = group_consecutive(평일_r1) + [(c, c) for c in 평일_r2]
    # row1용 그룹 / row2용 그룹 분리해야 COUNTIF 행번호가 맞음
    평일_grp_r1 = group_consecutive(평일_r1)
    평일_grp_r2 = group_consecutive(평일_r2)
    토_grp_r1   = group_consecutive(토_r1)
    토_grp_r2   = group_consecutive(토_r2)
    일_grp_r1   = group_consecutive(일_r1)
    일_grp_r2   = group_consecutive(일_r2)

    # AD~BK 열의 구분(평일/토/일) 결정
    # Row2에서 '평일', '토요일', '일요일' 레이블로 구분
    구분 = {}   # 열번호 → '평일'|'토요일'|'일요일'
    current = None
    for j in range(COL_AD, COL_BK + 1):
        lbl = ws.cell(2, j).value
        if lbl in ('평일', '토요일', '일요일'):
            current = lbl
        if current:
            구분[j] = current

    formulas = {}
    for j in range(COL_AD, COL_BK + 1):
        cat = 구분.get(j)
        if cat == '평일':
            parts  = make_countif_parts(평일_grp_r1, row1, j)
            parts += make_countif_parts(평일_grp_r2, row2, j)
        elif cat == '토요일':
            parts  = make_countif_parts(토_grp_r1, row1, j)
            parts += make_countif_parts(토_grp_r2, row2, j)
        elif cat == '일요일':
            parts  = make_countif_parts(일_grp_r1, row1, j)
            parts += make_countif_parts(일_grp_r2, row2, j)
        else:
            continue

        if parts:
            formulas[j] = "=" + "+".join(parts)
        else:
            formulas[j] = "=0"

    return formulas


# ─────────────────────────────────────────
# 시트 헤더 블록 찾기
# ─────────────────────────────────────────
def find_header_rows(ws) -> list:
    """'직종'/'성  명' 헤더가 있는 행 번호 목록"""
    headers = []
    for i in range(1, ws.max_row + 1):
        if ws.cell(i, 1).value == '직종' and ws.cell(i, 2).value == '성  명':
            headers.append(i)
    return headers


# ─────────────────────────────────────────
# 데이터 행 목록 추출
# ─────────────────────────────────────────
def find_data_rows(ws, header_rows: list) -> list:
    """
    이름이 있는 1번줄 행 목록
    (B열에 값이 있고, '소 계' 등 제외)
    """
    skip = {'소 계', '소계', None}
    data_rows = []
    for i in range(1, ws.max_row + 1):
        b = ws.cell(i, 2).value
        if b is None or b in skip:
            continue
        if isinstance(b, str) and b.startswith('주    소'):
            continue
        # 헤더행의 '성  명' 제외
        if b == '성  명':
            continue
        # i가 짝수행이면 2번줄(스킵)
        # 1번줄 여부: 직전 헤더행 이후 홀수번째 데이터
        data_rows.append(i)

    # 2번줄(이름행+1) 제거: 이름행은 홀수 오프셋을 가짐
    # 단순히 이름(B열 값)이 있는 행만 수집하면 됨 -> 이미 위에서 처리됨
    # 단, 2번줄(b=None)은 자동 제외
    return data_rows


# ─────────────────────────────────────────
# 메인 처리
# ─────────────────────────────────────────
def process():
    print(f"파일 로딩: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE)

    노임_sheets = [s for s in wb.sheetnames if '노임' in s]
    print(f"처리할 시트: {노임_sheets}\n")

    total_cells = 0

    for sname in 노임_sheets:
        ws = wb[sname]

        # 년도/월 추출
        year  = ws['V1'].value
        month = ws['X1'].value
        if not (year and month):
            print(f"  [{sname}] 년도/월 정보 없음 - 스킵")
            continue

        print(f"[{sname}] {year}년 {month}월 처리 중...", end="")

        # 데이터 행 찾기
        data_rows = find_data_rows(ws, find_header_rows(ws))
        count = 0

        for row1 in data_rows:
            formulas = build_formulas(year, month, row1, ws)
            for col, formula in formulas.items():
                ws.cell(row1, col).value = formula
                count += 1

        total_cells += count
        print(f" {len(data_rows)}명, {count}개 셀 입력 완료")

    print(f"\n총 {total_cells}개 수식 입력 완료")
    print(f"저장 중: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print("저장 완료!")


if __name__ == "__main__":
    process()
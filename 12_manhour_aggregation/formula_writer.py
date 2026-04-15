"""
formula_writer.py — 노임 시트 AD~BK 열 COUNTIF 수식 자동 입력
-------------------------------------------------------------
(기존 manhour_aggregation.py 에서 경로 설정만 분리)

날짜-요일 기반으로 평일/토요일/일요일 COUNTIF 수식을 생성하여
각 노임 시트의 AD~BK 영역에 입력합니다.
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime

# D~S 열 상수
COL_D = column_index_from_string("D")   # 4
COL_S = column_index_from_string("S")   # 19

# AD~BK 열 상수
COL_AD = column_index_from_string("AD") # 30
COL_BK = column_index_from_string("BK") # 63


# ── 날짜 → 열/줄 매핑 ────────────────────────────────────────────────────────
def get_date_col_info(year: int, month: int) -> dict:
    result = {}
    for day in range(1, 32):
        try:
            d = datetime.date(year, month, day)
        except ValueError:
            break
        col    = COL_D + (day - 1) if day <= 15 else COL_D + (day - 16)
        offset = 0 if day <= 15 else 1
        wd     = d.weekday()
        dtype  = 0 if wd < 5 else (1 if wd == 5 else 2)
        result[day] = (col, offset, dtype)
    return result


# ── 연속 열 구간 묶기 ─────────────────────────────────────────────────────────
def group_consecutive(cols: list) -> list:
    if not cols:
        return []
    cols = sorted(set(cols))
    groups = []
    start = prev = cols[0]
    for c in cols[1:]:
        if c == prev + 1:
            prev = c
        else:
            groups.append((start, prev))
            start = prev = c
    groups.append((start, prev))
    return groups


# ── COUNTIF 파트 문자열 생성 ──────────────────────────────────────────────────
def make_countif_parts(col_groups: list, data_row: int, target_col: int) -> list:
    parts = []
    for (sc, ec) in col_groups:
        sc_ltr  = get_column_letter(sc)
        ec_ltr  = get_column_letter(ec)
        tgt_ltr = get_column_letter(target_col)
        range_str = f"${sc_ltr}{data_row}" if sc == ec else f"${sc_ltr}{data_row}:${ec_ltr}{data_row}"
        parts.append(f"COUNTIF({range_str},{tgt_ltr}$1)")
    return parts


# ── 한 사람(2행)의 AD~BK 수식 생성 ───────────────────────────────────────────
def build_formulas(year: int, month: int, row1: int, ws) -> dict:
    row2 = row1 + 1
    date_info = get_date_col_info(year, month)

    평일_r1, 평일_r2 = [], []
    토_r1,   토_r2   = [], []
    일_r1,   일_r2   = [], []

    for day, (col, offset, dtype) in date_info.items():
        if dtype == 0:
            (평일_r1 if offset == 0 else 평일_r2).append(col)
        elif dtype == 1:
            (토_r1   if offset == 0 else 토_r2  ).append(col)
        else:
            (일_r1   if offset == 0 else 일_r2  ).append(col)

    평일_grp_r1 = group_consecutive(평일_r1)
    평일_grp_r2 = group_consecutive(평일_r2)
    토_grp_r1   = group_consecutive(토_r1)
    토_grp_r2   = group_consecutive(토_r2)
    일_grp_r1   = group_consecutive(일_r1)
    일_grp_r2   = group_consecutive(일_r2)

    # AD~BK 열의 구분 읽기 (row 2)
    구분 = {}
    current = None
    for j in range(COL_AD, COL_BK + 1):
        lbl = ws.cell(2, j).value
        if lbl in ('평일', '토요일', '일요일'):
            current = lbl
        if current:
            구분[j] = current

    formulas = {}
    for j in range(COL_AD, COL_BK + 1):
        cat = 구분.get(j)
        if cat == '평일':
            parts  = make_countif_parts(평일_grp_r1, row1, j)
            parts += make_countif_parts(평일_grp_r2, row2, j)
        elif cat == '토요일':
            parts  = make_countif_parts(토_grp_r1, row1, j)
            parts += make_countif_parts(토_grp_r2, row2, j)
        elif cat == '일요일':
            parts  = make_countif_parts(일_grp_r1, row1, j)
            parts += make_countif_parts(일_grp_r2, row2, j)
        else:
            continue
        formulas[j] = "=" + "+".join(parts) if parts else "=0"

    return formulas


# ── 시트 헤더/데이터 행 탐색 ──────────────────────────────────────────────────
def find_header_rows(ws) -> list:
    return [
        i for i in range(1, ws.max_row + 1)
        if ws.cell(i, 1).value == '직종' and ws.cell(i, 2).value == '성  명'
    ]


def find_data_rows(ws, header_rows: list) -> list:
    skip = {'소 계', '소계', None}
    data_rows = []
    for i in range(1, ws.max_row + 1):
        b = ws.cell(i, 2).value
        if b is None or b in skip:
            continue
        if isinstance(b, str) and (b.startswith('주    소') or b == '성  명'):
            continue
        data_rows.append(i)
    return data_rows


# ── 메인: 워크북 전체 수식 입력 ───────────────────────────────────────────────
def process_formulas(file_path: str) -> None:
    """
    file_path 의 워크북에서 노임 시트를 찾아 COUNTIF 수식 입력 후 저장
    """
    print(f"\n[수식 입력] {file_path}")
    wb = openpyxl.load_workbook(str(file_path))

    노임_sheets = [s for s in wb.sheetnames if '노임' in s]
    if not 노임_sheets:
        print("  [WARN] '노임' 포함 시트 없음 — 수식 입력 스킵")
        wb.close()
        return

    total_cells = 0

    for sname in 노임_sheets:
        ws = wb[sname]
        year  = ws.cell(1, 22).value   # V1
        month = ws.cell(1, 24).value   # X1

        if not (year and month):
            print(f"  [{sname}] 연도/월 정보 없음 — 스킵")
            continue

        print(f"  [{sname}] {year}년 {month}월 수식 입력 중...", end="")

        data_rows = find_data_rows(ws, find_header_rows(ws))
        count = 0
        for row1 in data_rows:
            for col, formula in build_formulas(year, month, row1, ws).items():
                ws.cell(row1, col).value = formula
                count += 1

        total_cells += count
        print(f" {len(data_rows)}명, {count}개 셀 완료")

    wb.save(str(file_path))
    print(f"[OK] 수식 {total_cells}개 입력 완료 → {file_path}")

"""
PDF → Excel 변환기
  - 지정 루트 폴더 내 모든 PDF를 재귀 탐색
  - 각 PDF → 동일 경로에 동일 이름 .xlsx 저장
  - 페이지별 시트 (시트명: p1, p2, ...)
  - None 셀 → 병합 셀로 복원
  - 테이블이 없는 페이지는 텍스트 그대로 삽입

사용:
  python pdf_to_excel.py [루트폴더]
  인자 없으면 스크립트 내 ROOT 경로 사용
"""

import os
import sys
import io
import re
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent / ".env", override=False)
except ImportError:
    pass

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# 기본 루트 폴더: .env 의 PDF_ROOT 또는 CLI 인자로 지정
# .env 설정 방법: .env.example 참고
ROOT = Path(os.environ["PDF_ROOT"]) if "PDF_ROOT" in os.environ else None


# ── 병합 감지 ─────────────────────────────────────────────────────────────────

def _detect_merges(table: list[list]) -> list[tuple[int, int, int, int]]:
    """
    pdfplumber 테이블의 None 셀로부터 병합 영역을 감지.
    우선순위:
      1) 바로 위 셀이 앵커(값 있음) → 세로 연장
      2) 바로 왼쪽 셀이 앵커       → 가로 연장
      3) 위쪽 셀의 오너 계승       → 세로 체인 연장
      4) 왼쪽 셀의 오너 계승       → 가로 체인 연장
    반환: [(min_row, min_col, max_row, max_col), ...]  0-indexed
    """
    nrows = len(table)
    ncols = max((len(row) for row in table), default=0)

    def cell_val(r, c):
        return table[r][c] if c < len(table[r]) else None

    # owner[r][c] = (owner_r, owner_c)
    owner = [[None] * ncols for _ in range(nrows)]

    for r in range(nrows):
        for c in range(ncols):
            val = cell_val(r, c)
            if val is not None:
                owner[r][c] = (r, c)
            else:
                above_is_anchor = r > 0 and cell_val(r - 1, c) is not None
                left_is_anchor  = c > 0 and cell_val(r, c - 1) is not None
                if above_is_anchor:
                    owner[r][c] = (r - 1, c)
                elif left_is_anchor:
                    owner[r][c] = owner[r][c - 1]
                elif c > 0 and owner[r][c - 1] is not None and owner[r][c - 1][0] == r:
                    # 같은 행의 오너가 있으면 가로 체인 우선
                    owner[r][c] = owner[r][c - 1]
                elif r > 0 and owner[r - 1][c] is not None:
                    owner[r][c] = owner[r - 1][c]
                elif c > 0 and owner[r][c - 1] is not None:
                    owner[r][c] = owner[r][c - 1]
                else:
                    owner[r][c] = (r, c)  # 고아 셀: 자기 자신

    # 오너별로 영역 수집
    regions: dict[tuple, list[tuple]] = {}
    for r in range(nrows):
        for c in range(ncols):
            key = owner[r][c]
            regions.setdefault(key, []).append((r, c))

    merges = []
    for cells in regions.values():
        if len(cells) <= 1:
            continue
        rows = [r for r, c in cells]
        cols = [c for r, c in cells]
        min_r, max_r = min(rows), max(rows)
        min_c, max_c = min(cols), max(cols)
        if min_r < max_r or min_c < max_c:
            merges.append((min_r, min_c, max_r, max_c))

    return merges


# ── 시트 쓰기 ─────────────────────────────────────────────────────────────────

def _write_sheet(ws, table: list[list], page_num: int) -> None:
    """테이블 데이터를 워크시트에 기록 + 병합 적용."""
    if not table:
        return

    merges = _detect_merges(table)

    # 데이터 기록
    for r, row in enumerate(table, start=1):
        for c, val in enumerate(row, start=1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)

    # 병합 적용 (0-indexed → 1-indexed)
    for min_r, min_c, max_r, max_c in merges:
        start = f"{get_column_letter(min_c + 1)}{min_r + 1}"
        end   = f"{get_column_letter(max_c + 1)}{max_r + 1}"
        try:
            ws.merge_cells(f"{start}:{end}")
        except Exception:
            pass

    # 자동 열 너비 (최대 50)
    for c in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(c)
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value:
                text = str(cell.value)
                # 줄바꿈 포함 시 최대 줄 길이
                line_max = max(len(line) for line in text.splitlines()) if '\n' in text else len(text)
                max_len = max(max_len, line_max)
        ws.column_dimensions[col_letter].width = min(max_len * 1.3 + 2, 50)


def _write_text_sheet(ws, text: str) -> None:
    """텍스트를 줄 단위로 기록 (테이블 없는 페이지용)."""
    for i, line in enumerate(text.splitlines(), start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 80


# ── PDF 변환 ──────────────────────────────────────────────────────────────────

TABLE_STRATEGIES = [
    {"vertical_strategy": "lines",  "horizontal_strategy": "lines"},
    {"vertical_strategy": "lines",  "horizontal_strategy": "text"},
    {"vertical_strategy": "text",   "horizontal_strategy": "text"},
]


def _extract_tables_from_page(page) -> list[list[list]]:
    """여러 전략을 순차적으로 시도하여 테이블 추출."""
    for strat in TABLE_STRATEGIES:
        tables = page.extract_tables(strat)
        if tables:
            return tables
    return []


def convert_pdf(pdf_path: Path) -> Path:
    """PDF 한 파일을 xlsx로 변환. 저장 경로 반환."""
    xlsx_path = pdf_path.with_suffix(".xlsx")

    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    with pdfplumber.open(str(pdf_path)) as pdf:
        total = len(pdf.pages)
        for page_num, page in enumerate(pdf.pages, start=1):
            sheet_name = f"p{page_num}"
            ws = wb.create_sheet(title=sheet_name)

            tables = _extract_tables_from_page(page)

            if tables:
                # 여러 테이블이 있으면 세로로 쌓기 (빈 행 1개 간격)
                row_offset = 0
                for tbl in tables:
                    # row_offset 적용
                    if row_offset > 0:
                        row_offset += 2  # 테이블 간 빈 행

                    merges = _detect_merges(tbl)

                    for r, row in enumerate(tbl, start=row_offset + 1):
                        for c, val in enumerate(row, start=1):
                            if val is not None:
                                ws.cell(row=r, column=c, value=val)

                    for min_r, min_c, max_r, max_c in merges:
                        s = f"{get_column_letter(min_c+1)}{min_r + row_offset + 1}"
                        e = f"{get_column_letter(max_c+1)}{max_r + row_offset + 1}"
                        try:
                            ws.merge_cells(f"{s}:{e}")
                        except Exception:
                            pass

                    row_offset += len(tbl)
            else:
                # 테이블 없으면 텍스트로
                text = page.extract_text() or ""
                _write_text_sheet(ws, text)

        # 열 너비 자동 조정 (전체 시트)
        for ws in wb.worksheets:
            for c in range(1, ws.max_column + 1):
                max_len = 0
                col_letter = get_column_letter(c)
                for r in range(1, ws.max_row + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.value:
                        text = str(cell.value)
                        line_max = max((len(l) for l in text.splitlines()), default=0)
                        max_len = max(max_len, line_max)
                ws.column_dimensions[col_letter].width = min(max_len * 1.3 + 2, 50)

    if not wb.worksheets:
        wb.create_sheet("empty")

    wb.save(str(xlsx_path))
    return xlsx_path


# ── 메인 ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) > 1:
        root = Path(sys.argv[1])
    elif ROOT is not None:
        root = ROOT
    else:
        print("[오류] 루트 폴더를 지정해주세요.")
        print("  방법 1: python pdf_to_excel.py <폴더경로>")
        print("  방법 2: .env 파일에 PDF_ROOT=<폴더경로> 설정")
        sys.exit(1)

    if not root.exists():
        print(f"[오류] 폴더가 없음: {root}")
        sys.exit(1)

    pdfs = sorted(root.rglob("*.pdf"))
    print(f"총 {len(pdfs)}개 PDF 발견\n")

    ok = fail = 0
    for pdf_path in pdfs:
        rel = pdf_path.relative_to(root)
        try:
            xlsx = convert_pdf(pdf_path)
            print(f"  ✓ {rel}")
            ok += 1
        except Exception as e:
            print(f"  ✗ {rel}  [{e}]")
            fail += 1

    print(f"\n완료: {ok}개 성공, {fail}개 실패")


if __name__ == "__main__":
    main()

"""
readers/common.py — 업체별 노무비 시트 자동 파서
--------------------------------------------------
지원 포맷:
  A) 금풍건설/기장원자로 형식  : 날짜 E열(5)~S열 시작, 행1 텍스트에서 연/월 추출
  B) 대우건설 형식             : 날짜 C열(3)~Q열 시작, 시트명에서 연/월 추출

공통 구조:
  - 1인당 2행 (홀수행=이름+1~15일, 짝수행=16~31일)
  - 날짜 헤더행 (1,2,...,15 연속 숫자) 자동 탐지
"""

import re
import datetime
import openpyxl
from pathlib import Path


# ── 연/월 탐지 ────────────────────────────────────────────────────────────────
def _yearmonth_from_sheetname(name: str):
    """'25년8월', '2022년6월' 형식"""
    m = re.search(r'(\d{2,4})년\s*(\d{1,2})월', name)
    if m:
        y = int(m.group(1))
        year = 2000 + y if y < 100 else y
        return year, int(m.group(2))
    return None, None


def _yearmonth_from_cells(ws) -> tuple:
    """'일용노무비지급명세서(2022년 06월)' 같은 셀 텍스트에서 추출 (상위 5행)"""
    for r in range(1, 6):
        for c in range(1, 40):
            v = str(ws.cell(r, c).value or '')
            m = re.search(r'(\d{4})년\s*(\d{1,2})월', v)
            if m:
                return int(m.group(1)), int(m.group(2))
    return None, None


def detect_yearmonth(ws) -> tuple:
    year, month = _yearmonth_from_sheetname(ws.title)
    if year:
        return year, month
    return _yearmonth_from_cells(ws)


# ── 날짜 헤더 탐지 ────────────────────────────────────────────────────────────
def find_date_header(ws):
    """
    1~15 숫자가 연속으로 있는 헤더행 탐지
    Returns: (header_row1, header_row2, date_start_col) or (None, None, None)
    """
    for r in range(1, 20):
        # 이 행의 각 셀 값을 수집
        cell_map = {}
        for c in range(1, 35):
            v = ws.cell(r, c).value
            if isinstance(v, int) and 1 <= v <= 15:
                cell_map[v] = c

        # 1~15 연속 존재 확인
        if all(i in cell_map for i in range(1, 16)):
            date_start_col = cell_map[1]
            # 바로 다음 행에 16 이상 숫자가 같은 위치에 있는지 확인
            next_r = r + 1
            v16 = ws.cell(next_r, date_start_col).value
            if isinstance(v16, int) and v16 in (16, 17):
                return r, next_r, date_start_col

    return None, None, None


# ── 이름 열 탐지 ─────────────────────────────────────────────────────────────
def find_name_col(ws, date_start_col: int, header_rows: list) -> int:
    """
    날짜 열 왼쪽에서 '성명' 헤더 또는 한국어 이름 패턴으로 이름 열 탐지
    """
    # 방법1: 헤더 행들에서 '성  명' / '성명' 키워드
    for r in header_rows:
        for c in range(1, date_start_col):
            v = str(ws.cell(r, c).value or '')
            if re.search(r'성\s*명', v):
                return c

    # 방법2: 데이터 영역에서 한글 이름 패턴 (2~5자 한글)
    start_search = max(header_rows) + 1
    for r in range(start_search, min(start_search + 30, ws.max_row + 1)):
        for c in range(1, date_start_col):
            v = ws.cell(r, c).value
            if (isinstance(v, str)
                    and 2 <= len(v.strip()) <= 5
                    and re.search(r'^[가-힣]+$', v.strip())):
                return c

    return 2  # fallback: B열


# ── 스킵 판단 ─────────────────────────────────────────────────────────────────
_SKIP_NAMES = {
    '성  명', '성명', '소 계', '소계', '합 계', '합계',
    '총  계', '총계', '합    계', '소    계',
}

def _is_valid_name(v) -> bool:
    if not isinstance(v, str):
        return False
    v = v.strip()
    if not v or v in _SKIP_NAMES:
        return False
    # 숫자만이거나 특수문자만인 경우 제외
    if re.match(r'^[\d\s\-_.*]+$', v):
        return False
    return True


# ── 시트 한 장 파싱 ───────────────────────────────────────────────────────────
def read_manhour_sheet(ws) -> list:
    """
    노무비 시트 한 장에서 출역 데이터 추출

    Returns:
        list of dict:
            name       (str)  : 성명
            year       (int)  : 연도
            month      (int)  : 월
            attendance (dict) : {day(int): 공수(float)}
    """
    year, month = detect_yearmonth(ws)
    if not (year and month):
        return []

    hrow1, hrow2, date_start_col = find_date_header(ws)
    if not hrow1:
        return []

    name_col = find_name_col(ws, date_start_col, [hrow1, hrow2])

    # 해당 월의 마지막 날 계산 (31일 없는 달 처리)
    import calendar
    max_day = calendar.monthrange(year, month)[1]

    results = []
    r = hrow2 + 1  # 데이터 시작 행

    while r <= ws.max_row:
        name_val = ws.cell(r, name_col).value
        if not _is_valid_name(name_val):
            r += 1
            continue

        name = name_val.strip()
        attendance = {}

        # 1~15일 (현재 행)
        for day in range(1, 16):
            col = date_start_col + (day - 1)
            v = ws.cell(r, col).value
            if v is not None and v != 0 and isinstance(v, (int, float)):
                attendance[day] = float(v)

        # 16~31일 (다음 행)
        next_r = r + 1
        if next_r <= ws.max_row:
            for day in range(16, max_day + 1):
                col = date_start_col + (day - 16)
                v = ws.cell(next_r, col).value
                if v is not None and v != 0 and isinstance(v, (int, float)):
                    attendance[day] = float(v)

        if attendance:
            results.append({
                'name': name,
                'year': year,
                'month': month,
                'attendance': attendance,
            })

        r += 2  # 1인당 2행

    return results


# ── 파일 전체 파싱 ────────────────────────────────────────────────────────────
# 노무비 시트 판단용: 이 키워드가 시트명에 있으면 노무비 시트로 간주
_MANHOUR_KEYWORDS = ['노무비', '노임', '일용', '직영', '용역']
# 이 키워드가 시트명에 있으면 제외 (자재비, 목차 등)
_SKIP_KEYWORDS = ['목차', '자재', '검수', '기성', '정산', '청구', '분개']


def is_manhour_sheet(ws) -> bool:
    name = ws.title
    if any(k in name for k in _SKIP_KEYWORDS):
        return False
    if any(k in name for k in _MANHOUR_KEYWORDS):
        return True
    # 키워드 없어도 날짜 헤더가 있으면 파싱 시도
    hrow1, _, _ = find_date_header(ws)
    return hrow1 is not None


def read_file(path: Path) -> list:
    """
    엑셀 파일 전체에서 노무비 데이터 추출

    Returns: read_manhour_sheet() 결과의 합산 리스트
    """
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        print(f"  [ERR] {path.name} 열기 실패: {e}")
        return []

    all_records = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        if not is_manhour_sheet(ws):
            continue
        records = read_manhour_sheet(ws)
        if records:
            print(f"    [{sname}] {len(records)}명 추출")
            all_records.extend(records)

    return all_records

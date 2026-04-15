"""
대한건설협회 건설업체 정보 수집기
https://www.cak.or.kr/lay1/program/S1T56C252/biz/bizSearch.do

실행:
    python cak_crawler.py

출력:
    cak_business_list.csv   — 목록 데이터 (전체 ~18,451건)
    cak_business_detail.csv — 상세 데이터 (업종 및 등록번호 포함)
"""

import asyncio
import sys
import io
import time
from pathlib import Path

import aiohttp
from playwright.async_api import async_playwright
from tqdm import tqdm

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

BASE_URL   = "https://www.cak.or.kr"
LIST_URL   = f"{BASE_URL}/biz/ajax/srchBizList.do"
DETAIL_URL = f"{BASE_URL}/biz/ajax/bizSearchDetailView.do"
# 이 사이트는 무한스크롤 방식: cpage=N, firstIndex=1 → 누적으로 N×PAGE_UNIT건 반환
# 유효한 pageUnit 값: 10, 30, 50 (브라우저 드롭다운 기준)
PAGE_UNIT   = 50          # 유효한 최대값 (10, 30, 50 중 선택)
CONCURRENCY = 10          # 상세 조회 동시 요청 수
DELAY_SEC   = 0.3         # 상세 요청 간 최소 딜레이 (초)
TEST_CNT    = 0           # 0 = 전체 수집, N>0 = N건만 수집 (테스트용)

OUTPUT_XLSX = Path("대한건설협회_건설업체정보.xlsx")


# ── Step 1: Playwright로 세션 인증 정보 획득 ─────────────────────────────────

async def get_credentials() -> tuple[str, str]:
    """CSRF 토큰 + JSESSIONID 반환"""
    print("[1/4] 브라우저 실행 — 세션 초기화 중...")
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context()
        page = await ctx.new_page()

        captured_csrf = {}

        async def on_request(req):
            if "srchBizList" in req.url:
                captured_csrf["token"] = req.headers.get("x-csrf-token", "")

        page.on("request", on_request)

        await page.goto(
            f"{BASE_URL}/lay1/program/S1T56C252/biz/bizSearch.do",
            wait_until="networkidle",
        )
        await page.click('button[onclick="fn_search()"]')
        await page.wait_for_timeout(3000)

        # 쿠키 컨텍스트에서 직접 추출
        cookies = await ctx.cookies()
        jsessionid = next(
            (c["value"] for c in cookies if c["name"] == "JSESSIONID"), ""
        )

        await browser.close()

    csrf = captured_csrf.get("token", "")
    if not csrf or not jsessionid:
        raise RuntimeError(
            f"인증 정보 획득 실패 — CSRF={bool(csrf)}, JSESSIONID={bool(jsessionid)}"
        )

    cookie = f"JSESSIONID={jsessionid}"
    print(f"  CSRF  : {csrf[:20]}...")
    print(f"  Cookie: {cookie[:30]}...")
    return csrf, cookie


# ── Step 2: 목록 수집 ─────────────────────────────────────────────────────────

def _list_headers(csrf: str, cookie: str) -> dict:
    return {
        "Content-Type":    "application/x-www-form-urlencoded; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "X-CSRF-TOKEN":     csrf,
        "ajax":             "true",
        "Referer":          f"{BASE_URL}/lay1/program/S1T56C252/biz/bizSearch.do",
        "Cookie":           cookie,
    }


def _list_payload(total_pages: int) -> dict:
    """
    이 사이트는 무한스크롤 방식:
      cpage=N, pageUnit=50, firstIndex=1 → 첫 번째부터 N×50건 누적 반환
    전체 건수를 한 번에 얻으려면 cpage = ceil(totCnt / PAGE_UNIT) 으로 요청.
    """
    return {
        "srh_closed":        "commBiz",
        "srh_sort":          "sigong",
        "srh_sort_dir":      "DESC",
        "srh_sangho":        "",
        "srh_nm":            "",
        "srh_type_city":     "1",
        "srh_city":          "",
        "srh_sikun":         "",
        "srh_sigong":        "",
        "handoamt_min":      "",
        "handoamt_max":      "",
        "srh_upjong":        "",
        "srh_upjong_detail": "",
        "handoamt_min2":     "",
        "handoamt_max2":     "",
        "srh_young":         "",
        "srh_sangsi":        "",
        "srh_comp_type":     "",
        "srh_etc2":          "",
        "srh_etc3":          "",
        "srh_etc4":          "",
        "srh_etc5":          "",
        "cpage":             str(total_pages),
        "pageUnit":          str(PAGE_UNIT),
        "firstIndex":        "1",
    }


async def collect_list(csrf: str, cookie: str) -> list[dict]:
    """
    전체 목록 수집 — 무한스크롤 누적 방식.
    1단계: cpage=1로 총 건수(totCnt) 확인
    2단계: cpage=ceil(totCnt/PAGE_UNIT)으로 전체 한 번에 수집
    """
    print("[2/4] 목록 수집 시작...")

    connector = aiohttp.TCPConnector(ssl=False)
    async with aiohttp.ClientSession(connector=connector) as session:

        # ── Step A: 총 건수 확인 ──────────────────────────────────────────────
        async with session.post(
            LIST_URL,
            headers=_list_headers(csrf, cookie),
            data=_list_payload(1),
            timeout=aiohttp.ClientTimeout(total=30),
        ) as resp:
            first = await resp.json(content_type=None)

        tot_cnt = int(first.get("totCnt", 0))
        if tot_cnt == 0:
            raise RuntimeError(
                f"총 건수 0 — 세션이 만료됐거나 API 응답 이상.\n응답: {first}"
            )

        target_cnt = min(tot_cnt, TEST_CNT) if TEST_CNT > 0 else tot_cnt
        total_pages = (target_cnt + PAGE_UNIT - 1) // PAGE_UNIT
        print(f"  총 {tot_cnt:,}건 → {target_cnt:,}건 수집 예정 (cpage={total_pages})")

        # ── Step B: 전체 한 번에 수집 ─────────────────────────────────────────
        print("  전체 목록 요청 중... (서버 처리 시간에 따라 수십 초 소요)")
        async with session.post(
            LIST_URL,
            headers=_list_headers(csrf, cookie),
            data=_list_payload(total_pages),
            timeout=aiohttp.ClientTimeout(total=300),   # 최대 5분 대기
        ) as resp:
            data = await resp.json(content_type=None)

    all_items = data.get("bizList", [])
    print(f"  수집 완료: {len(all_items):,}건")
    return all_items


# ── Step 3: 상세 정보 수집 ────────────────────────────────────────────────────

async def fetch_detail(
    session: aiohttp.ClientSession,
    csrf: str,
    cookie: str,
    hwno: str,
    semaphore: asyncio.Semaphore,
) -> dict:
    async with semaphore:
        await asyncio.sleep(DELAY_SEC)
        try:
            async with session.post(
                DETAIL_URL,
                headers=_list_headers(csrf, cookie),
                data={"hwno": hwno},
                timeout=aiohttp.ClientTimeout(total=30),
            ) as resp:
                data = await resp.json(content_type=None)
                return {"hwno": hwno, **data} if isinstance(data, dict) else {"hwno": hwno, "raw": data}
        except Exception as e:
            return {"hwno": hwno, "error": str(e)}


async def collect_details(
    csrf: str, cookie: str, hwno_list: list[str]
) -> list[dict]:
    print(f"[3/4] 상세 정보 수집 — {len(hwno_list):,}건...")
    semaphore = asyncio.Semaphore(CONCURRENCY)
    results = []

    connector = aiohttp.TCPConnector(ssl=False)
    async with aiohttp.ClientSession(connector=connector) as session:
        tasks = [
            fetch_detail(session, csrf, cookie, hwno, semaphore)
            for hwno in hwno_list
        ]
        with tqdm(total=len(tasks), desc="  상세 페이지", unit="건") as bar:
            for coro in asyncio.as_completed(tasks):
                result = await coro
                results.append(result)
                bar.update(1)

    print(f"  수집 완료: {len(results):,}건")
    return results


# ── 업종 코드 매핑 ────────────────────────────────────────────────────────────

# 대한건설협회 종합건설업 업종코드 → 업종명
UPJONG_MAP = {
    "10": "토목건축공사업",
    "20": "토목공사업",
    "40": "건축공사업",
    "50": "산업·환경설비공사업",
    "64": "조경공사업",
}

# 토목건축공사업(10) 등록 시 → 토목공사업·건축공사업 자동 포함
UPJONG_10_IMPLIED = ["토목공사업", "건축공사업"]

MEMBER_MAP = {
    "10": "정회원",
    "20": "준회원",
    "35": "특별회원",
}


# ── Step 4: xlsx 저장 ─────────────────────────────────────────────────────────

def _get_comInfo1(d: dict) -> dict:
    """detail 응답에서 comInfo1 dict 추출"""
    c1 = d.get("comInfo1") or {}
    if isinstance(c1, list):
        c1 = c1[0] if c1 else {}
    return c1 if isinstance(c1, dict) else {}


def save_xlsx(items: list[dict], details: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # ── Sheet 1: 건설업체목록 ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "건설업체목록"

    headers1 = ["번호", "주업종", "등록번호", "상호", "대표자", "지역",
                "시공능력평가액(천원)", "회원여부"]
    ws1.append(headers1)

    for item in items:
        member_raw = str(item.get("member", "")).strip()
        ws1.append([
            item.get("no", ""),
            item.get("upjongnm", ""),
            str(item.get("upjongno", "")).strip(),
            item.get("sangho", ""),
            item.get("nm", ""),
            item.get("citynm", ""),
            item.get("handoamt", ""),
            MEMBER_MAP.get(member_raw, member_raw),
        ])

    # ── Sheet 2: 업종및등록번호 ───────────────────────────────────────────────
    ws2 = wb.create_sheet("업종및등록번호")
    headers2 = ["번호", "상호", "업종명", "등록번호"]
    ws2.append(headers2)

    # hwno → detail 조회 테이블
    detail_map: dict[str, dict] = {}
    for d in details:
        hwno = str(d.get("hwno", "")).strip()
        if hwno:
            detail_map[hwno] = d

    for item in items:
        hwno    = str(item.get("hwno", "")).strip()
        no      = item.get("no", "")
        sangho  = item.get("sangho", "")
        d       = detail_map.get(hwno, {})
        c1      = _get_comInfo1(d)

        # detail이 없으면 목록 데이터의 주업종만 표시
        if not c1:
            upjongnm = item.get("upjongnm", "")
            upjongno = str(item.get("upjongno", "")).strip()
            reg_str  = f"제 {upjongno}호" if upjongno else "-"
            ws2.append([no, sangho, upjongnm, reg_str])
            continue

        # 업종코드별 등록번호 전개
        for code, upjong_nm in UPJONG_MAP.items():
            lic_val = str(c1.get(f"licence{code}", "") or "").strip()
            if not lic_val:
                continue

            reg_str = f"제 {lic_val}호"
            ws2.append([no, sangho, upjong_nm, reg_str])

            # 토목건축공사업(10) → 토목·건축 포함 표시
            if code == "10":
                for implied in UPJONG_10_IMPLIED:
                    ws2.append([no, sangho, implied, "(토목건축공사업 포함)"])

    # ── 공통 스타일 ───────────────────────────────────────────────────────────
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    center   = Alignment(horizontal="center", vertical="center")

    for ws in [ws1, ws2]:
        # 헤더 스타일
        for cell in ws[1]:
            cell.fill      = hdr_font and hdr_fill
            cell.font      = hdr_font
            cell.alignment = center
        ws.row_dimensions[1].height = 18
        ws.freeze_panes = "A2"

        # 컬럼 너비 자동 조정 (최대 40)
        for col in ws.columns:
            max_w = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_w + 3, 40)

    wb.save(OUTPUT_XLSX)
    print(f"\n  저장 완료: {OUTPUT_XLSX}")
    print(f"  - 건설업체목록   : {ws1.max_row - 1:,}행")
    print(f"  - 업종및등록번호 : {ws2.max_row - 1:,}행")


# ── 메인 ──────────────────────────────────────────────────────────────────────

async def main():
    start = time.time()
    print("=" * 60)
    print("  대한건설협회 건설업체 정보 수집기")
    print("=" * 60)

    # 1. 세션 인증
    csrf, cookie = await get_credentials()

    # 2. 목록 수집
    items = await collect_list(csrf, cookie)

    # 3. 상세 수집 (hwno 추출)
    hwno_list = [str(it.get("hwno", "")).strip() for it in items if it.get("hwno")]
    hwno_list = list(dict.fromkeys(hwno_list))  # 중복 제거
    print(f"  고유 hwno: {len(hwno_list):,}개")

    details = await collect_details(csrf, cookie, hwno_list)

    # 4. xlsx 저장
    print("[4/4] xlsx 저장...")
    save_xlsx(items, details)

    elapsed = time.time() - start
    print(f"\n완료! 소요시간: {elapsed/60:.1f}분")
    print(f"  → {OUTPUT_XLSX}")


if __name__ == "__main__":
    asyncio.run(main())

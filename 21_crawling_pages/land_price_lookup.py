"""
개별공시지가 조회 스크립트 (vworld 개별공시지가속성조회 API)
- 입력: 토지목록_수원화성_팔달문_주변_성곽복원_사업대상지__coding.xlsx
- 출력: 원본 파일의 '비 고' 컬럼에 개별공시지가(원/㎡) 기입
"""

import requests
import xml.etree.ElementTree as ET
import re
import time
import os
import shutil
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment

load_dotenv()

# =============================================
# ⚙️ 설정
# =============================================
VWORLD_KEY  = os.getenv("VWORLD_API_KEY")
INPUT_FILE  = "토지목록(수원화성 팔달문 주변 성곽복원 사업대상지)_coding.xlsx"
OUTPUT_FILE = "토지목록_공시지가포함.xlsx"

if not VWORLD_KEY:
    raise ValueError("❌ .env 파일에 VWORLD_API_KEY가 설정되지 않았습니다.")

# =============================================
# 위치(동) → 법정동코드 매핑 (수원시 팔달구 41115)
# PNU = 시군구코드(5) + 읍면동코드(5) + 산여부(1) + 본번(4) + 부번(4)
# =============================================
DONG_CODE = {
    "팔달로 2가": "12100",  # 실제 확인된 법정동코드
    "팔달로2가":  "12100",
    "팔달로 3가": "12200",
    "팔달로3가":  "12200",
    "남창동":     "12300",
    "영동":       "12400",
}
SIGUNGU_CODE = "41115"


# 동코드로 찾히지 않는 예외 필지 직접 지정
PNU_OVERRIDE = {
    ("팔달로3가",  "133"): "4111512100101330000",  # 실제로 팔달로2가 코드
    ("팔달로 3가", "133"): "4111512100101330000",
}

def build_pnu(dong: str, jibun: str):
    dong = dong.strip()
    jibun_clean = str(jibun).strip()

    # 예외 필지 직접 지정
    override = PNU_OVERRIDE.get((dong, jibun_clean))
    if override:
        return override

    code = DONG_CODE.get(dong)
    if not code:
        return None

    if "-" in jibun_clean:
        bobn, bubn = jibun_clean.split("-", 1)
    else:
        bobn, bubn = jibun_clean, "0"

    bobn = re.sub(r"[^0-9]", "", bobn).zfill(4)
    bubn = re.sub(r"[^0-9]", "", bubn).zfill(4)
    return f"{SIGUNGU_CODE}{code}1{bobn}{bubn}"


# =============================================
# vworld 개별공시지가속성조회 API
# 엔드포인트: http://api.vworld.kr/ned/data/getIndvdLandPriceAttr
# =============================================
BASE_URL = "http://api.vworld.kr/ned/data/getIndvdLandPriceAttr"

def fetch_land_price(pnu: str) -> dict:
    """PNU로 개별공시지가 조회 → 가장 최근 기준년도 반환"""
    params = {
        "key":      VWORLD_KEY,
        "pnu":      pnu,
        "format":   "json",
        "numOfRows": "100",  # 전체 연도 가져와서 최신값 선택
        "pageNo":   "1",
    }
    try:
        resp = requests.get(BASE_URL, params=params, timeout=10)
        resp.raise_for_status()

        # JSON 응답 파싱
        try:
            data = resp.json()
        except Exception:
            # XML fallback
            return _parse_xml(resp.text)

        # 응답 구조 탐색
        items = []

        # indvdLandPrices.field 구조 파싱 (확인된 실제 응답 구조)
        fields = (data.get("indvdLandPrices") or {}).get("field", [])
        if isinstance(fields, dict):
            fields = [fields]

        for f in fields:
            year  = str(f.get("stdrYear") or "")
            price = f.get("pblntfPclnd") or ""   # 확인된 실제 공시지가 필드명
            if year and price:
                items.append((year, str(price)))

        if items:
            items.sort(key=lambda x: x[0], reverse=True)
            return {"ok": True, "year": items[0][0], "price": items[0][1]}

        return {"ok": False, "error": f"데이터없음 (totalCount:{data.get('response',{}).get('totalCount','?')})"}

    except Exception as e:
        return {"ok": False, "error": str(e)[:80]}


def _parse_xml(text: str) -> dict:
    """XML 응답 fallback 파서"""
    try:
        root = ET.fromstring(text)
        items = []
        for item in root.iter("item"):
            year  = item.findtext("stdrYear") or item.findtext("stdr_year") or ""
            price = item.findtext("lndcgPblntfPc") or item.findtext("indvdLandPc") or ""
            if year and price:
                items.append((year.strip(), price.strip()))
        if items:
            items.sort(key=lambda x: x[0], reverse=True)
            return {"ok": True, "year": items[0][0], "price": items[0][1]}
    except Exception as e:
        return {"ok": False, "error": f"XML파싱실패: {e}"}
    return {"ok": False, "error": "XML데이터없음"}


# =============================================
# 메인
# =============================================
def main():
    shutil.copy2(INPUT_FILE, OUTPUT_FILE)
    wb = load_workbook(OUTPUT_FILE)
    ws = wb["토지목록"]

    # 헤더 행 찾기
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[0] == "연번":
            header_row = i
            break
    if not header_row:
        print("❌ 헤더 행(연번)을 찾을 수 없습니다.")
        return

    headers = [ws.cell(header_row, c).value for c in range(1, 7)]
    col_map = {v: i+1 for i, v in enumerate(headers) if v}

    bigo_col  = next((v for k, v in col_map.items() if k and "고" in k), None)
    loc_col   = col_map.get("위치")
    jibun_col = col_map.get("지번")
    seq_col   = col_map.get("연번")

    if not bigo_col:
        print("❌ '비고' 컬럼을 찾을 수 없습니다.")
        return

    data_start = header_row + 1
    total = ws.max_row - data_start + 1
    success = fail = skip = 0

    print(f"\n📋 총 {total}건 조회 시작\n")

    for row_num in range(data_start, ws.max_row + 1):
        seq   = ws.cell(row_num, seq_col).value
        dong  = ws.cell(row_num, loc_col).value
        jibun = ws.cell(row_num, jibun_col).value

        if not seq or not dong or not jibun:
            skip += 1
            continue

        pnu = build_pnu(str(dong), str(jibun))
        if not pnu:
            print(f"  [{seq}] '{dong}' → DONG_CODE에 없는 동명")
            ws.cell(row_num, bigo_col).value = "동코드없음"
            fail += 1
            continue

        print(f"  [{seq}/{total}] {dong} {jibun} (PNU:{pnu}) ...", end=" ", flush=True)
        result = fetch_land_price(pnu)
        cell = ws.cell(row_num, bigo_col)

        if result["ok"]:
            price_val = int(float(str(result["price"])))
            cell.value = price_val
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
            print(f"✅ {result['year']}년  {price_val:,}원/㎡")
            success += 1
        else:
            cell.value = "조회실패"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            print(f"❌ {result['error']}")
            fail += 1

        time.sleep(0.3)

    wb.save(OUTPUT_FILE)

    print(f"\n{'='*55}")
    print(f"✅ 성공: {success}건  ❌ 실패: {fail}건  ⏭️ 건너뜀: {skip}건")
    print(f"💾 저장 완료: {OUTPUT_FILE}")
    print(f"{'='*55}")


if __name__ == "__main__":
    main()
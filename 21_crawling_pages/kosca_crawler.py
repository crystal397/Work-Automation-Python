"""
대한전문건설협회 전문건설현황 수집기
https://www.kosca.or.kr/const/proconst/search.do?menuId=MENU001135

동작:
  시도(11~39)를 하나씩 POST 검색 → 페이지 순회 → 테이블 파싱

출력:
  kosca_건설업체현황.xlsx  (시트: 전문건설업체현황)
  컬럼: 시도, 상호, 대표자, 주소, 업종, 평가액
"""

import re
import sys
import io
import json
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ── 설정 ──────────────────────────────────────────────────────────────────────

BASE_URL        = "https://www.kosca.or.kr"
SEARCH_URL      = f"{BASE_URL}/const/proconst/search.do"
OUTPUT          = Path("kosca_건설업체현황.xlsx")
CHECKPOINT_FILE = Path("kosca_checkpoint.json")

SIDO_MAP = {
    "11": "서울", "21": "부산", "22": "대구", "23": "인천",
    "24": "광주", "25": "대전", "26": "울산", "29": "세종",
    "31": "경기", "32": "강원", "33": "충북", "34": "충남",
    "35": "전북", "36": "전남", "37": "경북", "38": "경남", "39": "제주",
}

REQ_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Referer": f"{SEARCH_URL}?menuId=MENU001135",
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8",
}

DELAY        = 0.2   # 요청 간 딜레이 (초)
PAGE_SIZE    = 10    # 페이지당 레코드 수 (사이트 고정값)
SIDO_WORKERS = 4     # 병렬 시도 수집 스레드 수
PAGE_WORKERS = 8     # 시도별 병렬 페이지 수집 스레드 수


# ── 유틸 ──────────────────────────────────────────────────────────────────────

def _make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(REQ_HEADERS)
    return s


def _post_search(session: requests.Session, sido: str) -> tuple[str, int]:
    """
    시도 검색 POST → 리디렉션 후 기준 GET URL + 총 페이지 수 반환.
    """
    data = {
        "menuId":    "MENU001135",
        "company":   "",
        "name1":     "",
        "sido":      sido,
        "region":    "",
        "gubun":     "industry",
        "upjong":    "",
        "detailChk": "no",
        "priceFrom": "",
        "priceTo":   "",
    }
    resp = session.post(SEARCH_URL, data=data, allow_redirects=True, timeout=30)
    base_url = resp.url  # 리디렉션된 암호화 파라미터 포함 GET URL

    # 총 페이지: 페이지네이션 'goPage(N)' 중 최대값
    soup = BeautifulSoup(resp.text, "html.parser")
    total_pages = _parse_total_pages(soup)
    return base_url, total_pages


def _parse_total_pages(soup: BeautifulSoup) -> int:
    paging = soup.find("ul", class_="paging")
    if not paging:
        return 1
    max_page = 1
    for a in paging.find_all("a", href=True):
        m = re.search(r"goPage\(([\d.]+)\)", a["href"])
        if m:
            max_page = max(max_page, int(float(m.group(1))))
    return max_page


def _fetch_page(session: requests.Session, base_url: str, page: int) -> list[dict]:
    """recordPage=N 파라미터를 추가하여 GET, 테이블[1] 파싱."""
    # base_url 에 recordPage 가 이미 있으면 교체, 없으면 추가
    if "recordPage=" in base_url:
        url = re.sub(r"recordPage=\d+", f"recordPage={page}", base_url)
    else:
        url = f"{base_url}&recordPage={page}"
    try:
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
        return _parse_table(resp.text)
    except Exception as e:
        tqdm.write(f"    [오류] p{page}: {e}")
        return []


def _split_by_br(td) -> list[str]:
    """td 안의 <br> 태그를 기준으로 텍스트를 분리하여 리스트 반환.

    이 사이트는 <br>텍스트<br>텍스트</br></br> 식으로 <br>을 중첩해서 닫아
    BeautifulSoup이 중첩 트리 구조로 파싱한다.
    → <br>을 만날 때마다 현재 버퍼를 확정하고 재귀적으로 자식을 처리."""
    parts: list[str] = []
    buf: list[str] = []

    def _walk(node) -> None:
        if getattr(node, "name", None) == "br":
            # <br> 직전까지 모인 텍스트를 한 항목으로 확정
            text = " ".join(buf).strip()
            if text:
                parts.append(text)
            buf.clear()
            # <br> 내부 자식도 재귀 처리 (중첩 케이스)
            for child in node.children:
                _walk(child)
        elif hasattr(node, "children"):
            for child in node.children:
                _walk(child)
        else:
            chunk = str(node).strip()
            if chunk:
                buf.append(chunk)

    for elem in td.children:
        _walk(elem)

    tail = " ".join(buf).strip()
    if tail:
        parts.append(tail)
    return parts if parts else [""]


def _parse_table(html: str) -> list[dict]:
    """페이지 HTML에서 회사 정보 테이블(index 1) 파싱.
    - rowspan 셀(상호·대표자·주소)을 추적
    - 셀 내 <br>로 구분된 값(업종·평가액 등)은 별개 행으로 펼침"""
    soup = BeautifulSoup(html, "html.parser")
    tables = soup.find_all("table")
    if len(tables) < 2:
        return []

    tbl = tables[1]
    rows = []
    keys = ["상호", "대표자", "주소", "업종", "평가액"]
    NUM_COLS = len(keys)

    # pending: col_idx → [남은 rowspan 행 수, [값 리스트]]
    pending: dict[int, list] = {}

    for tr in tbl.find_all("tr")[1:]:   # 헤더 행 건너뜀
        cells = tr.find_all("td")
        row_values: list[list[str] | None] = [None] * NUM_COLS
        cell_iter = iter(cells)

        for col_idx in range(NUM_COLS):
            if col_idx in pending:
                remaining, values = pending[col_idx]
                row_values[col_idx] = values
                if remaining - 1 <= 0:
                    del pending[col_idx]
                else:
                    pending[col_idx][0] = remaining - 1
            else:
                td = next(cell_iter, None)
                if td is None:
                    break
                values = _split_by_br(td)
                span = int(td.get("rowspan", 1))
                if span > 1:
                    pending[col_idx] = [span - 1, values]
                row_values[col_idx] = values

        if any(v is None for v in row_values):
            continue

        # <br>로 분리된 값 중 최대 개수만큼 행 생성
        max_n = max(len(v) for v in row_values)   # type: ignore[arg-type]
        for i in range(max_n):
            row_data: dict[str, str] = {}
            for col_idx, key in enumerate(keys):
                vals = row_values[col_idx]          # type: ignore[index]
                row_data[key] = vals[i] if i < len(vals) else vals[-1]
            rows.append(row_data)

    return rows


# ── 체크포인트 ────────────────────────────────────────────────────────────────

def load_checkpoint() -> dict[str, list[dict]]:
    if CHECKPOINT_FILE.exists():
        return json.loads(CHECKPOINT_FILE.read_text(encoding="utf-8"))
    return {}

def save_checkpoint(done: dict[str, list[dict]]) -> None:
    CHECKPOINT_FILE.write_text(json.dumps(done, ensure_ascii=False, indent=2), encoding="utf-8")


# ── 시도별 수집 ───────────────────────────────────────────────────────────────

def collect_sido(sido: str, sido_name: str) -> list[dict]:
    session = _make_session()
    base_url, total_pages = _post_search(session, sido)
    tqdm.write(f"  {sido_name}: {total_pages}페이지")

    # 페이지별 결과를 순서대로 저장하기 위해 index 유지
    results: dict[int, list[dict]] = {}

    def _fetch_one(page: int) -> tuple[int, list[dict]]:
        time.sleep(DELAY * ((page - 1) % PAGE_WORKERS))  # 초기 요청 분산
        rows = _fetch_page(session, base_url, page)
        for row in rows:
            row["시도"] = sido_name
        return page, rows

    with ThreadPoolExecutor(max_workers=PAGE_WORKERS) as ex:
        futures = {ex.submit(_fetch_one, p): p for p in range(1, total_pages + 1)}
        for fut in as_completed(futures):
            try:
                page, rows = fut.result()
                results[page] = rows
            except Exception as e:
                tqdm.write(f"    [오류] {sido_name} p{futures[fut]}: {e}")

    # 페이지 순서대로 합치기
    all_rows: list[dict] = []
    for p in range(1, total_pages + 1):
        all_rows.extend(results.get(p, []))

    return all_rows


# ── xlsx 저장 ─────────────────────────────────────────────────────────────────

def save_xlsx(all_rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "전문건설업체현황"

    col_order = ["시도", "상호", "대표자", "주소", "업종", "평가액"]
    ws.append(col_order)

    for row in all_rows:
        ws.append([row.get(c, "") for c in col_order])

    # 헤더 스타일
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    center   = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = center
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    # 컬럼 너비 자동 조정
    col_widths = {"시도": 8, "상호": 30, "대표자": 12, "주소": 50, "업종": 30, "평가액": 15}
    for i, col in enumerate(col_order, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = col_widths.get(col, 15)

    wb.save(OUTPUT)
    print(f"\n  저장 완료: {OUTPUT}  ({len(all_rows):,}행)")


# ── 메인 ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  대한전문건설협회 전문건설현황 수집기")
    print("=" * 60)
    start = time.time()

    checkpoint = load_checkpoint()
    if checkpoint:
        tqdm.write(f"  체크포인트 발견 — {len(checkpoint)}개 시도 이미 완료, 이어서 수집합니다.")

    lock = threading.Lock()
    todo = [(sido, name) for sido, name in SIDO_MAP.items() if sido not in checkpoint]
    done_rows: list[dict] = [row for rows in checkpoint.values() for row in rows]

    tqdm.write(f"  수집 대상: {len(todo)}개 시도  (시도 {SIDO_WORKERS}개 × 페이지 {PAGE_WORKERS}개 병렬)")

    with tqdm(total=len(todo), desc="시도 전체", unit="시도") as pbar:
        def _worker(sido_name_pair):
            sido, name = sido_name_pair
            rows = collect_sido(sido, name)
            with lock:
                done_rows.extend(rows)
                checkpoint[sido] = rows
                save_checkpoint(checkpoint)
                tqdm.write(f"  [{name}] {len(rows):,}건 수집 (누계: {len(done_rows):,}건)")
                pbar.update(1)
            return sido, len(rows)

        with ThreadPoolExecutor(max_workers=SIDO_WORKERS) as executor:
            futures = {executor.submit(_worker, pair): pair for pair in todo}
            for fut in as_completed(futures):
                try:
                    fut.result()
                except Exception as e:
                    sido, name = futures[fut]
                    tqdm.write(f"  [오류] {name}({sido}): {e}")

    CHECKPOINT_FILE.unlink(missing_ok=True)  # 완료 후 체크포인트 삭제

    print("\n[저장] xlsx 파일 생성...")
    save_xlsx(done_rows)

    elapsed = time.time() - start
    print(f"\n완료! 총 {len(done_rows):,}건  소요시간: {elapsed/60:.1f}분")
    print(f"  → {OUTPUT}")


def test_parse():
    """업종이 여러 개인 업체를 찾아 해당 업체의 모든 행을 출력."""
    print("=== 파싱 테스트 — 업종 다중 행 업체 검색 ===")
    session = _make_session()
    base_url, total_pages = _post_search(session, "11")
    print(f"서울 전체 {total_pages}페이지\n")

    for page in range(1, min(total_pages + 1, 50)):
        rows = _fetch_page(session, base_url, page)
        # 같은 상호가 2회 이상 등장하면 다중 업종 업체
        from collections import Counter
        counts = Counter(r.get("상호", "") for r in rows)
        multi = [name for name, cnt in counts.items() if cnt >= 2]
        if multi:
            # 업종이 서로 다른 경우만 출력 (같은 업종 중복 등록 제외)
            for target in multi:
                target_rows = [r for r in rows if r.get("상호") == target]
                업종_set = {r.get("업종", "") for r in target_rows}
                if len(업종_set) > 1:   # 업종이 실제로 다를 때만
                    print(f"  발견 (p{page}): {target}")
                    for r in target_rows:
                        print(f"    업종={r.get('업종','')} | 평가액={r.get('평가액','')}")
                    break
            else:
                time.sleep(0.2)
                continue
            break
        time.sleep(0.2)
    else:
        print("  50페이지 내 다중 업종 업체 없음")
    print("\n=== 테스트 완료 ===")


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "test":
        test_parse()
    else:
        main()

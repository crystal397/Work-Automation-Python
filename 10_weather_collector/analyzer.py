import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine
from config import DB_PATH, SITES, BASE_DIR
from flags import FLAG_COMPUTATIONS, FLAG_BY_ID

engine = create_engine(f"sqlite:///{DB_PATH}")

_OP_SYM   = {">=": "≥", "<=": "≤", "<": "<", ">": ">"}
_OP_ARROW = {">=": "↑", "<=": "↓", "<": "<", ">": ">"}

_FLAG_UNITS = {
    "precipitation":  "mm", "wind_max":       "m/s", "max_ins_wind": "m/s",
    "snow_depth":     "cm", "temp_max":       "℃",  "temp_min":     "℃",
    "sunshine_hours": "hr", "ground_temp":    "℃",  "evaporation":  "mm",
}

# ── 컬럼 한글 매핑 (기본 관측값용) ───────────────────────────────
COLUMN_LABELS = {
    "date":             "날짜",
    "station_code":     "관측소코드",
    "temp_max":         "최고기온(℃)",
    "temp_min":         "최저기온(℃)",
    "precipitation":    "일강수량(mm)",
    "wind_avg":         "평균풍속(m/s)",
    "wind_max":         "최대풍속(m/s)",
    "max_ins_wind":     "순간최대풍속(m/s)",
    "snow_depth":       "최대적설(cm)",
    "humidity_avg":     "평균습도(%)",
    "sunshine_hours":   "일조시간(hr)",
    "ground_temp":      "지면온도(℃)",
    "evaporation":      "증발량(mm)",
    "pressure":         "평균기압(hPa)",
    "rain_yn":          "강수유무",
    "snow_yn":          "강설유무",
    "fog_yn":           "안개유무",
}

# 플래그 한글명 (요약 시트용)
FLAG_NAMES = {
    "is_rain_day":      "우천 (10mm 이상)",
    "is_wind_day":      "강풍 (14m/s 이상)",
    "is_wind_crane":    "크레인 제한 (순간 10m/s 이상)",
    "is_snow_day":      "적설 (1cm 이상)",
    "is_heat_day":      "폭염 (35℃ 이상)",
    "is_cold_day":      "한파 (-10℃ 이하)",
    "is_no_sunshine":   "일조 부족 (2시간 미만)",
    "is_freeze_day":    "지면 동결 (0℃ 이하)",
    "is_high_evap_day": "증발 과다 (10mm 이상)",
    "rain_yn":          "강수 유무",
    "snow_yn":          "강설 유무",
    "fog_yn":           "안개",
}

# 플래그 기본 단위 표시 (요약용 레이블 기준명)
_FLAG_BASE_NAMES = {
    "is_rain_day": "우천", "is_wind_day": "강풍", "is_wind_crane": "크레인제한",
    "is_snow_day": "적설", "is_heat_day": "폭염", "is_cold_day": "한파",
    "is_no_sunshine": "일조부족", "is_freeze_day": "지면동결",
    "is_high_evap_day": "증발과다",
}


# ── 엑셀 서식 상수 ────────────────────────────────────────────────
_FILL_SECTION = PatternFill("solid", fgColor="D6E4F7")   # 공종 헤더 (파랑)
_FILL_BAD     = PatternFill("solid", fgColor="FFCCCC")   # 작업불가일 (빨강)
_FILL_GOOD    = PatternFill("solid", fgColor="CCFFDD")   # 작업가능일 (초록)
_FILL_HEADER  = PatternFill("solid", fgColor="E8E8E8")   # 헤더 행 (회색)
_FONT_BOLD    = Font(bold=True)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


# ── 헬퍼 함수 ────────────────────────────────────────────────────

def _apply_threshold(series: pd.Series, op: str, threshold: float) -> pd.Series:
    """결측값(NaN/None)은 False 처리, 나머지는 비교 연산 적용"""
    numeric = pd.to_numeric(series, errors="coerce")
    ops = {
        ">=": numeric >= threshold,
        "<=": numeric <= threshold,
        "<":  numeric < threshold,
        ">":  numeric > threshold,
    }
    return ops.get(op, pd.Series(False, index=series.index)).fillna(False)


def _flag_col_label(flag: str, thresholds: dict) -> str:
    """실제 사용 기준값이 반영된 상세 시트 컬럼 헤더명 반환"""
    if flag not in FLAG_COMPUTATIONS:
        return COLUMN_LABELS.get(flag, flag)
    col, op, default_t = FLAG_COMPUTATIONS[flag]
    t = thresholds.get(flag, default_t)
    unit = _FLAG_UNITS.get(col, "")
    t_str = str(int(t)) if t == int(t) else str(t)
    name = _FLAG_BASE_NAMES.get(flag, flag)
    return f"{name}({t_str}{unit}{_OP_ARROW[op]})"


# ── 엑셀 서식 적용 함수 ──────────────────────────────────────────

def _format_summary_ws(ws) -> None:
    """요약 시트 서식 적용"""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    # 헤더 행 (pandas가 쓴 컬럼명 행)
    for cell in ws[1]:
        cell.font = _FONT_BOLD
        cell.fill = _FILL_HEADER

    # 데이터 행
    for row in ws.iter_rows(min_row=2):
        a, b = row[0], row[1]
        v = str(a.value or "")

        if v.startswith("[ ") and v.endswith(" ]"):
            for c in (a, b):
                c.fill = _FILL_SECTION
                c.font = _FONT_BOLD
        elif v == "작업불가일":
            for c in (a, b):
                c.fill = _FILL_BAD
                c.font = _FONT_BOLD
        elif v == "작업가능일":
            for c in (a, b):
                c.fill = _FILL_GOOD
        elif v in ("현장명", "수집기간"):
            a.font = _FONT_BOLD


def _format_detail_ws(ws, n_flag_cols: int) -> None:
    """상세 시트 서식 적용
    n_flag_cols: 플래그 컬럼 수 + 작업불가일 컬럼 (= 우측 끝 컬럼들)
    """
    max_col = ws.max_column
    if not max_col:
        return

    flag_start = max_col - n_flag_cols + 1  # 1-indexed

    # 헤더 행 서식
    for cell in ws[1]:
        cell.font = _FONT_BOLD
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # 컬럼 너비: 헤더 + 내용 기준 자동 설정
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=6,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 24)

    # 플래그 컬럼: 중앙 정렬
    for col_idx in range(flag_start, max_col + 1):
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            cell.alignment = _ALIGN_CENTER

    # 작업불가일 = "O"인 행: 빨간 배경
    imp_col_idx = max_col - 1  # 0-indexed in row tuple
    for row in ws.iter_rows(min_row=2):
        if row[imp_col_idx].value == "O":
            for cell in row:
                cell.fill = _FILL_BAD


# ── 데이터 조회 ──────────────────────────────────────────────────

def get_weather_df(site_id: str, start: str, end: str) -> pd.DataFrame:
    """기간 내 기상 데이터 조회"""
    query = """
        SELECT *
        FROM weather_daily
        WHERE site_id = :site_id
          AND date BETWEEN :start AND :end
        ORDER BY date
    """
    return pd.read_sql(query, engine, params={
        "site_id": site_id,
        "start":   start,
        "end":     end,
    })


# ── 공종별 산정 ──────────────────────────────────────────────────

def analyze_work(df: pd.DataFrame, work: dict) -> dict | None:
    """
    공종별 작업불가일 산정
    work: {"name": ..., "start": ..., "end": ..., "flags": [...],
           "thresholds": {"is_rain_day": 5.0, ...}}  ← 기준값 생략 시 기본값 사용
    """
    mask = (df["date"] >= work["start"]) & (df["date"] <= work["end"])
    wdf  = df[mask].copy()

    if wdf.empty:
        return None

    flags             = work["flags"]
    custom_thresholds = work.get("thresholds", {})

    # 수치 플래그: 원시 관측값 컬럼에서 재계산
    #   커스텀 기준값 우선, 없으면 FLAG_COMPUTATIONS 기본값
    #   결측값(NULL/NaN)은 False 처리 (오판정 방지)
    for flag in flags:
        if flag in FLAG_COMPUTATIONS:
            col, op, default_threshold = FLAG_COMPUTATIONS[flag]
            threshold = custom_thresholds.get(flag, default_threshold)
            wdf[flag] = _apply_threshold(wdf[col], op, threshold)
        # yn 플래그 (rain_yn, snow_yn, fog_yn): iscs 파싱 기반, DB 저장값 그대로 사용

    wdf["is_work_impossible"] = wdf[flags].any(axis=1)

    total_days      = len(wdf)
    impossible_days = int(wdf["is_work_impossible"].sum())
    workable_days   = total_days - impossible_days

    flag_counts = {
        flag: int(wdf[flag].sum())
        for flag in flags
        if flag in wdf.columns
    }

    return {
        "name":            work["name"],
        "start":           work["start"],
        "end":             work["end"],
        "total_days":      total_days,
        "workable_days":   workable_days,
        "impossible_days": impossible_days,
        "flag_counts":     flag_counts,
        "thresholds":      custom_thresholds,
        "df":              wdf,
    }


# ── 시트 데이터 구성 ─────────────────────────────────────────────

def build_summary_sheet(site: dict, results: list) -> pd.DataFrame:
    """요약 시트 데이터 구성"""
    rows = []
    rows.append(("현장명",   site["name"]))
    rows.append(("수집기간", f"{site['start']} ~ {site['end']}"))
    rows.append(("",         ""))

    for work, r in zip(site.get("works", []), results):
        rows.append((f"[ {work['name']} ]", f"{work['start']} ~ {work['end']}"))

        if r is None:
            rows.append(("", "해당 기간 수집된 데이터가 없습니다."))
        else:
            rows.append(("총 일수",    f"{r['total_days']}일"))
            rows.append(("작업가능일", f"{r['workable_days']}일"))
            rows.append(("작업불가일", f"{r['impossible_days']}일"))
            rows.append(("",          "[ 사유별 집계 ]"))

            custom = r.get("thresholds", {})
            for flag, cnt in r["flag_counts"].items():
                label = FLAG_NAMES.get(flag, flag)
                # 커스텀 기준값이 기본값과 다르면 표시
                if flag in FLAG_COMPUTATIONS:
                    col_key, op, default_t = FLAG_COMPUTATIONS[flag]
                    t = custom.get(flag, default_t)
                    if t != default_t:
                        unit = _FLAG_UNITS.get(col_key, "")
                        t_str = str(int(t)) if t == int(t) else str(t)
                        label += f" (기준 {_OP_SYM[op]}{t_str}{unit})"
                rows.append((f"  {label}", f"{cnt}일"))

        rows.append(("", ""))

    rows.append(("※ 비고", "동일 날짜에 여러 사유가 겹쳐도 작업불가일은 1일로 산정"))
    return pd.DataFrame(rows, columns=["항목", "내용"])


def build_detail_sheet(df: pd.DataFrame, work: dict) -> tuple[pd.DataFrame, int]:
    """
    일별 상세 시트 데이터 구성
    반환: (DataFrame, n_flag_cols)  ← n_flag_cols는 서식 적용에 사용
    """
    work_flags        = work["flags"]
    custom_thresholds = work.get("thresholds", {})

    base_cols = [
        "date", "station_code",
        "temp_max", "temp_min", "precipitation",
        "wind_avg", "wind_max", "max_ins_wind",
        "snow_depth", "humidity_avg", "sunshine_hours",
        "ground_temp", "evaporation", "pressure",
    ]

    flag_cols = [f for f in work_flags if f in df.columns]
    imp_col   = ["is_work_impossible"] if "is_work_impossible" in df.columns else []

    cols   = [c for c in base_cols if c in df.columns] + flag_cols + imp_col
    df_out = df[cols].copy()

    # Boolean 컬럼 → "O" / "" 변환
    for col in flag_cols + imp_col:
        df_out[col] = df_out[col].map(lambda x: "O" if x else "")

    # 컬럼명 한글 변환 (플래그는 실제 기준값 반영)
    dynamic_flag_labels = {
        f: _flag_col_label(f, custom_thresholds) for f in flag_cols
    }
    label_map = {**COLUMN_LABELS, **dynamic_flag_labels, "is_work_impossible": "작업불가일"}
    df_out = df_out.rename(columns=label_map)

    n_flag_cols = len(flag_cols) + len(imp_col)
    return df_out, n_flag_cols


# ── 메인 출력 함수 ───────────────────────────────────────────────

def summarize(site: dict):
    site_id = site["id"]
    start   = site["start"]
    end     = site["end"]
    name    = site["name"]
    works   = site.get("works", [])

    df = get_weather_df(site_id, start, end)
    if df.empty:
        print(f"[{name}] 수집된 데이터가 없습니다.")
        return

    results = [analyze_work(df, work) for work in works]

    output_path = BASE_DIR / f"{site_id}_작업불가일.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # 요약 시트
        df_summary = build_summary_sheet(site, results)
        df_summary.to_excel(writer, sheet_name="요약", index=False)
        _format_summary_ws(writer.sheets["요약"])

        # 공종별 상세 시트
        for work, r in zip(works, results):
            sheet_name = work["name"][:31]
            if r is None:
                pd.DataFrame(
                    [["해당 기간 수집된 데이터가 없습니다."]]
                ).to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            else:
                df_detail, n_flag_cols = build_detail_sheet(r["df"], work)
                df_detail.to_excel(writer, sheet_name=sheet_name, index=False)
                _format_detail_ws(writer.sheets[sheet_name], n_flag_cols)

    # 콘솔 출력
    print(f"\n{'='*55}")
    print(f"  {name}")
    print(f"  작업불가일 산정 결과 ({start} ~ {end})")
    print(f"{'='*55}")

    for work, r in zip(works, results):
        print(f"\n  [ {work['name']} ]  {work['start']} ~ {work['end']}")
        if r is None:
            print("    해당 기간 수집된 데이터가 없습니다.")
        else:
            print(f"    총 일수     : {r['total_days']}일")
            print(f"    작업가능일  : {r['workable_days']}일")
            print(f"    작업불가일  : {r['impossible_days']}일")
            print("    사유별 집계 :")
            for flag, cnt in r["flag_counts"].items():
                if flag in FLAG_BY_ID:
                    _, lbl, col, op, default, unit = FLAG_BY_ID[flag]
                    if col is not None:
                        t = r.get("thresholds", {}).get(flag, default)
                        label = f"{lbl}({_OP_SYM[op]}{t}{unit})"
                    else:
                        label = lbl
                else:
                    label = FLAG_NAMES.get(flag, flag)
                print(f"      {label:<30} : {cnt}일")

    print(f"\n  엑셀 저장 완료 → {output_path}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    for site in SITES:
        summarize(site)

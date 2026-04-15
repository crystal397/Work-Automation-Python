"""
포맷별 텍스트 추출기 — 다중 폴백(fallback) 전략
- 각 포맷마다 여러 방법을 순서대로 시도
- WARN/FAIL 이 나오면 다음 방법으로 자동 재시도
- 모든 방법 소진 후에도 실패하면 보고서에 주의 표시
- 중단 없이 최대한 추출
"""

import io
import re
from dataclasses import dataclass, field
from pathlib import Path

import pytesseract
from PIL import Image, ImageFilter, ImageOps
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
TESSERACT_LANG = "kor+eng"

# ── 데이터 구조 ──────────────────────────────────────────────────────────────────

@dataclass
class Chunk:
    """추출된 텍스트 한 단위 (페이지/시트/단락 등)"""
    source: str           # 출처: "파일명.pdf | p.3"
    text: str
    method: str           # 사용한 추출 방법
    quality: str = "OK"   # OK / WARN / FAIL
    note: str = ""        # 품질 이슈 설명
    tried: list = field(default_factory=list)  # 시도한 방법 기록


@dataclass
class ExtractResult:
    """파일 하나의 추출 결과 전체"""
    file: str
    format: str
    chunks: list[Chunk] = field(default_factory=list)
    quality: str = "OK"    # OK / WARN / FAIL
    issues: list[str] = field(default_factory=list)

    @property
    def full_text(self) -> str:
        """출처 태그 포함 전체 텍스트 (분석기 입력용)"""
        parts = []
        for c in self.chunks:
            if c.text.strip():
                parts.append(f"[출처: {c.source}]\n{c.text.strip()}")
        return "\n\n".join(parts)

    @property
    def plain_text(self) -> str:
        return "\n".join(c.text for c in self.chunks if c.text.strip())


# ── 품질 판정 ────────────────────────────────────────────────────────────────────

def _assess(source: str, text: str, method: str,
            min_chars: int = 30, tried: list = None) -> Chunk:
    """텍스트 품질 평가 → Chunk"""
    tried = tried or [method]
    quality, note = "OK", ""

    if not text.strip():
        quality, note = "FAIL", "추출 결과 없음"

    elif len(text.strip()) < min_chars:
        quality, note = "WARN", f"텍스트 {len(text.strip())}자 — 내용 너무 짧음"

    else:
        garbage = len(re.findall(r'[?□▯\ufffd]{2,}', text))
        if garbage > 3:
            quality, note = "WARN", f"깨진 문자 패턴 {garbage}건"
        else:
            korean = len(re.findall(r'[\uac00-\ud7a3]', text))
            total  = max(len(re.sub(r'\s', '', text)), 1)
            if method in ("ocr", "com") and total > 50 and korean / total < 0.05:
                quality, note = "WARN", f"한글 비율 {korean/total:.0%} — 인식 불량 의심"

    return Chunk(source=source, text=text, method=method,
                 quality=quality, note=note, tried=tried)


def _best(chunks: list[Chunk]) -> Chunk:
    """여러 시도 중 가장 좋은 결과 선택 (OK > WARN > FAIL, 같으면 텍스트 길이)"""
    order = {"OK": 0, "WARN": 1, "FAIL": 2}
    return min(chunks, key=lambda c: (order[c.quality], -len(c.text)))


def _roll_up(chunks: list[Chunk]) -> str:
    if any(c.quality == "FAIL" for c in chunks): return "FAIL"
    if any(c.quality == "WARN" for c in chunks): return "WARN"
    return "OK"


# ── 폴백 실행기 ──────────────────────────────────────────────────────────────────

def _try_chain(methods: list[tuple], source: str,
               min_chars: int = 30) -> Chunk:
    """
    methods: [(이름, 함수)] 리스트를 순서대로 시도.
    OK가 나오면 즉시 반환. 모두 실패하면 가장 좋은 결과 반환.
    """
    attempts = []
    tried_names = []

    for name, fn in methods:
        tried_names.append(name)
        try:
            text = fn()
            chunk = _assess(source, text, name, min_chars, tried=list(tried_names))
            attempts.append(chunk)
            if chunk.quality == "OK":
                if len(tried_names) > 1:
                    chunk.note = f"[{tried_names[-2]} 실패 → {name} 성공]"
                return chunk
        except Exception as e:
            attempts.append(Chunk(
                source=source, text="", method=name,
                quality="FAIL", note=f"{name} 예외: {e}",
                tried=list(tried_names)
            ))

    # 모두 실패 → 가장 좋은 결과 반환, 시도 이력 기록
    result = _best(attempts)
    all_tried = " → ".join(tried_names)
    result.note = f"모든 방법 시도 후 최선: [{all_tried}] / {result.note}"
    result.tried = tried_names
    return result


# ── 이미지 전처리 유틸 ────────────────────────────────────────────────────────────

def _preprocess_enhance(img: Image.Image) -> Image.Image:
    """대비 강화"""
    return ImageOps.autocontrast(img.convert("L"))

def _preprocess_threshold(img: Image.Image) -> Image.Image:
    """이진화 (흑백 선명화)"""
    gray = img.convert("L")
    return gray.point(lambda x: 0 if x < 140 else 255, "1").convert("L")

def _preprocess_denoise(img: Image.Image) -> Image.Image:
    """노이즈 제거"""
    return img.convert("RGB").filter(ImageFilter.MedianFilter(3))

def _ocr_image(img: Image.Image, psm: int) -> str:
    return pytesseract.image_to_string(img, lang=TESSERACT_LANG,
                                       config=f"--psm {psm}")

def _ocr_with_confidence(img: Image.Image, psm: int) -> tuple[str, float]:
    data = pytesseract.image_to_data(img, lang=TESSERACT_LANG,
                                     config=f"--psm {psm}",
                                     output_type=pytesseract.Output.DICT)
    words = [w for w, c in zip(data["text"], data["conf"])
             if w.strip() and str(c).lstrip("-").isdigit() and int(c) > 0]
    confs = [int(c) for c in data["conf"]
             if str(c).lstrip("-").isdigit() and int(c) >= 0]
    text  = " ".join(words)
    avg   = sum(confs) / len(confs) if confs else 0
    return text, avg


# ── PDF ───────────────────────────────────────────────────────────────────────────

def _extract_pdf(path: Path) -> ExtractResult:
    import pdfplumber
    import fitz  # PyMuPDF
    import sys as _sys
    _sys.path.insert(0, str(Path(__file__).parent.parent.parent))
    import config as _config

    result = ExtractResult(file=path.name, format="pdf")

    # ── 파일 크기 확인 — 대용량이면 pdfplumber 없이 fitz만 사용 (속도 우선) ──────
    file_mb   = path.stat().st_size / (1024 * 1024)
    fast_mode = file_mb > _config.PDF_FAST_MAX_MB

    # 페이지 수 확인 — fast mode는 fitz로 빠르게, 일반은 pdfplumber
    if fast_mode:
        _doc_count = fitz.open(str(path))
        n_pages    = len(_doc_count)
        _doc_count.close()
        result.issues.append(
            f"{path.name}: {file_mb:.0f}MB — 대용량 PDF, pdfplumber 생략하고 fitz만 사용"
        )
    else:
        with pdfplumber.open(str(path)) as pdf:
            n_pages = len(pdf.pages)

    # 페이지 수 초과 시 OCR 생략 (스캔 PDF라도 텍스트 추출만 시도)
    ocr_allowed = n_pages <= _config.PDF_OCR_MAX_PAGES
    if not ocr_allowed:
        result.issues.append(
            f"{path.name}: {n_pages}페이지 — OCR 생략 (기준: {_config.PDF_OCR_MAX_PAGES}페이지). "
            f"텍스트 추출만 수행. 스캔 페이지는 FAIL 표시."
        )

    # fast_mode: fitz 문서를 한 번 열어두고 재사용
    fitz_doc = fitz.open(str(path)) if fast_mode else None

    for i in range(1, n_pages + 1):
        source = f"{path.name} | p.{i}"

        def _plumber_text(idx=i):
            with pdfplumber.open(str(path)) as p:
                t = p.pages[idx-1].extract_text() or ""
                if not t.strip():
                    raise ValueError("빈 페이지")
                return t

        def _pymupdf_text(idx=i):
            if fitz_doc is not None:
                t = fitz_doc[idx-1].get_text("text")
            else:
                doc = fitz.open(str(path))
                t   = doc[idx-1].get_text("text")
                doc.close()
            if not t.strip():
                raise ValueError("빈 페이지")
            return t

        def _plumber_ocr(idx=i):
            with pdfplumber.open(str(path)) as p:
                img = p.pages[idx-1].to_image(resolution=300).original
                buf = io.BytesIO(); img.save(buf, format="PNG"); buf.seek(0)
                return _ocr_image(Image.open(buf), psm=6)

        def _pymupdf_ocr(idx=i):
            if fitz_doc is not None:
                pix = fitz_doc[idx-1].get_pixmap(dpi=300)
            else:
                doc = fitz.open(str(path))
                pix = doc[idx-1].get_pixmap(dpi=300)
                doc.close()
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            return _ocr_image(img, psm=6)

        def _pymupdf_ocr_enhance(idx=i):
            if fitz_doc is not None:
                pix = fitz_doc[idx-1].get_pixmap(dpi=400)
            else:
                doc = fitz.open(str(path))
                pix = doc[idx-1].get_pixmap(dpi=400)
                doc.close()
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            return _ocr_image(_preprocess_enhance(img), psm=3)

        # ── 텍스트 추출 우선 시도 ─────────────────────────────────────────────────
        # fast_mode: pdfplumber 건너뛰고 fitz만 시도
        if not fast_mode:
            try:
                txt   = _plumber_text()
                chunk = _assess(source, txt, "pdfplumber-text", min_chars=5,
                                tried=["pdfplumber-text"])
                if chunk.text.strip():
                    result.chunks.append(chunk)
                    continue
            except Exception:
                pass

        try:
            txt   = _pymupdf_text()
            tried = ["pymupdf-text"] if fast_mode else ["pdfplumber-text", "pymupdf-text"]
            chunk = _assess(source, txt, "pymupdf-text", min_chars=5, tried=tried)
            if chunk.text.strip():
                result.chunks.append(chunk)
                continue
        except Exception:
            pass

        # 텍스트가 완전히 비어 있는 페이지 — OCR 허용 여부 확인 후 시도
        if not ocr_allowed:
            base_tried = ["pymupdf-text"] if fast_mode else ["pdfplumber-text", "pymupdf-text"]
            chunk = Chunk(source=source, text="", method="skipped-ocr",
                          quality="FAIL", note="스캔 페이지 (OCR 생략 — 페이지 수 초과)",
                          tried=base_tried)
        elif fast_mode:
            # fast_mode: pdfplumber-ocr 생략, fitz 기반 OCR만
            chunk = _try_chain([
                ("pymupdf-ocr",         _pymupdf_ocr),
                ("pymupdf-ocr-enhance", _pymupdf_ocr_enhance),
            ], source=source)
        else:
            chunk = _try_chain([
                ("pdfplumber-ocr",     _plumber_ocr),
                ("pymupdf-ocr",        _pymupdf_ocr),
                ("pymupdf-ocr-enhance",_pymupdf_ocr_enhance),
            ], source=source)

        result.chunks.append(chunk)

    if fitz_doc is not None:
        fitz_doc.close()

    return result


# ── Excel ─────────────────────────────────────────────────────────────────────────

def _extract_excel(path: Path) -> ExtractResult:
    import pandas as pd
    import openpyxl

    result = ExtractResult(file=path.name, format="excel")

    # 시트 목록 확보
    try:
        sheet_names = pd.ExcelFile(str(path)).sheet_names
    except Exception:
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            result.quality = "FAIL"
            result.issues.append(f"시트 목록 읽기 실패: {e}")
            return result

    for sheet in sheet_names:
        def _pandas(s=sheet):
            df = pd.read_excel(str(path), sheet_name=s, header=None).dropna(how="all")
            if df.empty: raise ValueError("빈 시트")
            rows = []
            for idx, (_, row) in enumerate(df.iterrows(), 1):
                cells = [str(v) for v in row if str(v) not in ("nan","NaT","")]
                if cells:
                    rows.append((idx, " | ".join(cells)))
            return rows

        def _openpyxl(s=sheet):
            wb  = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws  = wb[s]
            rows = []
            for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                cells = [str(v) for v in row if v is not None and str(v).strip()]
                if cells:
                    rows.append((idx, " | ".join(cells)))
            wb.close()
            if not rows: raise ValueError("빈 시트")
            return rows

        # 두 방법으로 행 리스트 확보 후 청크 생성
        rows = None
        tried = []
        for name, fn in [("pandas", _pandas), ("openpyxl", _openpyxl)]:
            tried.append(name)
            try:
                rows = fn(); break
            except Exception as e:
                pass

        if not rows:
            result.chunks.append(Chunk(
                source=f"{path.name} | {sheet}",
                text="", method="pandas+openpyxl",
                quality="FAIL",
                note=f"시트 '{sheet}' 읽기 실패 (pandas+openpyxl 모두 실패)",
                tried=tried
            ))
            continue

        for row_idx, row_text in rows:
            source = f"{path.name} | {sheet} | r{row_idx}"
            result.chunks.append(
                _assess(source, row_text, tried[-1], min_chars=2, tried=tried)
            )

    return result


# ── HTML ──────────────────────────────────────────────────────────────────────────

def _extract_html(path: Path) -> ExtractResult:
    from bs4 import BeautifulSoup

    result = ExtractResult(file=path.name, format="html")

    raw = path.read_bytes()

    def _lxml():
        soup = BeautifulSoup(raw, "lxml")
        for t in soup(["script","style","meta","link"]): t.decompose()
        return soup.get_text(separator="\n", strip=True)

    def _html_parser():
        soup = BeautifulSoup(raw, "html.parser")
        for t in soup(["script","style","meta","link"]): t.decompose()
        return soup.get_text(separator="\n", strip=True)

    def _regex():
        text = raw.decode("utf-8", errors="replace")
        text = re.sub(r'<script[^>]*>.*?</script>', '', text, flags=re.DOTALL)
        text = re.sub(r'<style[^>]*>.*?</style>',  '', text, flags=re.DOTALL)
        text = re.sub(r'<[^>]+>', ' ', text)
        return re.sub(r'\s+', '\n', text).strip()

    chunk = _try_chain([
        ("bs4-lxml",        _lxml),
        ("bs4-html.parser", _html_parser),
        ("regex",           _regex),
    ], source=f"{path.name} | doc")
    result.chunks.append(chunk)
    return result


# ── XML ───────────────────────────────────────────────────────────────────────────

def _extract_xml(path: Path) -> ExtractResult:
    from lxml import etree
    import xml.etree.ElementTree as ET
    from bs4 import BeautifulSoup

    result = ExtractResult(file=path.name, format="xml")
    raw    = path.read_bytes()

    def _lxml():
        tree  = etree.fromstring(raw)
        texts = [t.strip() for t in tree.itertext() if t.strip()]
        if not texts: raise ValueError("텍스트 노드 없음")
        return "\n".join(texts)

    def _stdlib():
        root  = ET.fromstring(raw)
        texts = [t.strip() for t in root.itertext() if t and t.strip()]
        if not texts: raise ValueError("텍스트 노드 없음")
        return "\n".join(texts)

    def _bs4():
        soup  = BeautifulSoup(raw, "xml")
        return soup.get_text(separator="\n", strip=True)

    chunk = _try_chain([
        ("lxml",   _lxml),
        ("stdlib", _stdlib),
        ("bs4-xml",_bs4),
    ], source=f"{path.name} | doc")
    result.chunks.append(chunk)
    return result


# ── HWP / HWPX ───────────────────────────────────────────────────────────────────

def _extract_hwp(path: Path) -> ExtractResult:
    import win32com.client
    import fitz

    result   = ExtractResult(file=path.name, format="hwp")
    abs_path = str(path.resolve())

    # ── HWP 프로세스 강제 종료 ────────────────────────────────────────────────────
    def _kill_hwp():
        """COM 실패 후 남은 HWP 프로세스를 강제 종료"""
        import subprocess
        try:
            subprocess.run(
                ["taskkill", "/F", "/IM", "Hwp.exe"],
                capture_output=True, timeout=5
            )
        except Exception:
            pass

    # ── COM 호출 타임아웃 래퍼 ────────────────────────────────────────────────────
    def _run_with_timeout(fn, timeout_sec: int = 60):
        """스레드로 fn 실행, timeout_sec 초 초과 시 TimeoutError.
        COM은 스레드별로 CoInitialize가 필요하므로 래퍼 안에서 초기화.

        주의: ThreadPoolExecutor는 스레드를 강제 종료할 수 없으므로,
        타임아웃 시 shutdown(wait=False)로 대기 없이 포기.
        멈춘 COM 스레드는 백그라운드에서 OS가 정리함.
        """
        import concurrent.futures

        def _com_wrapper():
            try:
                import pythoncom
                pythoncom.CoInitialize()
            except Exception:
                pass
            try:
                return fn()
            finally:
                try:
                    import pythoncom
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

        # with 구문 대신 직접 관리 — shutdown(wait=True)가 무한 블로킹하는 것을 방지
        ex = concurrent.futures.ThreadPoolExecutor(max_workers=1)
        future = ex.submit(_com_wrapper)
        try:
            result = future.result(timeout=timeout_sec)
            ex.shutdown(wait=False)
            return result
        except concurrent.futures.TimeoutError:
            future.cancel()
            ex.shutdown(wait=False)   # 멈춘 스레드를 기다리지 않고 포기
            _kill_hwp()               # HWP 프로세스 강제 종료
            raise TimeoutError(f"HWP 처리 {timeout_sec}초 초과 — 파일 건너뜀")
        except KeyboardInterrupt:
            future.cancel()
            ex.shutdown(wait=False)
            _kill_hwp()
            raise

    # ── 방법 1: HWP COM GetText ──────────────────────────────────────────────────
    def _com_gettext():
        hwp = win32com.client.Dispatch("HWPFrame.HwpObject")
        hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")
        hwp.XHwpWindows.Active_XHwpWindow.Visible = False
        hwp.Open(abs_path, "HWP", "forceopen:true")
        texts, para = [], 1
        try:
            hwp.InitScan()
            while True:
                state, text = hwp.GetText()
                if state == 1: break
                if text and text.strip():
                    texts.append((para, text)); para += 1
            hwp.ReleaseScan()
        finally:
            try: hwp.Quit()
            except: pass
        if not texts: raise ValueError("GetText 결과 없음")
        return texts  # [(para_idx, text)]

    # ── 방법 2: HWP COM → PDF 변환 후 PDF 추출 ──────────────────────────────────
    def _com_to_pdf():
        import tempfile, os
        hwp = win32com.client.Dispatch("HWPFrame.HwpObject")
        hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")
        hwp.XHwpWindows.Active_XHwpWindow.Visible = False
        hwp.Open(abs_path, "HWP", "forceopen:true")
        tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        tmp.close()
        try:
            hwp.SaveAs(tmp.name, "PDF")
            hwp.Quit()
            doc   = fitz.open(tmp.name)
            texts = [(i+1, doc[i].get_text("text")) for i in range(len(doc))
                     if doc[i].get_text("text").strip()]
            doc.close()
            if not texts: raise ValueError("PDF 변환 결과 없음")
            return texts
        finally:
            try: os.unlink(tmp.name)
            except: pass

    # ── 방법 3: hwp5txt (olefile 기반) ──────────────────────────────────────────
    def _hwp5txt():
        try:
            import subprocess, tempfile, os
            tmp = tempfile.NamedTemporaryFile(suffix=".txt", delete=False)
            tmp.close()
            r = subprocess.run(
                ["hwp5txt", "--output", tmp.name, abs_path],
                capture_output=True, timeout=30
            )
            if r.returncode != 0:
                raise RuntimeError(r.stderr.decode("utf-8", errors="replace"))
            text = Path(tmp.name).read_text(encoding="utf-8", errors="replace")
            os.unlink(tmp.name)
            if not text.strip(): raise ValueError("hwp5txt 결과 없음")
            return [(1, text)]
        except FileNotFoundError:
            raise RuntimeError("hwp5txt 미설치")

    # 각 방법 시도 — COM 실패 시 HWP 프로세스 정리 후 다음 방법으로
    para_rows = None
    tried_methods = []

    for name, fn in [("COM-GetText", _com_gettext),
                     ("COM-PDF변환", _com_to_pdf),
                     ("hwp5txt",    _hwp5txt)]:
        tried_methods.append(name)
        try:
            para_rows = _run_with_timeout(fn, timeout_sec=60); break
        except Exception as e:
            result.issues.append(f"HWP {name} 실패: {e}")
            # COM 관련 방법 실패 시 잔여 HWP 프로세스 정리
            if name.startswith("COM"):
                _kill_hwp()

    if not para_rows:
        result.chunks.append(Chunk(
            source=f"{path.name} | para.1",
            text="", method="hwp-all-failed",
            quality="FAIL",
            note="COM-GetText / COM-PDF변환 / hwp5txt 모두 실패 — 수동 확인 필요",
            tried=tried_methods
        ))
        return result

    for para_idx, text in para_rows:
        source = f"{path.name} | para.{para_idx}"
        result.chunks.append(
            _assess(source, text, tried_methods[-1], tried=tried_methods)
        )
    return result


# ── TIF / 이미지 OCR ─────────────────────────────────────────────────────────────

def _extract_image_ocr(path: Path) -> ExtractResult:
    result = ExtractResult(file=path.name, format="image")
    img    = Image.open(str(path))

    frame = 0
    while True:
        try:
            img.seek(frame)
        except EOFError:
            break

        page_img = img.copy().convert("RGB")
        source   = f"{path.name} | frame.{frame+1}"

        def _psm6(i=page_img):
            t, conf = _ocr_with_confidence(i, psm=6)
            if conf < 60: raise ValueError(f"신뢰도 {conf:.0f}%")
            return t

        def _psm3(i=page_img):
            t, conf = _ocr_with_confidence(i, psm=3)
            if conf < 60: raise ValueError(f"신뢰도 {conf:.0f}%")
            return t

        def _psm11(i=page_img):
            t, conf = _ocr_with_confidence(i, psm=11)
            if conf < 60: raise ValueError(f"신뢰도 {conf:.0f}%")
            return t

        def _enhance_psm6(i=page_img):
            t, conf = _ocr_with_confidence(_preprocess_enhance(i), psm=6)
            if conf < 50: raise ValueError(f"신뢰도 {conf:.0f}%")
            return t

        def _threshold_psm6(i=page_img):
            t, conf = _ocr_with_confidence(_preprocess_threshold(i), psm=6)
            if conf < 50: raise ValueError(f"신뢰도 {conf:.0f}%")
            return t

        def _denoise_psm6(i=page_img):
            t, conf = _ocr_with_confidence(_preprocess_denoise(i), psm=6)
            if conf < 45: raise ValueError(f"신뢰도 {conf:.0f}%")
            return t

        chunk = _try_chain([
            ("ocr-psm6",           _psm6),
            ("ocr-psm3",           _psm3),
            ("ocr-psm11",          _psm11),
            ("ocr-enhance-psm6",   _enhance_psm6),
            ("ocr-threshold-psm6", _threshold_psm6),
            ("ocr-denoise-psm6",   _denoise_psm6),
        ], source=f"{source} [OCR]", min_chars=10)

        result.chunks.append(chunk)
        frame += 1

    return result


# ── 메인 진입점 ───────────────────────────────────────────────────────────────────

SUPPORTED = {".pdf", ".xlsx", ".xls", ".html", ".htm", ".xml",
             ".hwp", ".hwpx", ".tif", ".tiff", ".png", ".jpg", ".jpeg"}

HANDLERS = {
    ".pdf":  _extract_pdf,
    ".xlsx": _extract_excel, ".xls":  _extract_excel,
    ".html": _extract_html,  ".htm":  _extract_html,
    ".xml":  _extract_xml,
    ".hwp":  _extract_hwp,   ".hwpx": _extract_hwp,
    ".tif":  _extract_image_ocr, ".tiff": _extract_image_ocr,
    ".png":  _extract_image_ocr, ".jpg":  _extract_image_ocr, ".jpeg": _extract_image_ocr,
}


def extract(file_path: str) -> ExtractResult:
    path = Path(file_path)
    ext  = path.suffix.lower()

    if ext not in HANDLERS:
        return ExtractResult(file=path.name, format=ext,
                             quality="FAIL",
                             issues=[f"지원하지 않는 포맷: {ext}"])
    try:
        result = HANDLERS[ext](path)
        if not result.chunks:
            result.quality = "FAIL"
            result.issues.append("추출된 내용 없음")
        else:
            result.quality = _roll_up(result.chunks)
            for c in result.chunks:
                if c.quality != "OK":
                    result.issues.append(f"{c.source}: {c.note}")
        return result
    except Exception as e:
        return ExtractResult(file=path.name, format=ext,
                             quality="FAIL",
                             issues=[f"처리 중 예외: {e}"])


def _cache_path(folder: Path) -> Path:
    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent.parent))
    import config
    return config.OUTPUT_DIR / "extract_cache.json"


def _load_cache(folder: Path) -> dict:
    """캐시 로드: {rel_path: mtime}"""
    import json
    p = _cache_path(folder)
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_cache(folder: Path, cache: dict):
    import json
    p = _cache_path(folder)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def extract_folder(folder_path: str) -> list[ExtractResult]:
    """
    폴더 내 모든 지원 파일을 재귀적으로 탐색하여 추출.
    하위 폴더 포함, 출처에 상대 경로 표기.
    이미 처리한 파일(경로+수정시각 동일)은 건너뜀.
    """
    folder = Path(folder_path)
    files  = sorted(
        f for f in folder.rglob("*")
        if f.is_file() and f.suffix.lower() in SUPPORTED
    )

    cache = _load_cache(folder)

    # 신규/변경 파일만 추출 대상
    to_process = []
    skipped = []
    for f in files:
        key   = str(f.relative_to(folder))
        mtime = str(f.stat().st_mtime)
        if cache.get(key) == mtime:
            skipped.append(f)
        else:
            to_process.append(f)

    print(f"총 {len(files)}개 파일 발견  |  신규/변경 {len(to_process)}개  |  캐시 스킵 {len(skipped)}개\n")

    if not to_process:
        print("모든 파일이 이미 처리되어 있습니다. output/extracted_for_analysis.md를 확인하세요.")
        # 캐시된 파일도 결과로 반환해야 하므로 빈 리스트 대신 스킵 처리
        return []

    # 폴더 구조 미리 보기
    dirs = sorted({f.parent.relative_to(folder) for f in to_process})
    for d in dirs:
        count = sum(1 for f in to_process if f.parent.relative_to(folder) == d)
        label = "(루트)" if str(d) == "." else str(d)
        print(f"  📁 {label}  —  {count}개 파일")
    print()

    try:
        from tqdm import tqdm
        _tqdm_available = True
    except ImportError:
        _tqdm_available = False

    results = []
    bar = tqdm(to_process, unit="파일", dynamic_ncols=True) if _tqdm_available else to_process

    try:
        for f in bar:
            rel   = f.relative_to(folder)
            fname = rel.name
            key   = str(rel)
            mtime = str(f.stat().st_mtime)

            if _tqdm_available:
                bar.set_description(f"⏳ {fname[:50]}")

            r = _extract_with_rel(f, rel)
            icon = {"OK": "✅", "WARN": "⚠️", "FAIL": "❌"}[r.quality]

            if _tqdm_available:
                bar.set_description(f"{icon} {fname[:50]}")
                tqdm.write(f"  {icon} {rel}  ({len(r.plain_text):,}자, {len(r.chunks)}청크)")
            else:
                print(f"  {icon} {rel}  ({len(r.plain_text):,}자, {len(r.chunks)}청크)")

            warn_issues  = [i for i in r.issues if "너무 짧음" not in i and "추출 결과 없음" not in i]
            short_issues = [i for i in r.issues if "너무 짧음" in i or "추출 결과 없음" in i]
            _print = tqdm.write if _tqdm_available else print
            for issue in warn_issues[:5]:
                _print(f"       → {issue}")
            if len(warn_issues) > 5:
                _print(f"       → ... 외 {len(warn_issues)-5}건 (extraction_report.md 참고)")
            if short_issues:
                _print(f"       → 짧은 셀/페이지 {len(short_issues)}건 (extraction_report.md 참고)")

            results.append(r)
            # 처리 완료 → 캐시 업데이트 (파일 하나씩 저장하여 중단 시 재개 가능)
            cache[key] = mtime
            _save_cache(folder, cache)

    except KeyboardInterrupt:
        if _tqdm_available:
            bar.close()
        done = len(results)
        remaining = len(to_process) - done
        print(f"\n\n[중단됨] 지금까지 {done}개 처리 완료, {remaining}개 미처리")
        print("  재실행하면 완료된 파일은 캐시에서 건너뛰고 이어서 진행합니다.")
        print("  (캐시 위치: extract_cache.json)")

    ok   = sum(1 for r in results if r.quality == "OK")
    warn = sum(1 for r in results if r.quality == "WARN")
    fail = sum(1 for r in results if r.quality == "FAIL")
    print(f"\n완료: ✅ {ok}  ⚠️ {warn}  ❌ {fail}  (캐시 스킵 {len(skipped)}개)")
    return results


def _extract_with_rel(path: Path, rel: Path) -> ExtractResult:
    """
    추출 후 모든 Chunk의 source를 상대 경로 기준으로 재설정.
    예) 계약서/본계약.pdf | p.1
    """
    result = extract(str(path))
    rel_str = str(rel).replace("\\", "/")

    # file 표시를 상대 경로로 변경
    result.file = rel_str

    # 각 청크의 source에서 파일명만 상대 경로로 교체
    for chunk in result.chunks:
        if path.name in chunk.source:
            chunk.source = chunk.source.replace(path.name, rel_str, 1)

    return result


if __name__ == "__main__":
    import sys
    results = extract_folder(sys.argv[1] if len(sys.argv) > 1 else "input")
    for r in results:
        print(f"\n{'='*60}\n[{r.file}] quality={r.quality}")
        print(r.full_text[:500])

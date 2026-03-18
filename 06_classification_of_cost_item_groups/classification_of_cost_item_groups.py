#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
물가변동 비목군 자동분류 프로그램 v2.0
=======================================
산출근거(산근) 엑셀 파일의 D열(품명)을 읽어 A열에 비목군 기호를 자동 표기

사용법:
    python 비목군_자동분류.py [엑셀파일경로]
    예) python 비목군_자동분류.py 남부내륙1공구_산근.xlsx

엑셀 구조 (고정):
    A열: 비목군 (출력)     B열: 호표     C열: 코드     D열: 품명
    E열: 재료비 단가       F열: 노무비 단가    G열: 경비 단가    H열: 합계 단가
    6행: 헤더, 8행부터 데이터

근거 규정:
    - 국가계약법 시행규칙 제74조
    - 정부입찰·계약 집행기준 제68조~제70조의5
    - 조달청 물가변동 질의응답집 (2025.10)

비목군: A(노무비) B(기계경비) C(광산품) D(공산품) E(전력수도가스폐기물)
        F(농림수산품) G1~G5(표준시장단가) H~N(각종 보험료) Z(기타)
"""

import sys
import os
import re
import openpyxl


# ================================================================
#  한국은행 생산자물가지수 기본분류 기반 키워드 사전
#  (질의응답집 71~74페이지 + 보강)
# ================================================================

# 광산품 (C) - 석탄/원유, 금속광업, 비금속광물광업
C_KEYWORDS = [
    "토사석", "석회석", "점토", "석재", "쇄석", "모래", "자갈",
    "소금", "고령토", "규조토", "화강암", "골재", "사석", "원석",
    "석분", "보조기층", "기층재", "필터재", "동상방지층재",
    "무연탄", "원유", "천연가스", "연탄",
    "주석", "니켈", "광석", "항만사석", "화강석 원석",
]

# 공산품 (D)
D_KEYWORDS = [
    # 섬유제품
    "섬유사", "로프", "망", "유리섬유", "포대", "도포", "천막",
    "생사", "어망", "장갑", "모피", "부직포",
    # 목재
    "제재목", "합판", "콜크", "침목", "각재", "판재", "지주목", "목재",
    # 코크스/석유
    "코크스", "타르", "윤활유", "그리스", "휘발유", "경유", "중유",
    "벙커", "기계유",
    # 펄프/종이/출판
    "펄프", "종이", "판지", "신문", "서적",
    # 화학
    "비료", "도료", "합성고무", "플라스틱", "프라스틱", "프라스틱제품", "아세틸렌",
    "압축산소", "혼합가스", "안료", "화약", "접착제", "의약품",
    "벤젤", "목탄", "페인트", "부엽토", "시너", "에폭시",
    "프라이머", "방청제", "실리콘", "우레탄", "폴리머",
    # 고무/플라스틱
    "타이어", "튜브", "고무", "비닐", "스티로폼",
    "PE필름", "P.E 필름", "PVC",
    # 비금속광물제품
    "유리", "석면", "암면", "타일", "연마지", "아스콘",
    "기와", "레미콘", "콘크리트제품", "인공골재", "시멘트",
    "모르타르", "벽돌", "블록", "맨홀", "흄관", "석재품", "도기시멘트",
    # 금속1차
    "철광괴", "철강재", "주물", "합금철강", "슬래그", "귀금속재",
    "이형철근", "철근", "형강", "강판", "철판", "강관", "파이프",
    "앵글", "플레이트", "와이어", "PC강연선", "강선",
    # 조립금속
    "탱크", "용기", "수공구", "보일러", "방열기", "철문",
    "철구조물", "못", "용접봉", "도금", "샤시문", "볼트", "너트",
    "와셔", "앵커", "스터드", "클램프", "브라켓", "행거",
    "난간", "펜스", "가드레일", "그레이팅",
    # 전기기계
    "전동기", "변압기", "전구", "조명기구", "배전반",
    "전기제어", "전기기계", "케이블", "전선", "배선",
    # 영상/음향
    "방송장치", "통신장비", "영상장비", "음향장비", "전자부품",
    # 기타 공산품
    "금속가구", "목재가구", "플라스틱가구", "간판", "광고물",
    "장식장", "침대", "매트리스", "운동용구",
    "폐철강재", "비철금속재생재", "재생재", "고철",
    "교명판", "설명판", "방음벽", "방음판", "울타리",
    "드레인보드", "지수판", "충진재", "실런트", "코킹",
    "백업재", "가스켓", "패킹", "스페이서",
    "H-Beam", "H-PILE", "SHEET-PILE", "시트파일",
    "복공판", "데크플레이트", "거더", "빔",
    "산소", "L.P.G", "LPG",
    "도장", "방수", "방식", "도막",
    "알루미늄", "동판", "스텐레스", "황동",
    "커플러", "슬리브", "다웰바", "캐노피", "루버", "차양",
    "폭약", "뇌관", "전색", "비트", "로드",
]

# 전력·수도·도시가스 및 폐기물 (E)
E_KEYWORDS = [
    "전력", "전기료", "전력요금", "전기요금", "전기사용료", "충전료",
    "수도", "급수", "생활용수", "공업용수", "수도요금",
    "도시가스", "가스요금", "난방용", "증기", "온수",
    "폐기물", "폐수", "분뇨처리", "하수", "폐기물처리",
    "폐기물수집", "폐기물운반", "폐기물 수집", "폐기물 운반",
]

# 농림수산품 (F)
F_KEYWORDS = [
    "묘목", "관상수", "정원수", "종자", "화훼", "파종",
    "화목", "원목", "비계목", "잔디",
    "수목", "소나무", "잡목", "지피류", "관목", "교목",
    "식재", "녹화", "객토",
]

# 기계경비 키워드
B_KEYWORDS = [
    "장비손료", "기계손료", "크레인", "굴삭기", "불도저",
    "로울러", "로더", "항타기", "펌프카", "콤프레셔",
    "다짐기", "발전기", "그라우트펌프",
]

E_CODES = ["K0810301B", "GS25W050", "GS26W060"]


# ================================================================
#  유틸리티 함수
# ================================================================
def _contains(text, keywords):
    text_upper = str(text or "").upper()
    for k in keywords:
        if k.upper() in text_upper:
            return True
    return False


def _is_std_code(code):
    return str(code or "").strip().startswith("_")


def _is_std_name(name):
    return "표준시장" in str(name or "")


def _strip(name):
    return str(name or "").strip()


def _is_expense_label(name):
    cleaned = re.sub(r'\s+', '', _strip(name)[:6])
    return cleaned.startswith('경비') or cleaned.startswith('-경비')


def _is_labor_label(name):
    cleaned = re.sub(r'\s+', '', _strip(name)[:6])
    return cleaned.startswith('노무비') or cleaned.startswith('-노무비')


def _is_material_label(name):
    cleaned = re.sub(r'\s+', '', _strip(name)[:8])
    return cleaned.startswith('재료비') or cleaned.startswith('-재료비')


def _scan_sub(ws, row, max_row, depth=80):
    has_std = has_gyun = False
    for sr in range(row + 1, min(row + depth, max_row + 1)):
        sb = ws.cell(row=sr, column=2).value
        if sb and str(sb).startswith("#."):
            break
        sc = str(ws.cell(row=sr, column=3).value or "").strip()
        sd = str(ws.cell(row=sr, column=4).value or "")
        if _is_std_code(sc) or _is_std_name(sd):
            has_std = True
        if "견적" in sd:
            has_gyun = True
    return has_std, has_gyun


def _std_nearby(ws, row):
    for sr in range(max(row - 5, 1), row):
        if "표준시장" in str(ws.cell(row=sr, column=4).value or ""):
            return True
    return False


# ================================================================
#  분류 함수
# ================================================================
def classify_material(name, code=""):
    if "전기뇌관" in str(name or ""):
        return "D"
    if _contains(name, C_KEYWORDS):
        return "C"
    if _contains(name, F_KEYWORDS):
        return "F"
    if _contains(name, E_KEYWORDS):
        return "E"
    return "D"


def classify_expense(code, name, parent):
    code = str(code or "").strip()
    name = _strip(name)

    if code in E_CODES:
        return "E"
    if code.startswith("L0010"):
        return "A"
    if code.startswith("GS25") or code.startswith("GS26"):
        return "E"

    if _contains(name, E_KEYWORDS) and "전기뇌관" not in name:
        return "E"
    # [p78⑤] 급수비(천공시 노무비의 00%) → E
    if "급수비" in name:
        return "E"

    # 사람 → A (노무비)
    if any(k in name for k in ["시험인", "품질관리원", "관리원", "기술사",
                                 "기능사", "기술자", "기사", "조사원"]):
        return "A"
    if _contains(name, B_KEYWORDS):
        return "B"

    if ("재료비" in name or "잡재료" in name or "잡품" in name) and ("%" in name or "/" in name):
        return "D"
    if "기구손료" in name or "공구손료" in name:
        return "D"
    if "인력품" in name:
        return "D"

    if _is_labor_label(name):
        return "A"
    if _is_expense_label(name):
        return "B"
    if _is_material_label(name):
        return "D"

    if "노무비" in name and ("%" in name or "/" in name):
        return "A"
    if "직접노무비" in name:
        return "B"
    if "간접노무비" in name:
        return "A"

    # [p79] Z비목 해당 항목
    if any(k in name for k in ["제경비", "기타경비", "산재보험", "고용보험",
                                 "건강보험", "연금보험", "노인장기요양",
                                 "퇴직공제", "환경보전비", "산업안전보건"]):
        return "Z"
    # [p79] 보상비, 임대료, 사용료, 한전불입금 → Z
    if any(k in name for k in ["보상비", "임대료", "사용료", "한전불입금",
                                 "구역화물"]):
        return "Z"

    if "견적" in code or "견적" in name:
        return parent

    return parent


def classify_G(name):
    name = str(name or "")
    # G5를 먼저 검사 (품명에 "전기설비"+"통신" 모두 포함될 수 있음)
    if any(k in name for k in ["통신설비", "신호설비", "정보통신", "CCTV", "감시설비"]):
        return "G5"
    if any(k in name for k in ["전력설비", "전기설비", "배전", "수변전",
                                 "접지", "전차선", "조명설비"]):
        return "G4"
    if any(k in name for k in ["기계설비", "배관", "펌프", "환기",
                                 "소방", "급배수", "보일러"]):
        return "G3"
    if any(k in name for k in ["건축", "미장", "타일", "도배", "유리",
                                 "창호", "유로폼", "거푸집", "문양거푸집"]):
        return "G2"
    return "G1"


# ================================================================
#  상위 품목(호표) 분류
# ================================================================
def classify_header(row, code, name, mat, lab, exp, ws, mx):
    code = str(code or "").strip()
    name = str(name or "").strip()
    mat = float(mat) if mat else 0
    lab = float(lab) if lab else 0
    exp = float(exp) if exp else 0
    nz = sum(1 for v in [mat, lab, exp] if v != 0)

    if _is_std_code(code) or _is_std_name(name):
        return classify_G(name)
    if "품질시험비" in name or code.startswith("K09509"):
        return "Z"
    if "비파괴" in name:
        return "Z"

    has_std, has_gyun = _scan_sub(ws, row, mx)

    # [p76②] 재료비 포함 설치비/제작비 견적 → G
    #   단, 설치비가 재료비의 00%이거나 자재공급자 설치 → 재료비 비목
    # [p76③] 1식 세부내역 없으면 → G
    # [p76④] 분류 곤란한 기타공사 견적 → 주공사의 G비목

    # 견적서 참조이고 2개↑ 비목 합산 → G
    if has_gyun and nz >= 2:
        return classify_G(name)
    if has_gyun and nz == 1:
        if mat != 0: return classify_material(name, code)
        if lab != 0: return "A"
        if exp != 0: return classify_expense(code, name, "Z")
    # 견적서인데 비목수 판단 불가(모두 0이지만 견적 참조) → G
    if has_gyun and nz == 0:
        return classify_G(name)

    if has_std and nz >= 2:
        return classify_G(name)

    if mat != 0 and lab == 0 and exp == 0:
        return classify_material(name, code)
    if lab != 0 and mat == 0 and exp == 0:
        return "A"
    if exp != 0 and mat == 0 and lab == 0:
        return classify_expense(code, name, "Z")

    if nz >= 2:
        return _complex(name, code, row, ws, mx)
    if nz == 0:
        return _zero(name)
    return "Z"


def _complex(name, code, row, ws, mx):
    """재+노+경 복합 품목 [제68조: 2개↑ 비목 합산 → G가 원칙]
    
    핵심 원칙: 재·노·경 중 2개 이상 비목이 합산된 단가는 G비목.
    예외: 재료비가 절대 주도(70%↑)이고 노무비가 단순 설치비(재료비의 %)인 경우 → 재료비 비목
    """
    name = str(name or "")

    # === 1. G 세부 분류 (전기/통신/기계/건축) ===
    if any(k in name for k in ["전력설비", "전기설비", "배전", "전차선"]): return "G4"
    if any(k in name for k in ["통신설비", "신호설비", "정보통신"]): return "G5"
    if any(k in name for k in ["기계설비", "소방설비"]): return "G3"

    # === 2. 하위에 표준시장단가 포함 → G ===
    has_std, _ = _scan_sub(ws, row, mx)
    if has_std:
        return classify_G(name)

    # === 3. Z 해당 항목 (복합이어도 Z) ===
    # [p79] 보상비, 임대료, 사용료, 한전불입금 → Z
    if any(k in name for k in ["보상비", "임대료", "한전불입금", "구역화물"]):
        return "Z"
    # [p78④] 계측비 → G1 (분류 가능 비목이지만 복합이면 G)
    if "계측" in name:
        return "G1"

    # === 4. 제68조 핵심 원칙 적용 ===
    # 재+노+경 2개↑ 비목 복합 → G가 원칙
    # 단, 순수 재료비 주도 + 설치비가 재료비의 %로 계상된 경우 → 재료비 비목
    # (p76: "설치비가 단순히 재료비의 00%로 계상이 되거나, 
    #  인도조건이 자재공급자가 설치하는 경우라면 재료비 비목으로 분류")

    # 농림수산품은 식재 공종으로 G1보다 F가 적합할 수 있음
    if _contains(name, F_KEYWORDS):
        return "F"

    # 기본: 복합 공종은 G (토목이 기본)
    return classify_G(name)


def _zero(name):
    name = str(name or "")
    if "운반" in name and "현장도착도" in name: return "D"
    if any(k in name for k in ["철거", "절단"]): return "D"
    if any(k in name for k in ["포함", "소운반"]): return "Z"
    if any(k in name for k in ["콘크리트", "라이닝"]): return "Z"
    if "와이어메쉬" in name: return "D"
    if "운반" in name: return "G1"
    return "Z"


# ================================================================
#  하위 행 분류
# ================================================================
def classify_sub(code, name, e, f, g, parent, p_nz, ws, row):
    code = str(code or "").strip()
    name = _strip(name)
    e = float(e) if e else 0
    f = float(f) if f else 0
    g = float(g) if g else 0

    # .M/.L/.E 분리코드
    if code.endswith(".M") or code.endswith(".L") or code.endswith(".E"):
        base = code[:-2]
        if _is_std_code(base) or _std_nearby(ws, row):
            return parent
        if base.startswith("GZZ"):
            if p_nz >= 2: return parent
            if code.endswith(".M"): return classify_material(name, base)
            if code.endswith(".L"): return "A"
            return "B"
        if code.endswith(".M"): return classify_material(name, base)
        if code.endswith(".L"): return "A"
        return "B"

    if _is_std_code(code):
        return parent

    if code.startswith("GZZ"):
        if p_nz >= 2: return parent
        if e != 0 and f == 0 and g == 0: return classify_material(name, code)
        if f != 0 and e == 0 and g == 0: return "A"
        if g != 0 and e == 0 and f == 0: return "B"
        return parent

    # E열만
    if e != 0 and f == 0 and g == 0:
        if code.startswith("L"): return "A"
        return classify_material(name, code)
    # F열만
    if f != 0 and e == 0 and g == 0:
        return "A"
    # G열만
    if g != 0 and e == 0 and f == 0:
        return classify_expense(code, name, parent)
    # 복합
    return parent


# ================================================================
#  메인
# ================================================================
def process_file(input_path):
    print(f"\n물가변동 비목군 자동분류 v2.0")
    print(f"{'='*50}")
    print(f"파일: {input_path}")

    wb_data = openpyxl.load_workbook(input_path, data_only=True)
    ws_data = wb_data.active
    wb_edit = openpyxl.load_workbook(input_path)
    ws_edit = wb_edit.active
    mx = ws_data.max_row
    print(f"시트: {ws_data.title} | 행수: {mx:,}")

    # PASS 1
    print("\n[1] 상위 품목 분류 중...")
    hc = 0
    for r in range(8, mx + 1):
        b = ws_data.cell(row=r, column=2).value
        if not (b and str(b).startswith("#.")): continue
        c = ws_data.cell(row=r, column=3).value
        d = ws_data.cell(row=r, column=4).value
        e = ws_data.cell(row=r, column=5).value
        f = ws_data.cell(row=r, column=6).value
        g = ws_data.cell(row=r, column=7).value
        ws_edit.cell(row=r, column=1).value = classify_header(r, c, d, e, f, g, ws_data, mx)
        hc += 1
    print(f"    → {hc:,}건 완료")

    # PASS 2
    print("[2] 하위 행 분류 중...")
    pb, pnz, sc = "Z", 0, 0
    for r in range(8, mx + 1):
        b = ws_data.cell(row=r, column=2).value
        h = ws_data.cell(row=r, column=8).value
        if b and str(b).startswith("#."):
            a = ws_edit.cell(row=r, column=1).value
            if a: pb = str(a)
            pe = float(ws_data.cell(row=r, column=5).value or 0)
            pf = float(ws_data.cell(row=r, column=6).value or 0)
            pg = float(ws_data.cell(row=r, column=7).value or 0)
            pnz = sum(1 for v in [pe, pf, pg] if v != 0)
            continue
        if h is not None and h != 0:
            c = ws_data.cell(row=r, column=3).value
            d = ws_data.cell(row=r, column=4).value
            e = ws_data.cell(row=r, column=5).value
            f = ws_data.cell(row=r, column=6).value
            g = ws_data.cell(row=r, column=7).value
            ws_edit.cell(row=r, column=1).value = classify_sub(c, d, e, f, g, pb, pnz, ws_data, r)
            sc += 1
        else:
            ws_edit.cell(row=r, column=1).value = None
    print(f"    → {sc:,}건 완료")

    # PASS 3: 판단 필요 항목에 ? 표시
    print("[3] 검토 필요(?) 항목 표시 중...")
    review_count = 0
    # 상위 비목군 추적용
    cur_parent = "Z"
    for r in range(8, mx + 1):
        b = ws_data.cell(row=r, column=2).value
        a = ws_edit.cell(row=r, column=1).value
        if not a:
            continue
        a_str = str(a)

        if b and str(b).startswith("#."):
            cur_parent = a_str
            e_val = float(ws_data.cell(row=r, column=5).value or 0)
            f_val = float(ws_data.cell(row=r, column=6).value or 0)
            g_val = float(ws_data.cell(row=r, column=7).value or 0)
            total = abs(e_val) + abs(f_val) + abs(g_val)
            nz = sum(1 for v in [e_val, f_val, g_val] if v != 0)
            d_val = str(ws_data.cell(row=r, column=4).value or "")

            need_review = False

            # [유형2] 재료비 85%↑인 복합 G → D일 수도 있음
            # p76 예외: 설치비가 재료비의 %이면 재료비 비목
            if a_str.startswith("G") and nz >= 2 and total > 0:
                mat_ratio = abs(e_val) / total * 100
                if mat_ratio > 85:
                    need_review = True

            # [유형3] Z인데 계측비 → B or G1 가능
            if a_str == "Z" and "계측" in d_val:
                need_review = True

            # [유형3] Z인데 인쇄/복사 → D 가능
            if a_str == "Z" and any(k in d_val for k in ["인쇄", "CD-ROM", "복사"]):
                need_review = True

            if need_review and "?" not in a_str:
                ws_edit.cell(row=r, column=1).value = a_str + "?"
                cur_parent = a_str + "?"
                review_count += 1

    print(f"    → {review_count:,}건 표시 완료")

    # 통계
    stats = {}
    tot = 0
    for r in range(8, mx + 1):
        a = ws_edit.cell(row=r, column=1).value
        if a and str(a).strip():
            k = str(a)
            stats[k] = stats.get(k, 0) + 1
            tot += 1

    labels = {
        "A": "노무비", "B": "기계경비", "C": "광산품", "D": "공산품",
        "E": "전력수도가스폐기물", "F": "농림수산품",
        "G1": "표준시장(토목)", "G2": "표준시장(건축)",
        "G3": "표준시장(기계)", "G4": "표준시장(전기)",
        "G5": "표준시장(정보통신)", "Z": "기타비목",
    }
    print(f"\n{'='*50}")
    print(f" 분류 결과  (총 {tot:,}건)")
    print(f"{'='*50}")
    review_total = 0
    for k in sorted(stats.keys()):
        lbl = labels.get(k.rstrip("?"), "")
        mark = " ← 검토필요" if "?" in k else ""
        print(f"  {k:>5} {lbl:<16}: {stats[k]:>6,}건{mark}")
        if "?" in k:
            review_total += stats[k]
    z_cnt = sum(v for k, v in stats.items() if k.startswith("Z"))
    z_pct = z_cnt / tot * 100 if tot else 0
    print(f"{'-'*50}")
    print(f"  Z(기타) 비율: {z_pct:.1f}%")
    print(f"  검토필요(?) : {review_total:,}건 ← 이 행만 확인하세요")
    print(f"{'='*50}")

    # 저장
    fn = os.path.basename(input_path)
    base, ext = os.path.splitext(fn)
    out = os.path.join(os.path.dirname(os.path.abspath(input_path)), f"{base}_비목분류{ext}")
    try:
        wb_edit.save(out)
    except OSError:
        out = os.path.join(os.getcwd(), f"{base}_비목분류{ext}")
        wb_edit.save(out)
    print(f"\n저장: {out}")
    return out


if __name__ == "__main__":
    if len(sys.argv) >= 2:
        # 드래그 & 드롭 또는 CMD에서 파일 경로 전달
        f = sys.argv[1]
    else:
        # 파일 선택 UI
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            print("\n  엑셀 파일을 선택하세요...")
            f = filedialog.askopenfilename(
                title="산근 엑셀 파일 선택",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            root.destroy()
            if not f:
                print("  파일을 선택하지 않았습니다.")
                input("\n  Enter를 누르면 종료합니다...")
                sys.exit(0)
        except Exception:
            print("물가변동 비목군 자동분류 v2.0")
            print("-" * 40)
            print("사용법: classification_of_cost_item_groups.exe [엑셀파일]")
            print("  또는 exe를 실행하면 파일 선택 창이 열립니다.")
            input("\n  Enter를 누르면 종료합니다...")
            sys.exit(1)

    if not os.path.exists(f):
        print(f"오류: 파일 없음 → {f}")
        input("\nEnter를 누르면 종료합니다...")
        sys.exit(1)

    process_file(f)
    input("\nEnter를 누르면 종료합니다...")
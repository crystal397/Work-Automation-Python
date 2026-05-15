"""월별급여현황표 xlsx → SalaryEntry[] 자동 추출.

㈜D 양식 (C공구(철도)) 기준 디폴트 매핑. 다른 회사 양식은 mapping yaml 로 오버라이드.

표 구조 (디폴트)
- 1행: 제목
- 2행: 현장명
- 3행: 사원명 | 직무 | 주민번호 | YY년N월 | ... | YY년N월
- 4행: 서브헤더 '지급액' 반복
- 5행~: 인원 데이터

산식 (보고서 4.4.2.2 / 4.3.2.2 일치 보장)
- 인원의 활동 월 = xlsx 에서 지급액 > 0 인 월
- 인원의 활동 시작/종료 = (첫 활동월 1일 ~ 마지막 활동월 말일) ∩ 산정기간
- ① 급여 실비 = Σ (월 지급액 × min(겹친 일수 / 그 월 일수, 1))
- ② 일수 = (종료 - 시작 + 1)
- ③ 퇴직급여충당금 = ① / 12
- ④ 소계 = ① + ③
- ⑤ 추정일수, ⑥ 1일 평균노무비, ⑦ 추정 소계 (B 컬럼) 는 별도 산식
"""

from __future__ import annotations

import re
from calendar import monthrange
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from contract_meta.models import Sourced
from cost_aggregation.audit import make_xlsx_source


@dataclass
class SalaryRecord:
    name: str
    role: str
    monthly: list[tuple[date, int, str]] = field(default_factory=list)   # [(month_start, amount, cell)]
    name_cell: str = ""
    role_cell: str = ""


@dataclass
class ExtractedSalary:
    name: Sourced[str]
    role: Sourced[str]
    salary_actual_krw: Sourced[int]            # ① 급여 실비 (일할 적용)
    days_actual: Sourced[int]                  # ② 일수
    severance_actual_krw: Sourced[int]         # ③ 퇴직급여충당금 = ① / 12
    subtotal_actual_krw: Sourced[int]          # ④ 소계 = ① + ③
    period_start_eff: date
    period_end_eff: date


def extract_salaries(
    xlsx_path: str | Path,
    period_start: date,
    period_end: date,
    *,
    sheet_index: int = 0,
    name_col: int = 0,
    role_col: int = 1,
    header_row: int | None = None,
    data_start_row: int | None = None,
) -> tuple[list[ExtractedSalary], list[str]]:
    """월별급여현황표 xlsx 에서 인원별 ① 급여 실비 (일할 적용) + ② 일수 + ③④ 산식 자동 계산."""
    wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    ws = wb.worksheets[sheet_index]
    sheet_name = ws.title
    file_name = Path(xlsx_path).name

    rows = list(ws.iter_rows(values_only=False))
    h_row, month_cols = _find_month_header(rows, header_row)
    if not month_cols:
        return [], [f"월별 헤더 찾기 실패: {xlsx_path}"]

    start_row = data_start_row if data_start_row is not None else h_row + 2

    out: list[ExtractedSalary] = []
    warnings: list[str] = []
    for row in rows[start_row:]:
        if not row or row[name_col].value is None:
            continue
        name = str(row[name_col].value).strip()
        if not name or name.startswith("계") or name.startswith("합계"):
            continue
        role_v = row[role_col].value if role_col < len(row) else None
        role = str(role_v).strip() if role_v else ""

        active_months: list[tuple[date, int, str]] = []
        for col_idx, month_date in month_cols:
            if col_idx >= len(row):
                continue
            cell = row[col_idx]
            v = cell.value
            if v is None:
                continue
            try:
                amount = int(round(float(v)))
            except (TypeError, ValueError):
                continue
            if amount <= 0:
                continue
            active_months.append((month_date, amount, cell.coordinate))

        if not active_months:
            continue

        first_month = active_months[0][0]
        last_month = active_months[-1][0]
        person_start = first_month
        person_end = date(last_month.year, last_month.month, monthrange(last_month.year, last_month.month)[1])
        eff_start = max(person_start, period_start)
        eff_end = min(person_end, period_end)
        if eff_start > eff_end:
            continue

        salary = 0
        for mdate, amount, _coord in active_months:
            m_first = mdate
            m_last = date(mdate.year, mdate.month, monthrange(mdate.year, mdate.month)[1])
            seg_start = max(m_first, eff_start)
            seg_end = min(m_last, eff_end)
            if seg_start > seg_end:
                continue
            m_days = (m_last - m_first).days + 1
            seg_days = (seg_end - seg_start).days + 1
            if seg_days == m_days:
                salary += amount
            else:
                salary += int(round(amount * seg_days / m_days))

        days = (eff_end - eff_start).days + 1
        severance = salary // 12 if salary else 0
        subtotal = salary + severance

        name_src = make_xlsx_source(file_name, sheet_name, row[name_col].coordinate, raw=name, label="사원명")
        role_src = make_xlsx_source(file_name, sheet_name, row[role_col].coordinate, raw=role, label="직무")
        salary_src = make_xlsx_source(
            file_name, sheet_name, active_months[0][2],
            raw=f"활동기간 {first_month}~{last_month}, 산정 {eff_start}~{eff_end}",
            label=f"① 급여 실비 (일할 적용, 활동월 {len(active_months)}개)",
        )
        days_src = make_xlsx_source(
            file_name, sheet_name, active_months[0][2],
            raw=f"{eff_start}~{eff_end}",
            label="② 일수 (산정기간 ∩ 활동기간)",
        )
        sev_src = make_xlsx_source(file_name, sheet_name, active_months[0][2], raw=f"{salary}/12", label="③ 퇴직급여충당금")
        sub_src = make_xlsx_source(file_name, sheet_name, active_months[0][2], raw=f"{salary}+{severance}", label="④ 소계")

        out.append(ExtractedSalary(
            name=Sourced[str](value=name, _source=name_src),
            role=Sourced[str](value=role, _source=role_src),
            salary_actual_krw=Sourced[int](value=salary, _source=salary_src),
            days_actual=Sourced[int](value=days, _source=days_src),
            severance_actual_krw=Sourced[int](value=severance, _source=sev_src),
            subtotal_actual_krw=Sourced[int](value=subtotal, _source=sub_src),
            period_start_eff=eff_start,
            period_end_eff=eff_end,
        ))

    return out, warnings


_MONTH_HEADER_PAT = re.compile(r"(\d{2})년\s*(\d{1,2})월")


def _find_month_header(rows, hint_row: int | None) -> tuple[int, list[tuple[int, date]]]:
    """헤더 행(YY년N월 패턴 다수 등장)을 찾고 (col_idx, month_first_date) 리스트 반환."""
    range_ = [hint_row] if hint_row is not None else range(len(rows))
    last_year = None
    for i in range_ if hint_row is not None else range(len(rows)):
        row = rows[i]
        cols: list[tuple[int, date]] = []
        last_year = None
        for c_idx, cell in enumerate(row):
            v = cell.value
            if v is None:
                continue
            s = str(v).strip()
            m = _MONTH_HEADER_PAT.search(s)
            if m:
                yy = int(m.group(1))
                mm = int(m.group(2))
                cols.append((c_idx, date(2000 + yy, mm, 1)))
                last_year = yy
            elif s.endswith("월") and last_year is not None:
                m2 = re.match(r"(\d{1,2})월", s)
                if m2:
                    cols.append((c_idx, date(2000 + last_year, int(m2.group(1)), 1)))
        if len(cols) >= 6:
            return i, cols
    return -1, []

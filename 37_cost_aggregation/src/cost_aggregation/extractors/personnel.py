"""인원투입현황 xlsx → Personnel[] 자동 추출.

표 구조 (C공구(철도) 케이스 기준)
- 헤더 행: 월별 컬럼 (2401, 2402, ..., 2512) — int 또는 'YYMM' 문자열
- 그룹 행: 직무 그룹명 (일반직 / 안전 / 품질 / 미화 / 서무 / 취사 / 관리 / 토목 ...) — 행 첫 셀이 그룹명, 나머지 셀은 그 그룹의 인원 합계
- 인원 행: 이름 + 월별 셀 (1 = 그 달 재직, None = 미재직)

산정 대상 기간(start, end)을 받아 각 인원의 실제 in/out 월을 산출.
"""

from __future__ import annotations

import re
from calendar import monthrange
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from cost_aggregation.audit import make_xlsx_source
from cost_aggregation.models import Personnel
from contract_meta.models import Sourced


_GROUP_LABELS = {"일반직", "현채직", "안전", "품질", "미화", "서무", "취사", "관리", "토목"}


@dataclass
class _ParsedRow:
    name: str
    role_group: str
    months_present: list[str]   # ['2412', '2501', ...]
    row_idx: int


def extract_personnel(
    xlsx_path: str | Path,
    *,
    affiliation: str,
    period_start: date,
    period_end: date,
    sheet_index: int = 0,
) -> tuple[list[Personnel], list[str]]:
    """인원투입현황 xlsx 에서 산정 대상 인원 리스트 추출.

    period_start ~ period_end 와 겹치는 월에 1 표시가 있는 인원만 포함.
    각 인원의 period_start/end 는 표시된 월 범위와 산정구간의 교집합.
    """
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.worksheets[sheet_index]
    sheet_name = ws.title

    rows = list(ws.iter_rows(values_only=False))
    header_row, month_columns = _find_header_row(rows)
    if header_row is None:
        return [], [f"헤더 행을 찾지 못함: {xlsx_path}"]

    current_group = None
    parsed: list[_ParsedRow] = []
    for r_idx, row in enumerate(rows[header_row + 1:], start=header_row + 2):
        if not row or row[0].value is None:
            continue
        first = str(row[0].value).strip()
        if first in _GROUP_LABELS:
            current_group = first
            continue
        if current_group is None:
            continue
        months: list[str] = []
        for col_idx, month_str in month_columns:
            cell = row[col_idx] if col_idx < len(row) else None
            if cell is not None and cell.value is not None:
                try:
                    if int(cell.value) == 1:
                        months.append(month_str)
                except (TypeError, ValueError):
                    pass
        if months:
            parsed.append(_ParsedRow(name=first, role_group=current_group, months_present=months, row_idx=r_idx))

    personnel: list[Personnel] = []
    warnings: list[str] = []
    for p in parsed:
        eff_start, eff_end = _intersect(p.months_present, period_start, period_end)
        if eff_start is None:
            continue
        src_name = make_xlsx_source(xlsx_path, sheet_name, f"A{p.row_idx}", raw=p.name)
        src_role = make_xlsx_source(xlsx_path, sheet_name, f"A{p.row_idx}", raw=p.role_group, label="직무 그룹")
        src_start = make_xlsx_source(xlsx_path, sheet_name, f"row{p.row_idx}", raw=p.months_present[0], label="월별 재직 표기")
        src_end = make_xlsx_source(xlsx_path, sheet_name, f"row{p.row_idx}", raw=p.months_present[-1], label="월별 재직 표기")
        personnel.append(Personnel(
            affiliation=Sourced[str](value=affiliation, _source=src_name),
            name=Sourced[str](value=p.name, _source=src_name),
            role=Sourced[str](value=p.role_group, _source=src_role),
            period_start=Sourced[date](value=eff_start, _source=src_start),
            period_end=Sourced[date](value=eff_end, _source=src_end),
        ))

    return personnel, warnings


def _find_header_row(rows) -> tuple[int | None, list[tuple[int, str]]]:
    """월별 헤더(2401, 2402...) 가 있는 행 인덱스와 (col_index, 'YYMM') 리스트 반환."""
    month_pat = re.compile(r"^2[0-9]{3}$")
    for i, row in enumerate(rows):
        cols: list[tuple[int, str]] = []
        for c_idx, cell in enumerate(row):
            v = cell.value
            if v is None:
                continue
            s = str(v).strip()
            if month_pat.match(s):
                cols.append((c_idx, s))
        if len(cols) >= 6:
            return i, cols
    return None, []


def _intersect(months: list[str], period_start: date, period_end: date) -> tuple[date | None, date | None]:
    """월 표기 리스트와 산정구간의 교집합을 (start, end) date 로 반환."""
    month_dates = [(_month_first(m), _month_last(m)) for m in months]
    overlapping = [(a, b) for (a, b) in month_dates if not (b < period_start or a > period_end)]
    if not overlapping:
        return None, None
    s = max(overlapping[0][0], period_start)
    e = min(overlapping[-1][1], period_end)
    return s, e


def _month_first(yymm: str) -> date:
    yy, mm = int(yymm[:2]), int(yymm[2:])
    return date(2000 + yy, mm, 1)


def _month_last(yymm: str) -> date:
    yy, mm = int(yymm[:2]), int(yymm[2:])
    last_day = monthrange(2000 + yy, mm)[1]
    return date(2000 + yy, mm, last_day)

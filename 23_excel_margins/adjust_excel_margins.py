"""
지정한 경로 내의 모든 엑셀 파일을 열어,
각 시트의 왼쪽 + 오른쪽 여백의 합이 2(인치)이면
왼쪽 1, 오른쪽 1로 변경하는 스크립트.

openpyxl의 여백 단위는 '인치'입니다.
 - 1 inch = 2.54 cm
 - 따라서 좌우 합 2인치 ≈ 5.08cm
만약 "cm 기준"으로 동작시키고 싶다면 TARGET_SUM 과 NEW_MARGIN 값을
cm → inch 로 환산해서 넣으면 됩니다. (예: 2cm → 2/2.54)
"""

import os
import sys
from pathlib import Path
from openpyxl import load_workbook

# ===== 설정값 (cm 단위로 입력) =====
TARGET_SUM_CM = 2.0   # 좌+우 여백 합이 이 값(cm)이면 조정
NEW_LEFT_CM = 1.0     # 새 왼쪽 여백 (cm)
NEW_RIGHT_CM = 1.0    # 새 오른쪽 여백 (cm)
TOLERANCE = 0.02      # 비교 허용 오차 (cm) — 반올림 오차 대응
# ===================================

# openpyxl 내부 단위는 inch이므로 cm → inch 변환
CM_PER_INCH = 2.54
TARGET_SUM = TARGET_SUM_CM / CM_PER_INCH
NEW_LEFT = NEW_LEFT_CM / CM_PER_INCH
NEW_RIGHT = NEW_RIGHT_CM / CM_PER_INCH


def adjust_margins_in_file(file_path: Path) -> dict:
    """
    하나의 엑셀 파일을 열어 모든 시트의 여백을 검사/수정.
    반환값: {'changed': 변경된 시트 수, 'checked': 검사한 시트 수, 'sheets': [상세내역]}
    """
    result = {"changed": 0, "checked": 0, "sheets": []}

    try:
        wb = load_workbook(file_path)
    except Exception as e:
        print(f"  [오류] 파일을 열 수 없습니다: {e}")
        return result

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        margins = ws.page_margins
        left = margins.left if margins.left is not None else 0
        right = margins.right if margins.right is not None else 0
        total = left + right
        result["checked"] += 1

        if abs(total - TARGET_SUM) < (TOLERANCE / CM_PER_INCH):
            margins.left = NEW_LEFT
            margins.right = NEW_RIGHT
            result["changed"] += 1
            result["sheets"].append(
                f"    ✓ [{sheet_name}] 좌+우={total*CM_PER_INCH:.2f}cm "
                f"→ 좌={NEW_LEFT_CM}cm, 우={NEW_RIGHT_CM}cm"
            )
        else:
            result["sheets"].append(
                f"    - [{sheet_name}] 좌+우={total*CM_PER_INCH:.2f}cm (대상 아님, 건너뜀)"
            )

    if result["changed"] > 0:
        try:
            wb.save(file_path)
        except Exception as e:
            print(f"  [오류] 저장 실패: {e}")
            result["changed"] = 0

    wb.close()
    return result


def process_folder(folder_path: str):
    """지정 폴더 내 모든 .xlsx 파일을 처리."""
    folder = Path(folder_path)

    if not folder.exists():
        print(f"[오류] 경로가 존재하지 않습니다: {folder}")
        return
    if not folder.is_dir():
        print(f"[오류] 폴더가 아닙니다: {folder}")
        return

    # 엑셀 파일 수집 (.xlsx, .xlsm). 임시파일(~$로 시작)은 제외
    excel_files = [
        f for f in folder.glob("*.xlsx") if not f.name.startswith("~$")
    ] + [
        f for f in folder.glob("*.xlsm") if not f.name.startswith("~$")
    ]

    if not excel_files:
        print(f"[알림] {folder} 안에 엑셀 파일이 없습니다.")
        return

    print(f"[시작] {folder} (대상 파일 {len(excel_files)}개)\n")

    total_changed_sheets = 0
    total_changed_files = 0

    for file_path in excel_files:
        print(f"▶ {file_path.name}")
        r = adjust_margins_in_file(file_path)
        for line in r["sheets"]:
            print(line)
        if r["changed"] > 0:
            total_changed_files += 1
            total_changed_sheets += r["changed"]
        print()

    print("=" * 50)
    print(f"처리 완료: 파일 {total_changed_files}/{len(excel_files)}개 변경, "
          f"시트 총 {total_changed_sheets}개 수정")


def _run_gui():
    """tkinter GUI 모드 — 더블클릭 실행 시 진입."""
    import tkinter as tk
    from tkinter import filedialog, scrolledtext

    root = tk.Tk()
    root.title("엑셀 여백 조정")
    root.resizable(False, False)

    # ── 상단: 경로 입력 + 찾아보기 ──────────────────────
    frame_top = tk.Frame(root, padx=10, pady=8)
    frame_top.pack(fill="x")

    tk.Label(frame_top, text="폴더 경로:").pack(side="left")

    path_var = tk.StringVar()
    entry = tk.Entry(frame_top, textvariable=path_var, width=55)
    entry.pack(side="left", padx=(4, 4))

    def browse():
        d = filedialog.askdirectory(title="엑셀 파일이 있는 폴더 선택")
        if d:
            path_var.set(d)

    tk.Button(frame_top, text="찾아보기", command=browse).pack(side="left")

    # ── 중단: 결과 텍스트 ────────────────────────────────
    log = scrolledtext.ScrolledText(root, width=80, height=22, state="disabled",
                                    font=("Consolas", 9))
    log.pack(padx=10, pady=(0, 4))

    def write_log(msg: str):
        log.configure(state="normal")
        log.insert("end", msg + "\n")
        log.see("end")
        log.configure(state="disabled")
        root.update_idletasks()

    # ── 하단: 실행 버튼 ──────────────────────────────────
    frame_bot = tk.Frame(root, padx=10, pady=6)
    frame_bot.pack()

    def run():
        folder = path_var.get().strip()
        if not folder:
            write_log("[오류] 폴더를 선택하세요.")
            return

        btn_run.configure(state="disabled")
        log.configure(state="normal")
        log.delete("1.0", "end")
        log.configure(state="disabled")

        # process_folder의 print를 로그 창으로 리다이렉트
        import io, contextlib
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            process_folder(folder)
        for line in buf.getvalue().splitlines():
            write_log(line)

        btn_run.configure(state="normal")

    btn_run = tk.Button(frame_bot, text="  실행  ", command=run,
                        bg="#0078d4", fg="white", font=("", 10, "bold"),
                        padx=16, pady=4)
    btn_run.pack()

    root.mainloop()


if __name__ == "__main__":
    # CLI 모드: 인자가 있으면 터미널에서 직접 실행
    if len(sys.argv) >= 2:
        process_folder(sys.argv[1])
    else:
        _run_gui()
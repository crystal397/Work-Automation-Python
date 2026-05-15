"""월별 작업자 × 일자 매트릭스 xlsx 출력 (v0.1).

24 v12 의 정교한 양식(인쇄 영역·머지 셀·페이지 나눔·휴일 표시)은 v0.2 에서 본격 재구현.
v0.1 은 시연용 단순 시트 — 한 시트당 한 공종 (예: '01.가시설') 의 작업자 × 일자 매트릭스 + 공수 + 노무비.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from crash_aggregation.models import MonthlyCrash


def write_monthly_matrix(monthly: MonthlyCrash, out_path: str | Path) -> Path:
    out_path = Path(out_path)
    wb = Workbook()
    wb.remove(wb.active)

    # 공종별 시트
    by_gongjong: dict[str, list] = defaultdict(list)
    for w in monthly.workers:
        by_gongjong[w.gongjong.value].append(w)

    for gongjong, workers in sorted(by_gongjong.items()):
        ws = wb.create_sheet(gongjong[:31])
        _render_matrix(ws, monthly.year, monthly.month, workers)

    # 요약 시트 — 카테고리별 합계
    s = wb.create_sheet("00.요약", 0)
    _render_summary(s, monthly)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _render_matrix(ws, year: int, month: int, workers: list) -> None:
    # 헤더
    ws["A1"] = f"{year}년 {month}월 — {ws.title}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")

    # 표 헤더
    ws["A3"] = "No"
    ws["B3"] = "성명"
    ws["C3"] = "단가"
    ws["D3"] = "공수합계"
    ws["E3"] = "노무비합계"
    ws["F3"] = "카테고리(주)"
    ws["G3"] = "비고"
    for col in "ABCDEFG":
        ws[f"{col}3"].font = Font(bold=True)
        ws[f"{col}3"].alignment = Alignment(horizontal="center")
        ws[f"{col}3"].fill = PatternFill("solid", fgColor="DDDDDD")

    # 일자 헤더 (1~31)
    for d in range(1, 32):
        col = get_column_letter(7 + d)
        ws[f"{col}3"] = d
        ws[f"{col}3"].font = Font(bold=True)
        ws[f"{col}3"].alignment = Alignment(horizontal="center")

    # 데이터
    for i, w in enumerate(workers, start=1):
        r = 3 + i
        ws[f"A{r}"] = i
        ws[f"B{r}"] = w.name.value
        ws[f"C{r}"] = w.unit_price_per_day.value if w.unit_price_per_day else None
        ws[f"D{r}"] = round(w.total_manday, 2)
        ws[f"E{r}"] = w.total_krw
        # 카테고리 첫 1개
        cats = {d.category.value for d in w.days}
        ws[f"F{r}"] = ", ".join(sorted(cats))

        day_md: dict[int, float] = {}
        for d in w.days:
            day_md[d.day] = day_md.get(d.day, 0) + d.manday.value
        for d in range(1, 32):
            col = get_column_letter(7 + d)
            v = day_md.get(d)
            if v is not None:
                ws[f"{col}{r}"] = round(v, 2)

    # 열 너비
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    for col in "CDE":
        ws.column_dimensions[col].width = 12
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    for d in range(1, 32):
        ws.column_dimensions[get_column_letter(7 + d)].width = 5


def _render_summary(ws, monthly: MonthlyCrash) -> None:
    ws["A1"] = f"{monthly.year}년 {monthly.month:02d}월 노무비 요약"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "구분"
    ws["B3"] = "값"
    ws["A3"].font = Font(bold=True)
    ws["B3"].font = Font(bold=True)

    rows = [
        ("총 작업자(고유)", len({w.name.value for w in monthly.workers})),
        ("총 공수(일)", round(sum(w.total_manday for w in monthly.workers), 2)),
        ("총 노무비(원)", sum(w.total_krw for w in monthly.workers)),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v

    ws["A8"] = "공종별"
    ws["A8"].font = Font(bold=True)
    ws["A9"] = "공종"
    ws["B9"] = "인원(고유)"
    ws["C9"] = "공수합계"
    ws["D9"] = "노무비합계"
    for col in "ABCD":
        ws[f"{col}9"].font = Font(bold=True)
        ws[f"{col}9"].fill = PatternFill("solid", fgColor="DDDDDD")

    g_buckets: dict[str, dict] = defaultdict(lambda: {"names": set(), "md": 0.0, "krw": 0})
    for w in monthly.workers:
        g_buckets[w.gongjong.value]["names"].add(w.name.value)
        g_buckets[w.gongjong.value]["md"] += w.total_manday
        g_buckets[w.gongjong.value]["krw"] += w.total_krw

    for i, (g, v) in enumerate(sorted(g_buckets.items(), key=lambda x: -x[1]["krw"]), start=10):
        ws[f"A{i}"] = g
        ws[f"B{i}"] = len(v["names"])
        ws[f"C{i}"] = round(v["md"], 2)
        ws[f"D{i}"] = v["krw"]

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16

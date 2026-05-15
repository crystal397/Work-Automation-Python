"""24 패턴: 월별 공종별 xlsx → WorkerMonth[] 자동 추출.

24 v12 의 read_source_file 로직을 audit 규약 위에서 재구현.

시트 구조
- 시트 이름: 'NN.공종명' 패턴 (`01.가시설`, `02.형틀목공`, ...)
- 시트 안에 여러 작업자 '블록' — B열 '작업자명' 헤더로 구분
- 블록 구조:
    [hdr]    B='작업자명'  C='No.'  D='공수'  E='계'  F='구분'  G='작업내용'
    [hdr+1]  B=<작업자명>  C=1     D=공수1   E=노무비1 F=카테고리1
    [hdr+2]  B='일 기준'   C=2     D=공수2   ...
    [hdr+3]  B=<단가>      C=3     D=공수3   ...
    [hdr+4]  B='시간 기준' C=4     D=공수4   ...
    [hdr+5]  B=<시간단가>  C=5     D=공수5   ...
    ... (1~31일 데이터)

v12 핵심 수정: hdr 행 자체에 1일차 데이터가 들어 있어 hdr+1 부터 스캔하면 누락. hdr 부터 스캔.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from contract_meta.models import Sourced
from crash_aggregation.audit import make_xlsx_source
from crash_aggregation.models import WorkerDay, WorkerMonth


SHEET_PATTERN = re.compile(r"^(0?\d+)\.\s*(.+)$")


@dataclass
class _BlockBounds:
    hdr_idx: int        # 0-base
    name_idx: int
    end_idx: int        # exclusive


def extract_workers(
    xlsx_path: str | Path,
    *,
    year: int,
    month: int,
) -> tuple[list[WorkerMonth], list[str]]:
    """월별 xlsx → 공종별 작업자 데이터."""
    wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    file_name = Path(xlsx_path).name
    out: list[WorkerMonth] = []
    warnings: list[str] = []

    for sheet_name in wb.sheetnames:
        m = SHEET_PATTERN.match(sheet_name.strip())
        if not m:
            continue
        gongjong = m.group(2).strip()
        ws = wb[sheet_name]

        rows = []
        for r in range(1, ws.max_row + 1):
            rows.append({
                "b": ws.cell(r, 2).value,
                "c": ws.cell(r, 3).value,
                "d": ws.cell(r, 4).value,
                "e": ws.cell(r, 5).value,
                "f": ws.cell(r, 6).value,
                "g": ws.cell(r, 7).value,
                "h": ws.cell(r, 8).value,
            })

        worker_hdr_idxs = [
            i for i, row in enumerate(rows)
            if isinstance(row["b"], str) and row["b"].strip() == "작업자명"
        ]

        for bi, hdr_idx in enumerate(worker_hdr_idxs):
            name_idx = hdr_idx + 1
            if name_idx >= len(rows):
                continue
            worker_name_raw = rows[name_idx]["b"]
            if worker_name_raw is None:
                continue
            worker_name = str(worker_name_raw).strip()
            if not worker_name or worker_name == "작업자":
                continue
            end_idx = worker_hdr_idxs[bi + 1] if bi + 1 < len(worker_hdr_idxs) else len(rows)
            bounds = _BlockBounds(hdr_idx=hdr_idx, name_idx=name_idx, end_idx=end_idx)

            worker = _extract_worker_block(rows, bounds, sheet_name, file_name, gongjong)
            if worker is not None:
                out.append(worker)

    return out, warnings


def _extract_worker_block(
    rows: list[dict],
    b: _BlockBounds,
    sheet_name: str,
    file_name: str,
    gongjong: str,
) -> WorkerMonth | None:
    name_row_excel = b.name_idx + 1
    worker_name = str(rows[b.name_idx]["b"]).strip()

    name_src = make_xlsx_source(
        file_name, sheet_name, f"B{name_row_excel}",
        raw=worker_name, label="작업자명",
    )
    gongjong_src = make_xlsx_source(
        file_name, sheet_name, f"sheet={sheet_name}",
        raw=gongjong, label="공종",
    )

    unit_day_price: Sourced[int] | None = None
    unit_hour_price: Sourced[int] | None = None
    for j in range(b.hdr_idx, b.end_idx):
        b_val = rows[j]["b"]
        if isinstance(b_val, str):
            if "일 기준" in b_val and j + 1 < b.end_idx:
                up = rows[j + 1]["b"]
                if isinstance(up, (int, float)) and up:
                    unit_day_price = Sourced[int](
                        value=int(up),
                        _source=make_xlsx_source(
                            file_name, sheet_name, f"B{j+2}",
                            raw=str(up), label="일 기준 단가",
                        ),
                    )
            elif "시간 기준" in b_val and j + 1 < b.end_idx:
                up = rows[j + 1]["b"]
                if isinstance(up, (int, float)) and up:
                    unit_hour_price = Sourced[int](
                        value=int(up),
                        _source=make_xlsx_source(
                            file_name, sheet_name, f"B{j+2}",
                            raw=str(up), label="시간 기준 단가",
                        ),
                    )

    days: list[WorkerDay] = []
    for j in range(b.hdr_idx, b.end_idx):
        r = rows[j]
        c_val, d_val, e_val = r["c"], r["d"], r["e"]
        if d_val is None:
            continue
        try:
            day = int(c_val)
            manday = float(d_val)
        except (TypeError, ValueError):
            continue
        if not (1 <= day <= 31):
            continue
        try:
            amount = int(round(float(e_val))) if e_val is not None else 0
        except (TypeError, ValueError):
            amount = 0
        category = _classify(r["f"], r["h"])
        excel_row = j + 1
        days.append(WorkerDay(
            day=day,
            manday=Sourced[float](
                value=manday,
                _source=make_xlsx_source(
                    file_name, sheet_name, f"D{excel_row}",
                    raw=str(manday), label=f"{day}일 공수",
                ),
            ),
            amount_krw=Sourced[int](
                value=amount,
                _source=make_xlsx_source(
                    file_name, sheet_name, f"E{excel_row}",
                    raw=str(amount), label=f"{day}일 노무비",
                ),
            ),
            category=Sourced[str](
                value=category,
                _source=make_xlsx_source(
                    file_name, sheet_name, f"F{excel_row}",
                    raw=f"F={r['f']}, H={r['h']}", label="카테고리",
                ),
            ),
            work_content=_optional_sourced(r["g"], file_name, sheet_name, f"G{excel_row}", "작업내용"),
        ))

    if not days:
        return None

    return WorkerMonth(
        name=Sourced[str](value=worker_name, _source=name_src),
        gongjong=Sourced[str](value=gongjong, _source=gongjong_src),
        unit_price_per_day=unit_day_price,
        unit_price_per_hour=unit_hour_price,
        days=days,
        block_row_range=(b.hdr_idx + 1, b.end_idx),
    )


def _optional_sourced(value, file_name, sheet_name, cell, label) -> Sourced[str] | None:
    if value is None or value == "":
        return None
    return Sourced[str](
        value=str(value),
        _source=make_xlsx_source(file_name, sheet_name, cell, raw=str(value), label=label),
    )


def _classify(f_val, h_val) -> str:
    """24 v12 의 classify_category 단순 버전. f열·h열의 문자열을 그대로 카테고리로 사용."""
    candidates = [s for s in (f_val, h_val) if isinstance(s, str) and s.strip()]
    if not candidates:
        return "기타"
    return candidates[0].strip()

# -*- coding: utf-8 -*-
"""
돌관공사비 산출근거 - 노무비 출력일보 자동 작성 스크립트 (v8)

[v8 주요 변경 사항]
1) 합계행 E열 '명' 표시 - 실제 근로자 수 (예: "25명") 직접 기입
2) 한국 공휴일 자동 '휴' 표시 (holidays 라이브러리 사용)
   - 대체공휴일, 선거일, 임시공휴일 모두 포함
3) 파일 분할 저장 (연도별)
   - 전체 통합 파일 1개 + '노무비 출력일보' 시트만 연도별 별도 파일
   - 예: 노무비출력일보_2023.xlsx, 노무비출력일보_2024.xlsx, ...

[v7 주요 변경]
- 페이지 분할 (21명/페이지), 각 페이지마다 헤더 재출력
- AB(단가), AC(총액) 소스에서 자동 추출

[사용 방법]
1. 아래 '사용자 설정'의 경로를 본인 환경에 맞게 수정
2. 터미널에서:
       pip install openpyxl holidays
       python mandays_report_automation_v8.py
"""

import os
import re
import glob
import shutil
import calendar
import time
import zipfile
import io
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Color

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))
except ImportError:
    pass


# ==============================================================
# 사용자 설정  (.env 또는 아래 기본값)
# ==============================================================
_BASE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.environ.get('CRASH_TEMPLATE_PATH',
    os.path.join(_BASE, 'sample_돌관공사비 산출근거.xlsx'))
SOURCE_DIR    = os.environ.get('CRASH_SOURCE_DIR',
    os.path.join(_BASE, 'source'))
OUTPUT_PATH   = os.environ.get('CRASH_OUTPUT_PATH',
    os.path.join(_BASE, '돌관공사비_산출근거_자동생성_v8.xlsx'))


# ==============================================================
# 소스 파일 구조 설정
# ==============================================================
SRC_SHEET_PATTERN = re.compile(r'^(0?\d+)\.\s*(.+)$')


# ==============================================================
# 템플릿 구조 설정 (케이씨산업 샘플 기준)
# ==============================================================
TEMPLATE_SHEET = '노무비 출력일보'
HEADER_ROWS = 6            # 섹션 헤더 높이
WORKER_ROW_SPAN = 2        # 근로자 1명당 행 수
PROTO_SECTION_SIZE = 48    # 견본 섹션 원본 크기 (20명용)
PROTO_WORKER_COUNT = 20

# 페이지 분할 (v7 신규)
WORKERS_PER_PAGE = 21      # 한 페이지당 최대 근로자 수
PAGE_ROWS = HEADER_ROWS + WORKERS_PER_PAGE * WORKER_ROW_SPAN  # 6 + 42 = 48

COL_LABEL = 2   # B: 카테고리 라벨
COL_ORDER = 3   # C: 순서/"계"
COL_TYPE  = 4   # D: 공종
COL_NAME  = 5   # E: 성명
DAY_COL_START = 6   # F: 1일 or 16일
COL_UNIT_PRICE = 28  # AB: 노무비 단가 (v8)
COL_TOTAL = 29       # AC: 노무비 총액 (v8)

CATEGORIES = ['본선', '복합', '삼성', '기타']
SECTION_TITLE_KEYWORDS = ['노무비 명세서', '노임내역']

COLOR_SATURDAY = '0000FF'
COLOR_SUNDAY   = 'FF0000'


# ==============================================================
# 유틸
# ==============================================================

def safe_set(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True


def extract_year_month_from_filename(filename):
    name = os.path.splitext(os.path.basename(filename))[0]
    for pattern in [
        r'(20\d{2})\s*년\s*(\d{1,2})\s*월',
        r'(20\d{2})[.\-_ ](\d{1,2})(?!\d)',
        r'(20\d{2})(\d{2})',
        r'(?<!\d)(2[3-9])[.\-_ ](\d{1,2})(?!\d)',
        r'(?<!\d)(2[3-9])\s*년\s*(\d{1,2})\s*월',
    ]:
        m = re.search(pattern, name)
        if m:
            y_raw, mo = int(m.group(1)), int(m.group(2))
            y = y_raw if y_raw >= 100 else 2000 + y_raw
            if 1 <= mo <= 12:
                return y, mo
    return None, None


def classify_category(f_val, h_val):
    combined = ''
    if f_val is not None:
        combined += str(f_val)
    if h_val is not None:
        combined += ' ' + str(h_val)
    if '본선' in combined:
        return '본선'
    if '복합' in combined:
        return '복합'
    if '삼성' in combined:
        return '삼성'
    return '기타'


# ==============================================================
# 소스 파일 읽기
# ==============================================================

def read_source_file(filepath):
    """소스 파일을 읽어 근로자별 데이터를 반환.

    반환 구조:
      result[category][gongjong][worker_name] = {
          'days': {day: manday, ...},   # 날짜별 공수 (기존 구조)
          'unit_price': float or None,  # B열 '일 기준' 아래 값 (AB열용)
          'total': float or None,       # E열 근로자 블록 마지막 합계 (AC열용)
      }
    """
    wb = load_workbook(filepath, data_only=True, read_only=False)  # read_only=False: max_row 안정
    result = {cat: {} for cat in CATEGORIES}

    try:
        for sheet_name in wb.sheetnames:
            m = SRC_SHEET_PATTERN.match(sheet_name.strip())
            if not m:
                continue
            gongjong = m.group(2).strip()
            ws = wb[sheet_name]

            # 모든 셀을 한 번에 메모리로 로드 (B, C, D, E, F, H열)
            # 시트 끝까지 순회하면서 근로자 블록을 찾음
            rows_data = []
            for r in range(1, ws.max_row + 1):
                rows_data.append({
                    'b': ws.cell(r, 2).value,
                    'c': ws.cell(r, 3).value,
                    'd': ws.cell(r, 4).value,
                    'e': ws.cell(r, 5).value,
                    'f': ws.cell(r, 6).value,
                    'h': ws.cell(r, 8).value,
                })

            # 블록 경계 찾기: B열에 '작업자명' 있는 행
            worker_header_idxs = [
                i for i, row in enumerate(rows_data)
                if isinstance(row['b'], str) and row['b'].strip() == '작업자명'
            ]

            for bi, hdr_idx in enumerate(worker_header_idxs):
                name_idx = hdr_idx + 1
                if name_idx >= len(rows_data):
                    continue
                worker_name = rows_data[name_idx]['b']
                if worker_name is None or not str(worker_name).strip():
                    continue
                worker_name = str(worker_name).strip()
                if worker_name == '작업자':  # dummy entry
                    continue

                # 블록 끝: 다음 '작업자명' 직전 (또는 시트 끝)
                if bi + 1 < len(worker_header_idxs):
                    block_end_idx = worker_header_idxs[bi + 1]
                else:
                    block_end_idx = len(rows_data)

                # 단가 추출: B열 '일 기준' 아래 행의 B열 값
                unit_price = None
                for j in range(hdr_idx, block_end_idx):
                    b = rows_data[j]['b']
                    if isinstance(b, str) and '일 기준' in b:
                        if j + 1 < block_end_idx:
                            up = rows_data[j + 1]['b']
                            if isinstance(up, (int, float)):
                                unit_price = float(up)
                            else:
                                try:
                                    unit_price = float(up) if up else None
                                except (TypeError, ValueError):
                                    unit_price = None
                        break

                # 총액 추출: 블록 마지막 행들 중 E열에 숫자가 있는 마지막 값
                # (블록 하단 "계" 행의 E열이 숫자로 되어 있음)
                total = None
                for j in range(block_end_idx - 1, hdr_idx, -1):
                    e = rows_data[j]['e']
                    if isinstance(e, (int, float)) and e != 0:
                        total = float(e)
                        break

                # 날짜별 공수 수집 + 카테고리별 분리
                # 한 근로자가 날짜별로 다른 카테고리에 투입될 수 있으므로,
                # 날짜별로 F/H열을 보고 카테고리 분류
                daily_by_cat = {cat: {} for cat in CATEGORIES}
                for j in range(name_idx, block_end_idx):
                    row = rows_data[j]
                    c_val = row['c']
                    d_val = row['d']
                    f_val = row['f']
                    h_val = row['h']
                    if d_val is None:
                        continue
                    try:
                        day = int(c_val)
                        manday = float(d_val)
                    except (TypeError, ValueError):
                        continue
                    if not (1 <= day <= 31):
                        continue
                    cat = classify_category(f_val, h_val)
                    daily_by_cat[cat][day] = daily_by_cat[cat].get(day, 0) + manday

                # 카테고리별로 저장 (그 카테고리에 투입된 날이 있을 때만)
                for cat, days in daily_by_cat.items():
                    if not days:
                        continue
                    # 같은 근로자가 여러 공종에 있을 수 있으므로 setdefault
                    gongjong_bucket = result[cat].setdefault(gongjong, {})
                    if worker_name in gongjong_bucket:
                        # 동일 공종·동일 근로자의 날짜 합산
                        existing = gongjong_bucket[worker_name]
                        for day, m in days.items():
                            existing['days'][day] = existing['days'].get(day, 0) + m
                    else:
                        gongjong_bucket[worker_name] = {
                            'days': dict(days),
                            'unit_price': unit_price,
                            'total': total,
                        }
    finally:
        wb.close()
    return result
    return result


# ==============================================================
# 템플릿 블록 스냅샷
# ==============================================================

def snapshot_cells(ws, start_row, num_rows, max_col):
    cells = []
    for offset in range(num_rows):
        r = start_row + offset
        for col in range(1, max_col + 1):
            cell = ws.cell(row=r, column=col)
            style_snap = None
            if cell.has_style:
                style_snap = {
                    'font': copy(cell.font),
                    'border': copy(cell.border),
                    'fill': copy(cell.fill),
                    'number_format': cell.number_format,
                    'alignment': copy(cell.alignment),
                }
            val = cell.value
            # ArrayFormula는 특별한 마커로 저장 (paste 시 다시 ArrayFormula로 만들기 위해)
            # Excel에서 SUMPRODUCT+IF 같은 수식은 ArrayFormula여야 제대로 계산됨
            is_array = False
            if val is not None and type(val).__name__ == 'ArrayFormula':
                try:
                    t = val.text
                    val = t if t.startswith('=') else '=' + t
                    is_array = True
                except Exception:
                    val = None
            if val is not None or style_snap is not None:
                cells.append((offset, col, val, style_snap, is_array))

    merges = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= start_row and mr.max_row < start_row + num_rows:
            merges.append((mr.min_row - start_row, mr.max_row - start_row,
                           mr.min_col, mr.max_col))

    row_heights = {}
    for offset in range(num_rows):
        r = start_row + offset
        if r in ws.row_dimensions and ws.row_dimensions[r].height is not None:
            row_heights[offset] = ws.row_dimensions[r].height

    return {
        'cells': cells, 'merges': merges, 'row_heights': row_heights,
        'orig_start': start_row, 'num_rows': num_rows,
    }


_formula_row_re = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')

def shift_formula_rows(formula, delta, shift_absolute=False, absolute_delta=None):
    """수식의 행 번호를 이동.
    - 상대 참조(예: F7)는 delta만큼 이동
    - 절대 참조(예: $F$3):
        shift_absolute=False (기본): 그대로 유지
        shift_absolute=True: absolute_delta만큼 이동 (None이면 delta 사용)
    """
    if shift_absolute and absolute_delta is None:
        absolute_delta = delta
    def repl(m):
        col_abs, col, row_abs, row_num = m.group(1), m.group(2), m.group(3), m.group(4)
        if row_abs == '$':
            if shift_absolute:
                return f'{col_abs}{col}{row_abs}{int(row_num) + absolute_delta}'
            return m.group(0)
        return f'{col_abs}{col}{row_abs}{int(row_num) + delta}'
    return _formula_row_re.sub(repl, formula)


def paste_block(ws, block, dst_start_row, formula_base_row=None, shift_absolute=False, absolute_delta=None):
    if formula_base_row is None:
        formula_base_row = block['orig_start']
    delta = dst_start_row - formula_base_row

    for item in block['cells']:
        if len(item) == 5:
            offset, col, value, style_snap, is_array = item
        else:
            offset, col, value, style_snap = item
            is_array = False
        r = dst_start_row + offset
        cell = ws.cell(row=r, column=col)
        if isinstance(cell, MergedCell):
            continue
        if isinstance(value, str) and value.startswith('='):
            shifted = shift_formula_rows(value, delta,
                                          shift_absolute=shift_absolute,
                                          absolute_delta=absolute_delta)
            if is_array:
                from openpyxl.worksheet.formula import ArrayFormula
                from openpyxl.utils import get_column_letter
                ref = f'{get_column_letter(col)}{r}'
                cell.value = ArrayFormula(ref, shifted)
            else:
                cell.value = shifted
        else:
            cell.value = value
        if style_snap is not None:
            cell.font = style_snap['font']
            cell.border = style_snap['border']
            cell.fill = style_snap['fill']
            cell.number_format = style_snap['number_format']
            cell.alignment = style_snap['alignment']

    for offset, height in block['row_heights'].items():
        ws.row_dimensions[dst_start_row + offset].height = height

    # 병합: ws.merge_cells() 대신 직접 CellRange 추가 (더 빠름)
    from openpyxl.worksheet.cell_range import CellRange
    for (min_off, max_off, min_col, max_col) in block['merges']:
        try:
            rng = CellRange(
                min_col=min_col, min_row=dst_start_row + min_off,
                max_col=max_col, max_row=dst_start_row + max_off
            )
            ws.merged_cells.ranges.add(rng)
        except Exception:
            pass


# ==============================================================
# 요일/휴일 처리
# ==============================================================

def apply_weekend_colors(ws, header_start_row, year, month):
    """섹션 헤더의 일자 셀에 요일 색상 적용.
    헤더 레이아웃 (header_start_row 기준 0-based):
      offset 0: 타이틀
      offset 1: 구분
      offset 2: 컬럼명
      offset 3: 1~15일의 '휴' 표시행 (각 열이 해당 일자 위)
      offset 4: 1~15일 숫자
      offset 5: 16~31일의 '휴' 표시행
      offset 6: 16~31일 숫자  ← 실제로는 offset 5가 위쪽 "휴", offset 6이 숫자가 아님

    실제 템플릿 구조:
      offset 2 (row 3): 1~15일 '휴' 행 (숫자 위)
      offset 3 (row 4): 1~15일 숫자
      offset 4 (row 5): 16~31일 '휴' 행 (숫자 위)
      offset 5 (row 6): 16~31일 숫자
    """
    days_in_month = calendar.monthrange(year, month)[1]
    # v8: 공휴일 정보 (한국 공휴일 + 대체공휴일 + 선거일)
    holiday_days = _get_holidays_for_month(year, month)

    row_hol_1_15  = header_start_row + 2  # 1~15일 '휴' 표시 (숫자 위)
    row_num_1_15  = header_start_row + 3  # 1~15일 숫자
    row_hol_16_31 = header_start_row + 4  # 16~31일 '휴' 표시
    row_num_16_31 = header_start_row + 5  # 16~31일 숫자

    # 1) 1~15일 '휴' 표시행 전체 초기화 (헤더 텍스트 F3~U3에 있을 수도 있으니 F열부터만)
    # 단, F3~K3 같은 열은 '순 서'/'공종'/'성 명' 등 헤더가 아닌 날짜 위 자리임.
    # 안전하게: F~U 범위 (col 6~21)만 처리.
    for col in range(DAY_COL_START, 22):  # F(6) ~ U(21)
        # 1~15일 휴 행 초기화 (기본값 None으로)
        safe_set(ws, row_hol_1_15, col, None)
        # 16~31일 휴 행 초기화
        safe_set(ws, row_hol_16_31, col, None)

    # 2) 1~15일
    for day in range(1, 16):
        col = DAY_COL_START + (day - 1)  # F=6 for day=1
        if day > days_in_month:
            safe_set(ws, row_num_1_15, col, None)
            safe_set(ws, row_hol_1_15, col, None)
            continue
        safe_set(ws, row_num_1_15, col, day)
        wd = calendar.weekday(year, month, day)
        is_holiday = day in holiday_days
        num_cell = ws.cell(row=row_num_1_15, column=col)
        hol_cell = ws.cell(row=row_hol_1_15, column=col)
        if wd == 6 or is_holiday:  # 일요일 or 공휴일: 빨강 + "휴"
            _apply_font_color(num_cell, COLOR_SUNDAY)
            safe_set(ws, row_hol_1_15, col, '휴')
            _apply_font_color(hol_cell, COLOR_SUNDAY)
        elif wd == 5:  # 토요일: 파랑 (휴 표시 없음)
            _apply_font_color(num_cell, COLOR_SATURDAY)
        else:
            _apply_font_color(num_cell, '000000')

    # 3) 16~31일
    for day in range(16, 32):
        col = DAY_COL_START + (day - 16)  # F=6 for day=16
        if day > days_in_month:
            safe_set(ws, row_num_16_31, col, None)
            safe_set(ws, row_hol_16_31, col, None)
            continue
        safe_set(ws, row_num_16_31, col, day)
        wd = calendar.weekday(year, month, day)
        is_holiday = day in holiday_days
        num_cell = ws.cell(row=row_num_16_31, column=col)
        hol_cell = ws.cell(row=row_hol_16_31, column=col)
        if wd == 6 or is_holiday:
            _apply_font_color(num_cell, COLOR_SUNDAY)
            safe_set(ws, row_hol_16_31, col, '휴')
            _apply_font_color(hol_cell, COLOR_SUNDAY)
        elif wd == 5:
            _apply_font_color(num_cell, COLOR_SATURDAY)
        else:
            _apply_font_color(num_cell, '000000')


# v8: 한국 공휴일 조회 (holidays 라이브러리)
_holidays_cache = {}

def _get_holidays_for_month(year, month):
    """해당 년/월의 공휴일(일) 집합을 반환.
    대체공휴일, 선거일, 임시공휴일 모두 포함.
    holidays 라이브러리가 없으면 빈 set 반환 (일요일만 '휴' 표시).
    """
    key = (year, month)
    if key in _holidays_cache:
        return _holidays_cache[key]
    try:
        import holidays as _h
        kr = _h.KR(years=[year])
        result = set()
        for d in kr:
            if d.year == year and d.month == month:
                result.add(d.day)
        _holidays_cache[key] = result
        return result
    except Exception:
        _holidays_cache[key] = set()
        return set()


def _apply_font_color(cell, rgb_hex):
    if isinstance(cell, MergedCell):
        return
    old = cell.font
    cell.font = Font(
        name=old.name, size=old.size, bold=old.bold, italic=old.italic,
        vertAlign=old.vertAlign, underline=old.underline, strike=old.strike,
        color=Color(rgb='FF' + rgb_hex),
    )


# ==============================================================
# 섹션 생성 (가변 길이)
# ==============================================================

def build_section(ws, dst_start_row, header_block, worker_block, sum_block,
                  workers, year, month, category_label,
                  proto_first_worker_row, proto_section_start):
    """
    섹션을 페이지 분할 방식으로 생성.

    workers: [(gongjong, name, {'days': {...}, 'unit_price': ..., 'total': ...}), ...]

    레이아웃:
      페이지 1: 헤더(6행) + 근로자 21명(42행) [마지막 페이지가 아니면 합계 없음]
      페이지 2: 헤더(6행) + 근로자 21명 또는 남은 인원 + 빈 슬롯들 + 합계(2행, 마지막 페이지만)

    근로자가 M명일 때 필요한 페이지 수:
      pages = ceil(M / WORKERS_PER_PAGE)
      모든 페이지는 헤더 + 21명 슬롯 크기 고정 (48행)
      마지막 페이지는 그 아래에 합계 2행 추가

    합계는 "첫 페이지 근로자 첫 행"부터 "마지막 페이지 근로자 마지막 실제 행"까지.
    단, 빈 슬롯이 있는 페이지에서는 슬롯 범위 전체를 합계 범위로 잡아도 무방 (빈 행은 0).
    이 구현은 모든 페이지의 전체 슬롯 범위를 합계 범위로 사용 (계산상 동일).
    """
    M = len(workers)
    if M == 0:
        return dst_start_row
    # 섹션 전체 행 이동량 (절대 참조 헤더 이동용)
    section_delta = dst_start_row - proto_section_start

    # 필요한 페이지 수
    import math
    num_pages = max(1, math.ceil(M / WORKERS_PER_PAGE))

    # 각 페이지의 첫 행 (모든 페이지가 HEADER_ROWS + WORKERS_PER_PAGE*2 = 48행 고정)
    page_start_rows = [dst_start_row + p * PAGE_ROWS for p in range(num_pages)]

    # 페이지별 헤더 + 근로자 슬롯 생성
    for p in range(num_pages):
        page_start = page_start_rows[p]
        # 각 페이지는 견본 섹션의 "헤더 위치"와 대응하므로
        # absolute_delta = page_start - proto_section_start
        page_delta = page_start - proto_section_start

        # 1) 헤더
        paste_block(ws, header_block, page_start)

        # 타이틀 C열 연월 치환
        title_cell = ws.cell(row=page_start, column=COL_ORDER)
        if not isinstance(title_cell, MergedCell) and isinstance(title_cell.value, str):
            title_cell.value = re.sub(r'20\d{2}년\s*\d{1,2}월',
                                        f'{year}년 {month:02d}월',
                                        title_cell.value)
        # AE 기준 문구
        ae_cell = ws.cell(row=page_start, column=31)
        if not isinstance(ae_cell, MergedCell) and isinstance(ae_cell.value, str):
            ae_cell.value = f'{year}년 {month}월 기준'

        # 요일 색상
        apply_weekend_colors(ws, page_start, year, month)

        # 2) 이 페이지의 근로자 슬롯 21개 - 실제 있는 근로자만큼은 데이터 기입,
        # 없는 슬롯은 빈 슬롯으로 둠 (수식만 존재)
        worker_first_row_on_page = page_start + HEADER_ROWS
        for slot_idx in range(WORKERS_PER_PAGE):
            row_top = worker_first_row_on_page + slot_idx * WORKER_ROW_SPAN
            row_bot = row_top + 1
            global_idx = p * WORKERS_PER_PAGE + slot_idx  # 전체 중 몇 번째 근로자

            # 근로자 슬롯 블록 paste (절대 참조는 이 페이지 헤더로 이동)
            paste_block(ws, worker_block, row_top,
                        formula_base_row=proto_first_worker_row,
                        shift_absolute=True,
                        absolute_delta=page_delta)

            # 견본의 일자 데이터 제거 (F~U, 수식이 아닌 값만)
            for col in range(DAY_COL_START, 22):
                for rr in (row_top, row_bot):
                    cell = ws.cell(row=rr, column=col)
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith('=')):
                        cell.value = None
            # AB/AC/AD의 기본값(270000 등)도 제거
            for col in (COL_UNIT_PRICE, COL_TOTAL, 30):  # AB, AC, AD
                for rr in (row_top, row_bot):
                    cell = ws.cell(row=rr, column=col)
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith('=')):
                        cell.value = None
            # 견본의 C(순서) D(공종) E(성명) 값 제거 (빈 슬롯/실데이터 모두에 대해)
            # 실데이터 근로자가 있으면 뒤에서 다시 기입됨
            for col in (COL_ORDER, COL_TYPE, COL_NAME):
                for rr in (row_top, row_bot):
                    cell = ws.cell(row=rr, column=col)
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith('=')):
                        cell.value = None

            # 실제 근로자가 있으면 데이터 기입
            if global_idx < M:
                gongjong, name, info = workers[global_idx]
                day_dict = info.get('days', {}) if isinstance(info, dict) else info
                unit_price = info.get('unit_price') if isinstance(info, dict) else None
                total = info.get('total') if isinstance(info, dict) else None

                safe_set(ws, row_top, COL_ORDER, global_idx + 1)
                safe_set(ws, row_top, COL_TYPE, gongjong)
                safe_set(ws, row_top, COL_NAME, name)
                # 일자별 공수
                for day, manday in day_dict.items():
                    if 1 <= day <= 15:
                        col = DAY_COL_START + (day - 1)
                        safe_set(ws, row_top, col, manday)
                    elif 16 <= day <= 31:
                        col = DAY_COL_START + (day - 16)
                        safe_set(ws, row_bot, col, manday)
                # AB 단가, AC 총액 (v7 신규)
                if unit_price is not None:
                    safe_set(ws, row_top, COL_UNIT_PRICE, unit_price)
                if total is not None:
                    safe_set(ws, row_top, COL_TOTAL, total)

    # 3) 합계 (마지막 페이지 슬롯 끝 바로 다음, 2행)
    last_page_start = page_start_rows[-1]
    sum_first_row = last_page_start + PAGE_ROWS  # 마지막 페이지 헤더6+슬롯42 = 48행 다음

    paste_block(ws, sum_block, sum_first_row)

    # 합계 수식 범위 재조정:
    # 첫 페이지 근로자 첫 행 ~ 마지막 페이지 근로자 마지막 슬롯 행
    # 모든 페이지가 21명 슬롯을 가지므로, 마지막 페이지 슬롯 끝 = sum_first_row - 1
    # 첫 페이지 근로자 첫 행 = dst_start_row + HEADER_ROWS
    worker_first_abs = dst_start_row + HEADER_ROWS
    worker_last_abs = sum_first_row - 1

    range_re = re.compile(r'(\$?[A-Z]{1,3}\$?)(\d+)(:)(\$?[A-Z]{1,3}\$?)(\d+)')
    for offset in range(sum_block['num_rows']):
        r = sum_first_row + offset
        for col in range(1, 35):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith('='):
                def repl(m, _end=worker_last_abs, _start=worker_first_abs):
                    col1, row1, sep, col2, row2 = m.group(1), int(m.group(2)), m.group(3), m.group(4), int(m.group(5))
                    if row2 == _end:
                        row1 = _start
                    return f'{col1}{row1}{sep}{col2}{row2}'
                cell.value = range_re.sub(repl, cell.value)

    # 라벨 + "계" + "N명"
    safe_set(ws, sum_first_row, COL_LABEL, f'{year}년{month:02d}월({category_label})')
    safe_set(ws, sum_first_row, COL_ORDER, '계')
    # E열: 실제 근로자 수 ("25명" 형태로 직접 기입, 수식 아님)
    safe_set(ws, sum_first_row, COL_NAME, f'{M}명')

    # 섹션 전체 행 수 반환
    total_rows = num_pages * PAGE_ROWS + 2  # 페이지들 + 합계 2행
    return dst_start_row + total_rows


# ==============================================================
# External Link 제거 (속도 최적화의 핵심)
# ==============================================================

def strip_external_links(xlsx_path):
    """xlsx 파일에서 external link 관련 XML + definedNames 전부 제거.
    openpyxl은 external link 때문에 로드·저장이 수십 초 느려짐.
    또한 external link를 참조하는 definedNames는 Excel에서 복구 경고 원인.
    """
    tmp_path = xlsx_path + '.stripped.tmp'

    removed = 0
    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                # external link 관련 파일 제외
                if 'externalLink' in item:
                    removed += 1
                    continue
                data = zin.read(item)

                # workbook.xml 수정 (externalReferences + definedNames 제거)
                if item == 'xl/workbook.xml':
                    text = data.decode('utf-8')
                    text = re.sub(r'<externalReferences>.*?</externalReferences>',
                                  '', text, flags=re.DOTALL)
                    # definedNames 전체 제거 (대부분 external link 참조)
                    text = re.sub(r'<definedNames>.*?</definedNames>',
                                  '', text, flags=re.DOTALL)
                    data = text.encode('utf-8')

                elif item == 'xl/_rels/workbook.xml.rels':
                    text = data.decode('utf-8')
                    text = re.sub(r'<Relationship[^/]+externalLink[^/]+/>',
                                  '', text)
                    data = text.encode('utf-8')

                elif item == '[Content_Types].xml':
                    text = data.decode('utf-8')
                    text = re.sub(r'<Override[^/]+externalLink[^/]+/>',
                                  '', text)
                    data = text.encode('utf-8')

                zout.writestr(item, data)

    os.replace(tmp_path, xlsx_path)
    return removed


# ==============================================================
# 메인
# ==============================================================

def main():
    t_start = time.time()
    print("=" * 70)
    print("돌관공사비 노무비 출력일보 자동 작성 (v8)")
    print("=" * 70)

    if not os.path.exists(TEMPLATE_PATH):
        print(f"❌ 템플릿 파일 없음: {TEMPLATE_PATH}")
        return
    if not os.path.exists(SOURCE_DIR):
        print(f"❌ 소스 폴더 없음: {SOURCE_DIR}")
        return

    out_dir = os.path.dirname(OUTPUT_PATH)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    t0 = time.time()
    shutil.copy2(TEMPLATE_PATH, OUTPUT_PATH)
    print(f"✓ 템플릿 복사: {OUTPUT_PATH} ({time.time()-t0:.1f}초)")

    # External link 제거 (속도 최적화)
    t0 = time.time()
    removed = strip_external_links(OUTPUT_PATH)
    print(f"✓ External link 제거: {removed}개 ({time.time()-t0:.1f}초)")

    t0 = time.time()
    wb = load_workbook(OUTPUT_PATH)
    print(f"✓ 워크북 로드: {time.time()-t0:.1f}초")

    if TEMPLATE_SHEET not in wb.sheetnames:
        print(f"❌ '{TEMPLATE_SHEET}' 시트 없음")
        return
    ws = wb[TEMPLATE_SHEET]

    # 견본 섹션 위치
    proto_section_start = None
    for r in range(1, min(ws.max_row + 1, 200)):
        v = ws.cell(row=r, column=COL_ORDER).value
        if isinstance(v, str) and any(kw in v for kw in SECTION_TITLE_KEYWORDS):
            proto_section_start = r
            break
    if proto_section_start is None:
        print(f"❌ 견본 섹션 없음")
        return
    print(f"✓ 견본 섹션 위치: row {proto_section_start}")

    # 세 블록 스냅샷
    max_col = ws.max_column
    header_block = snapshot_cells(ws, proto_section_start, HEADER_ROWS, max_col)
    proto_first_worker_row = proto_section_start + HEADER_ROWS
    worker_block = snapshot_cells(ws, proto_first_worker_row, WORKER_ROW_SPAN, max_col)
    proto_sum_start = proto_section_start + PROTO_SECTION_SIZE - 2
    sum_block = snapshot_cells(ws, proto_sum_start, 2, max_col)

    print(f"  헤더 블록: {len(header_block['cells'])}셀")
    print(f"  근로자 블록: {len(worker_block['cells'])}셀")
    print(f"  합계 블록: {len(sum_block['cells'])}셀")

    # 시트 재생성
    del wb[TEMPLATE_SHEET]
    ws = wb.create_sheet(TEMPLATE_SHEET)
    print("✓ 기존 시트 삭제 후 재생성")

    # 소스 파일
    source_files = []
    for ext in ('*.xlsx', '*.xlsm'):
        source_files.extend(glob.glob(os.path.join(SOURCE_DIR, ext)))
    source_files = [f for f in source_files if not os.path.basename(f).startswith('~$')]
    source_files.sort()
    if not source_files:
        print(f"❌ 소스 폴더에 엑셀 파일 없음")
        return
    print(f"✓ 소스 파일 {len(source_files)}개 발견")

    # 소스 파일 읽기 (시간 측정)
    t_src = time.time()
    monthly_data = {}
    for idx_f, fp in enumerate(source_files, 1):
        fn = os.path.basename(fp)
        y, m = extract_year_month_from_filename(fn)
        if y is None:
            print(f"  ⚠ 연/월 인식 실패: {fn}")
            continue
        t_file = time.time()
        try:
            data = read_source_file(fp)
        except Exception as e:
            print(f"    ❌ 읽기 실패: {e}")
            continue
        monthly_data[(y, m)] = data
        print(f"  [{idx_f}/{len(source_files)}] {y}년 {m:02d}월 ({time.time()-t_file:.1f}초)")
    print(f"✓ 소스 파일 읽기 완료: {time.time()-t_src:.1f}초")

    if not monthly_data:
        print("❌ 읽은 데이터 없음")
        return

    # 섹션 생성 (가변 길이)
    t_sec = time.time()
    months_sorted = sorted(monthly_data.keys())
    cur_row = 1
    section_log = []
    print(f"\n[섹션 생성 시작] 총 {len(months_sorted)}개월 처리")

    for idx_m, (y, m) in enumerate(months_sorted, start=1):
        t_month = time.time()
        month_data = monthly_data[(y, m)]
        cat_counts = {}
        for cat in CATEGORIES:
            tot = sum(len(w) for w in month_data.get(cat, {}).values())
            if tot > 0:
                cat_counts[cat] = tot
        summary = ', '.join(f"{c}={n}명" for c, n in cat_counts.items())

        for cat in CATEGORIES:
            cat_data = month_data.get(cat, {})
            flat = []
            for gj in sorted(cat_data.keys()):
                for worker in sorted(cat_data[gj].keys()):
                    flat.append((gj, worker, cat_data[gj][worker]))
            if not flat:
                continue

            next_row = build_section(
                ws, cur_row,
                header_block, worker_block, sum_block,
                flat, y, m, cat,
                proto_first_worker_row=proto_first_worker_row,
                proto_section_start=proto_section_start,
            )
            section_log.append((cur_row, y, m, cat, len(flat), next_row - cur_row))
            cur_row = next_row

        print(f"  [{idx_m}/{len(months_sorted)}] {y}년 {m:02d}월  → {summary or '데이터 없음'} ({time.time()-t_month:.1f}초, 누적 {cur_row}행)")

    print(f"\n✓ 총 {len(section_log)}개 섹션 생성 완료: {time.time()-t_sec:.1f}초")
    for (r, y, m, cat, n, length) in section_log[:15]:
        print(f"   row {r:>6}: {y}년 {m:02d}월 [{cat}] {n}명 ({length}행)")
    if len(section_log) > 15:
        print(f"   ... (총 {len(section_log)}개)")

    # 저장 - 통합 파일
    t_save = time.time()
    print("\n통합 파일 저장 중...")
    wb.save(OUTPUT_PATH)
    print(f"✓ 통합 파일 저장 완료: {time.time()-t_save:.1f}초 ({os.path.getsize(OUTPUT_PATH)/1024/1024:.1f} MB)")

    # 연도별 분할 파일 생성 (v8 신규)
    # 각 연도별로 '노무비 출력일보' 시트만 포함하는 별도 파일 저장
    t_split = time.time()
    print("\n연도별 파일 분할 중...")
    years_data = {}  # {year: [(y,m,cat,flat), ...]}
    for (y, m) in months_sorted:
        month_data = monthly_data[(y, m)]
        for cat in CATEGORIES:
            cat_data = month_data.get(cat, {})
            flat = []
            for gj in sorted(cat_data.keys()):
                for worker in sorted(cat_data[gj].keys()):
                    flat.append((gj, worker, cat_data[gj][worker]))
            if not flat:
                continue
            years_data.setdefault(y, []).append((y, m, cat, flat))

    out_base = os.path.splitext(OUTPUT_PATH)[0]  # 확장자 제외 경로

    for year in sorted(years_data.keys()):
        year_path = f"{out_base}_{year}.xlsx"
        t_year = time.time()

        # 원본 템플릿에서 시작 (통합 파일이 아닌 것이 핵심)
        shutil.copy2(TEMPLATE_PATH, year_path)
        # external link + definedNames 제거
        strip_external_links(year_path)

        wb_y = load_workbook(year_path)
        # 다른 시트들 모두 삭제 (노무비 출력일보만 남기기 위해)
        # 주의: 삭제된 시트를 참조하는 수식은 #REF!가 되지만,
        # 이 파일은 '노무비 출력일보' 시트만 필요하므로 무관.
        for sheet_name in list(wb_y.sheetnames):
            if sheet_name != TEMPLATE_SHEET:
                del wb_y[sheet_name]
        # 노무비 출력일보 시트도 새로 생성
        if TEMPLATE_SHEET in wb_y.sheetnames:
            del wb_y[TEMPLATE_SHEET]
        ws_y = wb_y.create_sheet(TEMPLATE_SHEET)

        # 해당 연도 섹션만 paste
        cur_row_y = 1
        year_sec_count = 0
        for (y, m, cat, flat) in years_data[year]:
            cur_row_y = build_section(
                ws_y, cur_row_y,
                header_block, worker_block, sum_block,
                flat, y, m, cat,
                proto_first_worker_row=proto_first_worker_row,
                proto_section_start=proto_section_start,
            )
            year_sec_count += 1

        wb_y.save(year_path)
        size_mb = os.path.getsize(year_path) / 1024 / 1024
        print(f"  ✓ {year}년 ({year_sec_count}개 섹션, {cur_row_y-1}행): {year_path} [{size_mb:.2f} MB, {time.time()-t_year:.1f}초]")

    print(f"✓ 연도별 분할 완료: {time.time()-t_split:.1f}초")

    print(f"\n{'='*70}")
    print(f"✓ 완료!")
    print(f"  통합 파일: {OUTPUT_PATH}")
    print(f"  연도별 파일: {out_base}_YYYY.xlsx (노무비 출력일보 시트만)")
    print(f"  [총 소요 시간] {time.time()-t_start:.1f}초")
    print(f"{'='*70}")
    print("\n※ Excel에서 Ctrl+Alt+F9로 전체 재계산 실행하세요.")


if __name__ == '__main__':
    main()
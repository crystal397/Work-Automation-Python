# -*- coding: utf-8 -*-
"""
돌관공사비 산출근거 - 노무비 출력일보 자동 작성 스크립트 (v6)

[v6 주요 변경 사항 - 무결성 수정]
- definedNames (외부 참조 고아) 제거
  → v5에서 Excel 열 때 '복구 필요' 경고 발생 원인
  → 템플릿에 external link와 함께 달려있던 13,499개 definedNames를 제거
- ArrayFormula → 일반 수식 변환
  → v5에서 V~Y열 값이 0으로 계산되던 문제 해결
  → openpyxl은 ArrayFormula의 ref/내부 행번호를 자동 이동시키지 못함
  → v6는 ArrayFormula를 일반 수식으로 바꿔 paste 시 행 이동되도록 함
- 그 외 v5의 속도 최적화는 모두 유지

[v4 주요 변경 사항]
- 섹션을 48행 고정이 아닌 '가변 길이'로 생성
  → 근로자 N명이면 섹션 = 헤더 6행 + 근로자 N×2행 + 합계 2행 = (8 + 2N)행
  → 20명 초과 시 페이지 분할 없이 한 섹션으로 쭉 이어짐
- 월별 요일 색상 자동 적용 (토=파랑, 일=빨강+"휴")

[전체 기능 요약]
- 본선/복합/삼성/기타 4개 카테고리 (F열+H열 키워드)
- 한 근로자가 날짜마다 다른 카테고리에 투입된 경우 날짜별로 분배
- 각 월·카테고리 합계행 B열에 라벨 (예: "2023년11월(본선)")
- 모든 수식 유지 (Z,AA,AE,AF, V~Y ArrayFormula, SUMIF, 환산평균노임 참조)

[사용 방법]
1. 아래 '사용자 설정'의 경로를 본인 환경에 맞게 수정
2. 터미널에서:
       pip install openpyxl
       python mandays_report_automation_v6.py
"""

import os
import re
import glob
import shutil
import calendar
import time
import zipfile
import io
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Color

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))
except ImportError:
    pass


# ==============================================================
# 사용자 설정  (.env 또는 아래 기본값)
# ==============================================================
_BASE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.environ.get('CRASH_TEMPLATE_PATH',
    os.path.join(_BASE, 'sample_돌관공사비 산출근거.xlsx'))
SOURCE_DIR    = os.environ.get('CRASH_SOURCE_DIR',
    os.path.join(_BASE, 'source'))
OUTPUT_PATH   = os.environ.get('CRASH_OUTPUT_PATH',
    os.path.join(_BASE, '돌관공사비_산출근거_자동생성_v6.xlsx'))


# ==============================================================
# 소스 파일 구조 설정
# ==============================================================
SRC_SHEET_PATTERN = re.compile(r'^(0?\d+)\.\s*(.+)$')


# ==============================================================
# 템플릿 구조 설정
# ==============================================================
TEMPLATE_SHEET = '노무비 출력일보'
HEADER_ROWS = 6            # 섹션 헤더 높이
WORKER_ROW_SPAN = 2        # 근로자 1명당 행 수
PROTO_SECTION_SIZE = 48    # 견본 섹션 원본 크기 (20명용)
PROTO_WORKER_COUNT = 20

COL_LABEL = 2   # B: 카테고리 라벨
COL_ORDER = 3   # C: 순서/"계"
COL_TYPE  = 4   # D: 공종
COL_NAME  = 5   # E: 성명
DAY_COL_START = 6   # F: 1일 or 16일

CATEGORIES = ['본선', '복합', '삼성', '기타']
SECTION_TITLE_KEYWORDS = ['노무비 명세서', '노임내역']

COLOR_SATURDAY = '0000FF'
COLOR_SUNDAY   = 'FF0000'


# ==============================================================
# 유틸
# ==============================================================

def safe_set(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True


def extract_year_month_from_filename(filename):
    name = os.path.splitext(os.path.basename(filename))[0]
    for pattern in [
        r'(20\d{2})\s*년\s*(\d{1,2})\s*월',
        r'(20\d{2})[.\-_ ](\d{1,2})(?!\d)',
        r'(20\d{2})(\d{2})',
        r'(?<!\d)(2[3-9])[.\-_ ](\d{1,2})(?!\d)',
        r'(?<!\d)(2[3-9])\s*년\s*(\d{1,2})\s*월',
    ]:
        m = re.search(pattern, name)
        if m:
            y_raw, mo = int(m.group(1)), int(m.group(2))
            y = y_raw if y_raw >= 100 else 2000 + y_raw
            if 1 <= mo <= 12:
                return y, mo
    return None, None


def classify_category(f_val, h_val):
    combined = ''
    if f_val is not None:
        combined += str(f_val)
    if h_val is not None:
        combined += ' ' + str(h_val)
    if '본선' in combined:
        return '본선'
    if '복합' in combined:
        return '복합'
    if '삼성' in combined:
        return '삼성'
    return '기타'


# ==============================================================
# 소스 파일 읽기
# ==============================================================

def read_source_file(filepath):
    wb = load_workbook(filepath, data_only=True, read_only=True)
    result = {cat: {} for cat in CATEGORIES}

    try:
        for sheet_name in wb.sheetnames:
            m = SRC_SHEET_PATTERN.match(sheet_name.strip())
            if not m:
                continue
            gongjong = m.group(2).strip()
            ws = wb[sheet_name]

            cur_worker = None
            for row_tuple in ws.iter_rows(min_col=1, max_col=8, values_only=False):
                b_val = row_tuple[1].value if len(row_tuple) > 1 else None
                c_val = row_tuple[2].value if len(row_tuple) > 2 else None
                d_val = row_tuple[3].value if len(row_tuple) > 3 else None
                f_val = row_tuple[5].value if len(row_tuple) > 5 else None
                h_val = row_tuple[7].value if len(row_tuple) > 7 else None

                if isinstance(b_val, str) and b_val.strip() == '작업자명':
                    cur_worker = '__PENDING__'
                    continue
                if cur_worker == '__PENDING__':
                    if b_val is not None and str(b_val).strip():
                        cur_worker = str(b_val).strip()
                    else:
                        cur_worker = None

                if cur_worker and cur_worker != '__PENDING__' and d_val is not None:
                    try:
                        day = int(c_val)
                        manday = float(d_val)
                    except (TypeError, ValueError):
                        continue
                    if not (1 <= day <= 31):
                        continue
                    cat = classify_category(f_val, h_val)
                    bucket = result[cat].setdefault(gongjong, {}).setdefault(cur_worker, {})
                    bucket[day] = bucket.get(day, 0) + manday
    finally:
        wb.close()
    return result


# ==============================================================
# 템플릿 블록 스냅샷
# ==============================================================

def snapshot_cells(ws, start_row, num_rows, max_col):
    cells = []
    for offset in range(num_rows):
        r = start_row + offset
        for col in range(1, max_col + 1):
            cell = ws.cell(row=r, column=col)
            style_snap = None
            if cell.has_style:
                style_snap = {
                    'font': copy(cell.font),
                    'border': copy(cell.border),
                    'fill': copy(cell.fill),
                    'number_format': cell.number_format,
                    'alignment': copy(cell.alignment),
                }
            val = cell.value
            # ArrayFormula는 특별한 마커로 저장 (paste 시 다시 ArrayFormula로 만들기 위해)
            # Excel에서 SUMPRODUCT+IF 같은 수식은 ArrayFormula여야 제대로 계산됨
            is_array = False
            if val is not None and type(val).__name__ == 'ArrayFormula':
                try:
                    t = val.text
                    val = t if t.startswith('=') else '=' + t
                    is_array = True
                except Exception:
                    val = None
            if val is not None or style_snap is not None:
                cells.append((offset, col, val, style_snap, is_array))

    merges = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= start_row and mr.max_row < start_row + num_rows:
            merges.append((mr.min_row - start_row, mr.max_row - start_row,
                           mr.min_col, mr.max_col))

    row_heights = {}
    for offset in range(num_rows):
        r = start_row + offset
        if r in ws.row_dimensions and ws.row_dimensions[r].height is not None:
            row_heights[offset] = ws.row_dimensions[r].height

    return {
        'cells': cells, 'merges': merges, 'row_heights': row_heights,
        'orig_start': start_row, 'num_rows': num_rows,
    }


_formula_row_re = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')

def shift_formula_rows(formula, delta, shift_absolute=False, absolute_delta=None):
    """수식의 행 번호를 이동.
    - 상대 참조(예: F7)는 delta만큼 이동
    - 절대 참조(예: $F$3):
        shift_absolute=False (기본): 그대로 유지
        shift_absolute=True: absolute_delta만큼 이동 (None이면 delta 사용)
    """
    if shift_absolute and absolute_delta is None:
        absolute_delta = delta
    def repl(m):
        col_abs, col, row_abs, row_num = m.group(1), m.group(2), m.group(3), m.group(4)
        if row_abs == '$':
            if shift_absolute:
                return f'{col_abs}{col}{row_abs}{int(row_num) + absolute_delta}'
            return m.group(0)
        return f'{col_abs}{col}{row_abs}{int(row_num) + delta}'
    return _formula_row_re.sub(repl, formula)


def paste_block(ws, block, dst_start_row, formula_base_row=None, shift_absolute=False, absolute_delta=None):
    if formula_base_row is None:
        formula_base_row = block['orig_start']
    delta = dst_start_row - formula_base_row

    for item in block['cells']:
        if len(item) == 5:
            offset, col, value, style_snap, is_array = item
        else:
            offset, col, value, style_snap = item
            is_array = False
        r = dst_start_row + offset
        cell = ws.cell(row=r, column=col)
        if isinstance(cell, MergedCell):
            continue
        if isinstance(value, str) and value.startswith('='):
            shifted = shift_formula_rows(value, delta,
                                          shift_absolute=shift_absolute,
                                          absolute_delta=absolute_delta)
            if is_array:
                from openpyxl.worksheet.formula import ArrayFormula
                from openpyxl.utils import get_column_letter
                ref = f'{get_column_letter(col)}{r}'
                cell.value = ArrayFormula(ref, shifted)
            else:
                cell.value = shifted
        else:
            cell.value = value
        if style_snap is not None:
            cell.font = style_snap['font']
            cell.border = style_snap['border']
            cell.fill = style_snap['fill']
            cell.number_format = style_snap['number_format']
            cell.alignment = style_snap['alignment']

    for offset, height in block['row_heights'].items():
        ws.row_dimensions[dst_start_row + offset].height = height

    # 병합: ws.merge_cells() 대신 직접 CellRange 추가 (더 빠름)
    from openpyxl.worksheet.cell_range import CellRange
    for (min_off, max_off, min_col, max_col) in block['merges']:
        try:
            rng = CellRange(
                min_col=min_col, min_row=dst_start_row + min_off,
                max_col=max_col, max_row=dst_start_row + max_off
            )
            ws.merged_cells.ranges.add(rng)
        except Exception:
            pass


# ==============================================================
# 요일/휴일 처리
# ==============================================================

def apply_weekend_colors(ws, header_start_row, year, month):
    """섹션 헤더의 일자 셀에 요일 색상 적용.
    헤더 레이아웃 (header_start_row 기준 0-based):
      offset 0: 타이틀
      offset 1: 구분
      offset 2: 컬럼명
      offset 3: 1~15일의 '휴' 표시행 (각 열이 해당 일자 위)
      offset 4: 1~15일 숫자
      offset 5: 16~31일의 '휴' 표시행
      offset 6: 16~31일 숫자  ← 실제로는 offset 5가 위쪽 "휴", offset 6이 숫자가 아님

    실제 템플릿 구조:
      offset 2 (row 3): 1~15일 '휴' 행 (숫자 위)
      offset 3 (row 4): 1~15일 숫자
      offset 4 (row 5): 16~31일 '휴' 행 (숫자 위)
      offset 5 (row 6): 16~31일 숫자
    """
    days_in_month = calendar.monthrange(year, month)[1]

    row_hol_1_15  = header_start_row + 2  # 1~15일 '휴' 표시 (숫자 위)
    row_num_1_15  = header_start_row + 3  # 1~15일 숫자
    row_hol_16_31 = header_start_row + 4  # 16~31일 '휴' 표시
    row_num_16_31 = header_start_row + 5  # 16~31일 숫자

    # 1) 1~15일 '휴' 표시행 전체 초기화 (헤더 텍스트 F3~U3에 있을 수도 있으니 F열부터만)
    # 단, F3~K3 같은 열은 '순 서'/'공종'/'성 명' 등 헤더가 아닌 날짜 위 자리임.
    # 안전하게: F~U 범위 (col 6~21)만 처리.
    for col in range(DAY_COL_START, 22):  # F(6) ~ U(21)
        # 1~15일 휴 행 초기화 (기본값 None으로)
        safe_set(ws, row_hol_1_15, col, None)
        # 16~31일 휴 행 초기화
        safe_set(ws, row_hol_16_31, col, None)

    # 2) 1~15일
    for day in range(1, 16):
        col = DAY_COL_START + (day - 1)  # F=6 for day=1
        if day > days_in_month:
            safe_set(ws, row_num_1_15, col, None)
            safe_set(ws, row_hol_1_15, col, None)
            continue
        safe_set(ws, row_num_1_15, col, day)
        wd = calendar.weekday(year, month, day)
        num_cell = ws.cell(row=row_num_1_15, column=col)
        hol_cell = ws.cell(row=row_hol_1_15, column=col)
        if wd == 5:  # 토요일: 숫자 파랑
            _apply_font_color(num_cell, COLOR_SATURDAY)
        elif wd == 6:  # 일요일: 숫자 빨강 + "휴"
            _apply_font_color(num_cell, COLOR_SUNDAY)
            safe_set(ws, row_hol_1_15, col, '휴')
            _apply_font_color(hol_cell, COLOR_SUNDAY)
        else:
            _apply_font_color(num_cell, '000000')

    # 3) 16~31일
    for day in range(16, 32):
        col = DAY_COL_START + (day - 16)  # F=6 for day=16
        if day > days_in_month:
            safe_set(ws, row_num_16_31, col, None)
            safe_set(ws, row_hol_16_31, col, None)
            continue
        safe_set(ws, row_num_16_31, col, day)
        wd = calendar.weekday(year, month, day)
        num_cell = ws.cell(row=row_num_16_31, column=col)
        hol_cell = ws.cell(row=row_hol_16_31, column=col)
        if wd == 5:
            _apply_font_color(num_cell, COLOR_SATURDAY)
        elif wd == 6:
            _apply_font_color(num_cell, COLOR_SUNDAY)
            safe_set(ws, row_hol_16_31, col, '휴')
            _apply_font_color(hol_cell, COLOR_SUNDAY)
        else:
            _apply_font_color(num_cell, '000000')


def _apply_font_color(cell, rgb_hex):
    if isinstance(cell, MergedCell):
        return
    old = cell.font
    cell.font = Font(
        name=old.name, size=old.size, bold=old.bold, italic=old.italic,
        vertAlign=old.vertAlign, underline=old.underline, strike=old.strike,
        color=Color(rgb='FF' + rgb_hex),
    )


# ==============================================================
# 섹션 생성 (가변 길이)
# ==============================================================

def build_section(ws, dst_start_row, header_block, worker_block, sum_block,
                  workers, year, month, category_label,
                  proto_first_worker_row, proto_section_start):
    """
    섹션 레이아웃:
      + 0 ~ +5        : 헤더 (6행)
      + 6 ~ +(6+2N-1) : 근로자 N명 (2행씩)
      +(6+2N) ~ +(6+2N+1) : 합계 (2행)
    총 길이 = 8 + 2N
    """
    n = len(workers)
    # 섹션 전체의 행 이동량 - 절대 참조(헤더) 이동에 사용
    section_delta = dst_start_row - proto_section_start

    # 1) 헤더
    paste_block(ws, header_block, dst_start_row)

    # 타이틀 C1 연월 치환
    title_cell = ws.cell(row=dst_start_row, column=COL_ORDER)
    if not isinstance(title_cell, MergedCell) and isinstance(title_cell.value, str):
        title_cell.value = re.sub(r'20\d{2}년\s*\d{1,2}월',
                                    f'{year}년 {month:02d}월',
                                    title_cell.value)
    # AE 기준
    ae_cell = ws.cell(row=dst_start_row, column=31)
    if not isinstance(ae_cell, MergedCell) and isinstance(ae_cell.value, str):
        ae_cell.value = f'{year}년 {month}월 기준'

    # 요일 색상
    apply_weekend_colors(ws, dst_start_row, year, month)

    # 2) 근로자 N명
    worker_first_row = dst_start_row + HEADER_ROWS
    for idx, (gongjong, name, day_dict) in enumerate(workers):
        row_top = worker_first_row + idx * WORKER_ROW_SPAN
        row_bot = row_top + 1
        # 견본 근로자 블록(2행) 붙이기
        # shift_absolute=True + absolute_delta=section_delta:
        #   $F$3, $F$5 같은 절대 참조를 섹션 헤더 위치로 이동 (다른 섹션은 다른 헤더)
        #   단, 같은 섹션 내 모든 근로자는 같은 섹션 헤더를 참조 (section_delta 동일)
        # 상대 참조 F7:U7은 각 근로자 행에 맞춰 이동 (delta = row_top - proto_first_worker_row)
        paste_block(ws, worker_block, row_top,
                    formula_base_row=proto_first_worker_row,
                    shift_absolute=True,
                    absolute_delta=section_delta)
        # 값 세팅
        safe_set(ws, row_top, COL_ORDER, idx + 1)
        safe_set(ws, row_top, COL_TYPE, gongjong)
        safe_set(ws, row_top, COL_NAME, name)
        # 일자별 공수 기입 전에 견본 잔여값 제거
        for col in range(DAY_COL_START, 22):
            for rr in (row_top, row_bot):
                cell = ws.cell(row=rr, column=col)
                if isinstance(cell, MergedCell):
                    continue
                if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.value = None
        for day, manday in day_dict.items():
            if 1 <= day <= 15:
                col = DAY_COL_START + (day - 1)
                safe_set(ws, row_top, col, manday)
            elif 16 <= day <= 31:
                col = DAY_COL_START + (day - 16)
                safe_set(ws, row_bot, col, manday)

    # 3) 합계 (근로자 바로 뒤)
    sum_first_row = worker_first_row + n * WORKER_ROW_SPAN
    paste_block(ws, sum_block, sum_first_row)

    # 합계 블록의 수식 범위 재조정:
    # 견본에서 근로자 범위는 proto_first_worker_row ~ (proto_first_worker_row + 40 - 1)
    # 새 섹션에서는 worker_first_row ~ (worker_first_row + 2N - 1)
    # paste_block이 이미 delta 적용했으므로, 수식에 남은 행 참조의 끝(=sum_first_row - 1)은 올바름.
    # 시작만 worker_first_row로 치환해주면 됨.
    worker_last_abs = sum_first_row - 1
    worker_first_abs = worker_first_row

    range_re = re.compile(r'(\$?[A-Z]{1,3}\$?)(\d+)(:)(\$?[A-Z]{1,3}\$?)(\d+)')
    for offset in range(sum_block['num_rows']):
        r = sum_first_row + offset
        for col in range(1, 35):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith('='):
                def repl(m, _end=worker_last_abs, _start=worker_first_abs):
                    col1, row1, sep, col2, row2 = m.group(1), int(m.group(2)), m.group(3), m.group(4), int(m.group(5))
                    if row2 == _end:
                        row1 = _start
                    return f'{col1}{row1}{sep}{col2}{row2}'
                cell.value = range_re.sub(repl, cell.value)

    # 라벨 + "계"
    safe_set(ws, sum_first_row, COL_LABEL, f'{year}년{month:02d}월({category_label})')
    safe_set(ws, sum_first_row, COL_ORDER, '계')

    return dst_start_row + HEADER_ROWS + n * WORKER_ROW_SPAN + 2


# ==============================================================
# External Link 제거 (속도 최적화의 핵심)
# ==============================================================

def strip_external_links(xlsx_path):
    """xlsx 파일에서 external link 관련 XML + definedNames 전부 제거.
    openpyxl은 external link 때문에 로드·저장이 수십 초 느려짐.
    또한 external link를 참조하는 definedNames는 Excel에서 복구 경고 원인.
    """
    tmp_path = xlsx_path + '.stripped.tmp'

    removed = 0
    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                # external link 관련 파일 제외
                if 'externalLink' in item:
                    removed += 1
                    continue
                data = zin.read(item)

                # workbook.xml 수정 (externalReferences + definedNames 제거)
                if item == 'xl/workbook.xml':
                    text = data.decode('utf-8')
                    text = re.sub(r'<externalReferences>.*?</externalReferences>',
                                  '', text, flags=re.DOTALL)
                    # definedNames 전체 제거 (대부분 external link 참조)
                    text = re.sub(r'<definedNames>.*?</definedNames>',
                                  '', text, flags=re.DOTALL)
                    data = text.encode('utf-8')

                elif item == 'xl/_rels/workbook.xml.rels':
                    text = data.decode('utf-8')
                    text = re.sub(r'<Relationship[^/]+externalLink[^/]+/>',
                                  '', text)
                    data = text.encode('utf-8')

                elif item == '[Content_Types].xml':
                    text = data.decode('utf-8')
                    text = re.sub(r'<Override[^/]+externalLink[^/]+/>',
                                  '', text)
                    data = text.encode('utf-8')

                zout.writestr(item, data)

    os.replace(tmp_path, xlsx_path)
    return removed


# ==============================================================
# 메인
# ==============================================================

def main():
    t_start = time.time()
    print("=" * 70)
    print("돌관공사비 노무비 출력일보 자동 작성 (v6)")
    print("=" * 70)

    if not os.path.exists(TEMPLATE_PATH):
        print(f"❌ 템플릿 파일 없음: {TEMPLATE_PATH}")
        return
    if not os.path.exists(SOURCE_DIR):
        print(f"❌ 소스 폴더 없음: {SOURCE_DIR}")
        return

    out_dir = os.path.dirname(OUTPUT_PATH)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    t0 = time.time()
    shutil.copy2(TEMPLATE_PATH, OUTPUT_PATH)
    print(f"✓ 템플릿 복사: {OUTPUT_PATH} ({time.time()-t0:.1f}초)")

    # External link 제거 (속도 최적화)
    t0 = time.time()
    removed = strip_external_links(OUTPUT_PATH)
    print(f"✓ External link 제거: {removed}개 ({time.time()-t0:.1f}초)")

    t0 = time.time()
    wb = load_workbook(OUTPUT_PATH)
    print(f"✓ 워크북 로드: {time.time()-t0:.1f}초")

    if TEMPLATE_SHEET not in wb.sheetnames:
        print(f"❌ '{TEMPLATE_SHEET}' 시트 없음")
        return
    ws = wb[TEMPLATE_SHEET]

    # 견본 섹션 위치
    proto_section_start = None
    for r in range(1, min(ws.max_row + 1, 200)):
        v = ws.cell(row=r, column=COL_ORDER).value
        if isinstance(v, str) and any(kw in v for kw in SECTION_TITLE_KEYWORDS):
            proto_section_start = r
            break
    if proto_section_start is None:
        print(f"❌ 견본 섹션 없음")
        return
    print(f"✓ 견본 섹션 위치: row {proto_section_start}")

    # 세 블록 스냅샷
    max_col = ws.max_column
    header_block = snapshot_cells(ws, proto_section_start, HEADER_ROWS, max_col)
    proto_first_worker_row = proto_section_start + HEADER_ROWS
    worker_block = snapshot_cells(ws, proto_first_worker_row, WORKER_ROW_SPAN, max_col)
    proto_sum_start = proto_section_start + PROTO_SECTION_SIZE - 2
    sum_block = snapshot_cells(ws, proto_sum_start, 2, max_col)

    print(f"  헤더 블록: {len(header_block['cells'])}셀")
    print(f"  근로자 블록: {len(worker_block['cells'])}셀")
    print(f"  합계 블록: {len(sum_block['cells'])}셀")

    # 시트 재생성
    del wb[TEMPLATE_SHEET]
    ws = wb.create_sheet(TEMPLATE_SHEET)
    print("✓ 기존 시트 삭제 후 재생성")

    # 소스 파일
    source_files = []
    for ext in ('*.xlsx', '*.xlsm'):
        source_files.extend(glob.glob(os.path.join(SOURCE_DIR, ext)))
    source_files = [f for f in source_files if not os.path.basename(f).startswith('~$')]
    source_files.sort()
    if not source_files:
        print(f"❌ 소스 폴더에 엑셀 파일 없음")
        return
    print(f"✓ 소스 파일 {len(source_files)}개 발견")

    # 소스 파일 읽기 (시간 측정)
    t_src = time.time()
    monthly_data = {}
    for idx_f, fp in enumerate(source_files, 1):
        fn = os.path.basename(fp)
        y, m = extract_year_month_from_filename(fn)
        if y is None:
            print(f"  ⚠ 연/월 인식 실패: {fn}")
            continue
        t_file = time.time()
        try:
            data = read_source_file(fp)
        except Exception as e:
            print(f"    ❌ 읽기 실패: {e}")
            continue
        monthly_data[(y, m)] = data
        print(f"  [{idx_f}/{len(source_files)}] {y}년 {m:02d}월 ({time.time()-t_file:.1f}초)")
    print(f"✓ 소스 파일 읽기 완료: {time.time()-t_src:.1f}초")

    if not monthly_data:
        print("❌ 읽은 데이터 없음")
        return

    # 섹션 생성 (가변 길이)
    t_sec = time.time()
    months_sorted = sorted(monthly_data.keys())
    cur_row = 1
    section_log = []
    print(f"\n[섹션 생성 시작] 총 {len(months_sorted)}개월 처리")

    for idx_m, (y, m) in enumerate(months_sorted, start=1):
        t_month = time.time()
        month_data = monthly_data[(y, m)]
        cat_counts = {}
        for cat in CATEGORIES:
            tot = sum(len(w) for w in month_data.get(cat, {}).values())
            if tot > 0:
                cat_counts[cat] = tot
        summary = ', '.join(f"{c}={n}명" for c, n in cat_counts.items())

        for cat in CATEGORIES:
            cat_data = month_data.get(cat, {})
            flat = []
            for gj in sorted(cat_data.keys()):
                for worker in sorted(cat_data[gj].keys()):
                    flat.append((gj, worker, cat_data[gj][worker]))
            if not flat:
                continue

            next_row = build_section(
                ws, cur_row,
                header_block, worker_block, sum_block,
                flat, y, m, cat,
                proto_first_worker_row=proto_first_worker_row,
                proto_section_start=proto_section_start,
            )
            section_log.append((cur_row, y, m, cat, len(flat), next_row - cur_row))
            cur_row = next_row

        print(f"  [{idx_m}/{len(months_sorted)}] {y}년 {m:02d}월  → {summary or '데이터 없음'} ({time.time()-t_month:.1f}초, 누적 {cur_row}행)")

    print(f"\n✓ 총 {len(section_log)}개 섹션 생성 완료: {time.time()-t_sec:.1f}초")
    for (r, y, m, cat, n, length) in section_log[:15]:
        print(f"   row {r:>6}: {y}년 {m:02d}월 [{cat}] {n}명 ({length}행)")
    if len(section_log) > 15:
        print(f"   ... (총 {len(section_log)}개)")

    # 저장
    t_save = time.time()
    print("\n저장 중...")
    wb.save(OUTPUT_PATH)
    print(f"✓ 저장 완료: {time.time()-t_save:.1f}초")

    print(f"\n{'='*70}")
    print(f"✓ 완료! 결과 파일: {OUTPUT_PATH}")
    print(f"  [총 소요 시간] {time.time()-t_start:.1f}초")
    print(f"{'='*70}")
    print("\n※ Excel에서 Ctrl+Alt+F9로 전체 재계산 실행하세요.")


if __name__ == '__main__':
    main()
# -*- coding: utf-8 -*-
"""
돌관공사비 산출근거 - 노무비 출력일보 자동 작성 스크립트 (v11)

[v11 주요 변경 사항 - PDF 인쇄 설정 완전 복원]
1) 인쇄 설정 전부 sample에서 복원:
   - 용지: A4 (paperSize=9)
   - 방향: 가로(landscape)
   - 배율: 79% (fitToPage=True, fitToWidth=1, fitToHeight=0)
   - 여백: 좌우 1.8cm, 상하 1.9cm, 헤더/푸터 0.8cm
   - 가로 중앙 정렬
2) 인쇄 영역(print_area) 자동 설정: C1:AF{마지막행}
3) 각 페이지 경계마다 수동 페이지 나눔(row_breaks) 추가
   → PDF 출력 시 각 페이지가 정확히 48행 단위로 끊어짐

[v10 주요 변경]
1) 템플릿의 열 너비 유지
2) 마지막 페이지 20명 + 합계로 변경 (PDF 한 페이지 맞춤)

[v8 주요 변경]
1) 합계행 E열 '명' 표시 - 실제 근로자 수 (예: "25명")
2) 한국 공휴일 자동 '휴' 표시 (대체공휴일, 선거일, 임시공휴일 포함)
3) 파일 분할 저장 (연도별)

[사용 방법]
1. 아래 '사용자 설정'의 경로를 본인 환경에 맞게 수정
2. 터미널에서:
       pip install openpyxl holidays
       python mandays_report_automation_v11.py
"""

import os
import re
import glob
import shutil
import calendar
import time
import zipfile
import io
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Color

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))
except ImportError:
    pass


# ==============================================================
# 사용자 설정  (.env 또는 아래 기본값)
# ==============================================================
_BASE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.environ.get('CRASH_TEMPLATE_PATH',
    os.path.join(_BASE, 'sample_돌관공사비 산출근거.xlsx'))
SOURCE_DIR    = os.environ.get('CRASH_SOURCE_DIR',
    os.path.join(_BASE, 'source'))
OUTPUT_PATH   = os.environ.get('CRASH_OUTPUT_PATH',
    os.path.join(_BASE, 'output', '돌관공사비_산출근거_자동생성_v11.xlsx'))


# ==============================================================
# 소스 파일 구조 설정
# ==============================================================
SRC_SHEET_PATTERN = re.compile(r'^(0?\d+)\.\s*(.+)$')


# ==============================================================
# 템플릿 구조 설정
# ==============================================================
TEMPLATE_SHEET = '노무비 출력일보'
HEADER_ROWS = 6            # 섹션 헤더 높이
WORKER_ROW_SPAN = 2        # 근로자 1명당 행 수
PROTO_SECTION_SIZE = 48    # 견본 섹션 원본 크기 (20명용)
PROTO_WORKER_COUNT = 20

# 페이지 분할 (v7 신규)
WORKERS_PER_PAGE = 21      # 일반 페이지당 최대 근로자 수
LAST_PAGE_WORKERS_MAX = 20 # v10: 마지막 페이지는 20명 + 합계 2행 = 48행 (PDF 한 페이지)
PAGE_ROWS = HEADER_ROWS + WORKERS_PER_PAGE * WORKER_ROW_SPAN  # 48 (일반 페이지 높이)
LAST_PAGE_ROWS = HEADER_ROWS + LAST_PAGE_WORKERS_MAX * WORKER_ROW_SPAN + 2  # 48 (20명 + 합계 2)

COL_LABEL = 2   # B: 카테고리 라벨
COL_ORDER = 3   # C: 순서/"계"
COL_TYPE  = 4   # D: 공종
COL_NAME  = 5   # E: 성명
DAY_COL_START = 6   # F: 1일 or 16일
COL_UNIT_PRICE = 28  # AB: 노무비 단가 (v11)
COL_TOTAL = 29       # AC: 노무비 총액 (v11)

CATEGORIES = ['본선', '복합', '삼성', '기타']
SECTION_TITLE_KEYWORDS = ['노무비 명세서', '노임내역']

COLOR_SATURDAY = '0000FF'
COLOR_SUNDAY   = 'FF0000'


# ==============================================================
# 유틸
# ==============================================================

def safe_set(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True


def extract_year_month_from_filename(filename):
    name = os.path.splitext(os.path.basename(filename))[0]
    for pattern in [
        r'(20\d{2})\s*년\s*(\d{1,2})\s*월',
        r'(20\d{2})[.\-_ ](\d{1,2})(?!\d)',
        r'(20\d{2})(\d{2})',
        r'(?<!\d)(2[3-9])[.\-_ ](\d{1,2})(?!\d)',
        r'(?<!\d)(2[3-9])\s*년\s*(\d{1,2})\s*월',
    ]:
        m = re.search(pattern, name)
        if m:
            y_raw, mo = int(m.group(1)), int(m.group(2))
            y = y_raw if y_raw >= 100 else 2000 + y_raw
            if 1 <= mo <= 12:
                return y, mo
    return None, None


def classify_category(f_val, h_val):
    combined = ''
    if f_val is not None:
        combined += str(f_val)
    if h_val is not None:
        combined += ' ' + str(h_val)
    if '본선' in combined:
        return '본선'
    if '복합' in combined:
        return '복합'
    if '삼성' in combined:
        return '삼성'
    return '기타'


# ==============================================================
# 소스 파일 읽기
# ==============================================================

def read_source_file(filepath):
    """소스 파일을 읽어 근로자별 데이터를 반환.

    반환 구조:
      result[category][gongjong][worker_name] = {
          'days': {day: manday, ...},   # 날짜별 공수 (기존 구조)
          'unit_price': float or None,  # B열 '일 기준' 아래 값 (AB열용)
          'total': float or None,       # E열 근로자 블록 마지막 합계 (AC열용)
      }
    """
    wb = load_workbook(filepath, data_only=True, read_only=False)  # read_only=False: max_row 안정
    result = {cat: {} for cat in CATEGORIES}

    try:
        for sheet_name in wb.sheetnames:
            m = SRC_SHEET_PATTERN.match(sheet_name.strip())
            if not m:
                continue
            gongjong = m.group(2).strip()
            ws = wb[sheet_name]

            # 모든 셀을 한 번에 메모리로 로드 (B, C, D, E, F, H열)
            # 시트 끝까지 순회하면서 근로자 블록을 찾음
            rows_data = []
            for r in range(1, ws.max_row + 1):
                rows_data.append({
                    'b': ws.cell(r, 2).value,
                    'c': ws.cell(r, 3).value,
                    'd': ws.cell(r, 4).value,
                    'e': ws.cell(r, 5).value,
                    'f': ws.cell(r, 6).value,
                    'h': ws.cell(r, 8).value,
                })

            # 블록 경계 찾기: B열에 '작업자명' 있는 행
            worker_header_idxs = [
                i for i, row in enumerate(rows_data)
                if isinstance(row['b'], str) and row['b'].strip() == '작업자명'
            ]

            for bi, hdr_idx in enumerate(worker_header_idxs):
                name_idx = hdr_idx + 1
                if name_idx >= len(rows_data):
                    continue
                worker_name = rows_data[name_idx]['b']
                if worker_name is None or not str(worker_name).strip():
                    continue
                worker_name = str(worker_name).strip()
                if worker_name == '작업자':  # dummy entry
                    continue

                # 블록 끝: 다음 '작업자명' 직전 (또는 시트 끝)
                if bi + 1 < len(worker_header_idxs):
                    block_end_idx = worker_header_idxs[bi + 1]
                else:
                    block_end_idx = len(rows_data)

                # 단가 추출: B열 '일 기준' 아래 행의 B열 값
                unit_price = None
                for j in range(hdr_idx, block_end_idx):
                    b = rows_data[j]['b']
                    if isinstance(b, str) and '일 기준' in b:
                        if j + 1 < block_end_idx:
                            up = rows_data[j + 1]['b']
                            if isinstance(up, (int, float)):
                                unit_price = float(up)
                            else:
                                try:
                                    unit_price = float(up) if up else None
                                except (TypeError, ValueError):
                                    unit_price = None
                        break

                # 총액(전체) - v8까지 사용: 블록 마지막 E열 값 (참고용으로 남겨둠)
                total_all = None
                for j in range(block_end_idx - 1, hdr_idx, -1):
                    e = rows_data[j]['e']
                    if isinstance(e, (int, float)) and e != 0:
                        total_all = float(e)
                        break

                # v9: 카테고리별로 일자별 공수 + 카테고리별 E열 합계 수집
                # 한 근로자가 날짜마다 다른 카테고리에 투입될 수 있으므로,
                # 날짜별로 F/H열을 보고 카테고리를 분류하고,
                # 같은 행의 E열 값(=그 날짜의 노무비)을 해당 카테고리 총액에 더한다.
                daily_by_cat = {cat: {} for cat in CATEGORIES}
                total_by_cat = {cat: 0.0 for cat in CATEGORIES}
                for j in range(name_idx, block_end_idx):
                    row = rows_data[j]
                    c_val = row['c']
                    d_val = row['d']
                    e_val = row['e']
                    f_val = row['f']
                    h_val = row['h']
                    if d_val is None:
                        continue
                    try:
                        day = int(c_val)
                        manday = float(d_val)
                    except (TypeError, ValueError):
                        continue
                    if not (1 <= day <= 31):
                        continue
                    cat = classify_category(f_val, h_val)
                    daily_by_cat[cat][day] = daily_by_cat[cat].get(day, 0) + manday
                    # v9: 그 날의 E열 값을 카테고리 총액에 더함
                    if isinstance(e_val, (int, float)):
                        total_by_cat[cat] += float(e_val)

                # 카테고리별로 저장 (그 카테고리에 투입된 날이 있을 때만)
                for cat, days in daily_by_cat.items():
                    if not days:
                        continue
                    # 같은 근로자가 여러 공종에 있을 수 있으므로 setdefault
                    gongjong_bucket = result[cat].setdefault(gongjong, {})
                    if worker_name in gongjong_bucket:
                        # 동일 공종·동일 근로자의 날짜/총액 합산
                        existing = gongjong_bucket[worker_name]
                        for day, m in days.items():
                            existing['days'][day] = existing['days'].get(day, 0) + m
                        if existing.get('total') is not None and total_by_cat[cat]:
                            existing['total'] += total_by_cat[cat]
                        elif total_by_cat[cat]:
                            existing['total'] = total_by_cat[cat]
                    else:
                        gongjong_bucket[worker_name] = {
                            'days': dict(days),
                            'unit_price': unit_price,
                            'total': total_by_cat[cat] if total_by_cat[cat] else None,
                        }
    finally:
        wb.close()
    return result


# ==============================================================
# 템플릿 블록 스냅샷
# ==============================================================

def snapshot_cells(ws, start_row, num_rows, max_col):
    cells = []
    for offset in range(num_rows):
        r = start_row + offset
        for col in range(1, max_col + 1):
            cell = ws.cell(row=r, column=col)
            style_snap = None
            if cell.has_style:
                style_snap = {
                    'font': copy(cell.font),
                    'border': copy(cell.border),
                    'fill': copy(cell.fill),
                    'number_format': cell.number_format,
                    'alignment': copy(cell.alignment),
                }
            val = cell.value
            # ArrayFormula는 특별한 마커로 저장 (paste 시 다시 ArrayFormula로 만들기 위해)
            # Excel에서 SUMPRODUCT+IF 같은 수식은 ArrayFormula여야 제대로 계산됨
            is_array = False
            if val is not None and type(val).__name__ == 'ArrayFormula':
                try:
                    t = val.text
                    val = t if t.startswith('=') else '=' + t
                    is_array = True
                except Exception:
                    val = None
            if val is not None or style_snap is not None:
                cells.append((offset, col, val, style_snap, is_array))

    merges = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= start_row and mr.max_row < start_row + num_rows:
            merges.append((mr.min_row - start_row, mr.max_row - start_row,
                           mr.min_col, mr.max_col))

    row_heights = {}
    for offset in range(num_rows):
        r = start_row + offset
        if r in ws.row_dimensions and ws.row_dimensions[r].height is not None:
            row_heights[offset] = ws.row_dimensions[r].height

    return {
        'cells': cells, 'merges': merges, 'row_heights': row_heights,
        'orig_start': start_row, 'num_rows': num_rows,
    }


_formula_row_re = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')

def shift_formula_rows(formula, delta, shift_absolute=False, absolute_delta=None):
    """수식의 행 번호를 이동.
    - 상대 참조(예: F7)는 delta만큼 이동
    - 절대 참조(예: $F$3):
        shift_absolute=False (기본): 그대로 유지
        shift_absolute=True: absolute_delta만큼 이동 (None이면 delta 사용)
    """
    if shift_absolute and absolute_delta is None:
        absolute_delta = delta
    def repl(m):
        col_abs, col, row_abs, row_num = m.group(1), m.group(2), m.group(3), m.group(4)
        if row_abs == '$':
            if shift_absolute:
                return f'{col_abs}{col}{row_abs}{int(row_num) + absolute_delta}'
            return m.group(0)
        return f'{col_abs}{col}{row_abs}{int(row_num) + delta}'
    return _formula_row_re.sub(repl, formula)


# ==============================================================
# v10: 열 너비 백업/복원
# ==============================================================

def snapshot_column_dimensions(ws):
    """시트의 column_dimensions 정보를 백업.
    각 항목: (key, width, min, max, hidden, customWidth, bestFit, outlineLevel)
    """
    dims = []
    for key, d in ws.column_dimensions.items():
        dims.append({
            'key': key,
            'width': d.width,
            'min': d.min,
            'max': d.max,
            'hidden': d.hidden,
            'customWidth': d.customWidth,
            'bestFit': d.bestFit,
            'outlineLevel': d.outlineLevel,
        })
    return dims


def restore_column_dimensions(ws, dims):
    """snapshot한 column_dimensions를 새 시트에 복원."""
    from openpyxl.worksheet.dimensions import ColumnDimension
    for d in dims:
        cd = ColumnDimension(
            ws,
            min=d['min'],
            max=d['max'],
            width=d['width'],
            hidden=d['hidden'] or False,
            customWidth=d['customWidth'] or False,
            bestFit=d['bestFit'] or False,
            outlineLevel=d['outlineLevel'] or 0,
        )
        ws.column_dimensions[d['key']] = cd


def snapshot_row_default(ws):
    """기본 행 높이 정보 백업."""
    if ws.sheet_format is not None:
        return {
            'defaultRowHeight': ws.sheet_format.defaultRowHeight,
            'defaultColWidth': ws.sheet_format.defaultColWidth,
        }
    return {}


def restore_row_default(ws, info):
    """기본 행 높이 복원."""
    if info and ws.sheet_format is not None:
        if info.get('defaultRowHeight') is not None:
            ws.sheet_format.defaultRowHeight = info['defaultRowHeight']
        if info.get('defaultColWidth') is not None:
            ws.sheet_format.defaultColWidth = info['defaultColWidth']


# ==============================================================
# v11: 인쇄 설정 백업/복원
# ==============================================================

def snapshot_print_settings(ws):
    """인쇄 관련 설정을 snapshot."""
    ps = ws.page_setup
    pm = ws.page_margins
    po = ws.print_options
    snap = {
        'orientation': ps.orientation,
        'paperSize': ps.paperSize,
        'paperHeight': ps.paperHeight,
        'paperWidth': ps.paperWidth,
        'fitToWidth': ps.fitToWidth,
        'fitToHeight': ps.fitToHeight,
        'scale': ps.scale,
        'margin_left': pm.left,
        'margin_right': pm.right,
        'margin_top': pm.top,
        'margin_bottom': pm.bottom,
        'margin_header': pm.header,
        'margin_footer': pm.footer,
        'horizontalCentered': po.horizontalCentered,
        'verticalCentered': po.verticalCentered,
        'fitToPage': (ws.sheet_properties.pageSetUpPr.fitToPage
                      if ws.sheet_properties.pageSetUpPr else None),
    }
    return snap


def restore_print_settings(ws, snap):
    """인쇄 설정 복원."""
    if not snap:
        return
    ps = ws.page_setup
    pm = ws.page_margins
    po = ws.print_options
    if snap.get('orientation') is not None:
        ps.orientation = snap['orientation']
    if snap.get('paperSize') is not None:
        ps.paperSize = snap['paperSize']
    if snap.get('paperHeight') is not None:
        ps.paperHeight = snap['paperHeight']
    if snap.get('paperWidth') is not None:
        ps.paperWidth = snap['paperWidth']
    if snap.get('fitToWidth') is not None:
        ps.fitToWidth = snap['fitToWidth']
    if snap.get('fitToHeight') is not None:
        ps.fitToHeight = snap['fitToHeight']
    if snap.get('scale') is not None:
        ps.scale = snap['scale']
    # margins
    if snap.get('margin_left') is not None:
        pm.left = snap['margin_left']
    if snap.get('margin_right') is not None:
        pm.right = snap['margin_right']
    if snap.get('margin_top') is not None:
        pm.top = snap['margin_top']
    if snap.get('margin_bottom') is not None:
        pm.bottom = snap['margin_bottom']
    if snap.get('margin_header') is not None:
        pm.header = snap['margin_header']
    if snap.get('margin_footer') is not None:
        pm.footer = snap['margin_footer']
    # print options
    if snap.get('horizontalCentered') is not None:
        po.horizontalCentered = snap['horizontalCentered']
    if snap.get('verticalCentered') is not None:
        po.verticalCentered = snap['verticalCentered']
    # fitToPage
    if snap.get('fitToPage') is not None and ws.sheet_properties.pageSetUpPr is not None:
        ws.sheet_properties.pageSetUpPr.fitToPage = snap['fitToPage']


def set_print_area_and_breaks(ws, last_row, page_boundary_rows):
    """인쇄 영역 설정 + 각 페이지 경계에 수동 페이지 나눔 추가.

    ws: 대상 시트
    last_row: 마지막 행 (인쇄 영역 끝)
    page_boundary_rows: 각 페이지 '다음 페이지 시작 행' 리스트
                       예: 48, 96, 144 (48행 단위 페이지)
    """
    # 인쇄 영역: C1:AF{last_row}
    ws.print_area = f'C1:AF{last_row}'

    # 기존 row_breaks 초기화
    from openpyxl.worksheet.pagebreak import Break
    ws.row_breaks.brk = []

    # 각 경계 행 이전에서 나눔 추가 (id = 나눔 직전 행)
    for boundary in page_boundary_rows:
        # boundary = 다음 페이지의 시작 행
        # Excel에서 row break는 "이 행의 아래에서 끊는다"는 의미
        # boundary가 새 페이지 시작이면 boundary-1 뒤에서 끊어야 함
        br = Break(id=boundary - 1, min=2, max=31, man=True)
        ws.row_breaks.append(br)


def paste_block(ws, block, dst_start_row, formula_base_row=None, shift_absolute=False, absolute_delta=None):
    if formula_base_row is None:
        formula_base_row = block['orig_start']
    delta = dst_start_row - formula_base_row

    for item in block['cells']:
        if len(item) == 5:
            offset, col, value, style_snap, is_array = item
        else:
            offset, col, value, style_snap = item
            is_array = False
        r = dst_start_row + offset
        cell = ws.cell(row=r, column=col)
        if isinstance(cell, MergedCell):
            continue
        if isinstance(value, str) and value.startswith('='):
            shifted = shift_formula_rows(value, delta,
                                          shift_absolute=shift_absolute,
                                          absolute_delta=absolute_delta)
            if is_array:
                from openpyxl.worksheet.formula import ArrayFormula
                from openpyxl.utils import get_column_letter
                ref = f'{get_column_letter(col)}{r}'
                cell.value = ArrayFormula(ref, shifted)
            else:
                cell.value = shifted
        else:
            cell.value = value
        if style_snap is not None:
            cell.font = style_snap['font']
            cell.border = style_snap['border']
            cell.fill = style_snap['fill']
            cell.number_format = style_snap['number_format']
            cell.alignment = style_snap['alignment']

    for offset, height in block['row_heights'].items():
        ws.row_dimensions[dst_start_row + offset].height = height

    # 병합: ws.merge_cells() 대신 직접 CellRange 추가 (더 빠름)
    from openpyxl.worksheet.cell_range import CellRange
    for (min_off, max_off, min_col, max_col) in block['merges']:
        try:
            rng = CellRange(
                min_col=min_col, min_row=dst_start_row + min_off,
                max_col=max_col, max_row=dst_start_row + max_off
            )
            ws.merged_cells.ranges.add(rng)
        except Exception:
            pass


# ==============================================================
# 요일/휴일 처리
# ==============================================================

def apply_weekend_colors(ws, header_start_row, year, month):
    """섹션 헤더의 일자 셀에 요일 색상 적용.
    헤더 레이아웃 (header_start_row 기준 0-based):
      offset 0: 타이틀
      offset 1: 구분
      offset 2: 컬럼명
      offset 3: 1~15일의 '휴' 표시행 (각 열이 해당 일자 위)
      offset 4: 1~15일 숫자
      offset 5: 16~31일의 '휴' 표시행
      offset 6: 16~31일 숫자  ← 실제로는 offset 5가 위쪽 "휴", offset 6이 숫자가 아님

    실제 템플릿 구조:
      offset 2 (row 3): 1~15일 '휴' 행 (숫자 위)
      offset 3 (row 4): 1~15일 숫자
      offset 4 (row 5): 16~31일 '휴' 행 (숫자 위)
      offset 5 (row 6): 16~31일 숫자
    """
    days_in_month = calendar.monthrange(year, month)[1]
    # v8: 공휴일 정보 (한국 공휴일 + 대체공휴일 + 선거일)
    holiday_days = _get_holidays_for_month(year, month)

    row_hol_1_15  = header_start_row + 2  # 1~15일 '휴' 표시 (숫자 위)
    row_num_1_15  = header_start_row + 3  # 1~15일 숫자
    row_hol_16_31 = header_start_row + 4  # 16~31일 '휴' 표시
    row_num_16_31 = header_start_row + 5  # 16~31일 숫자

    # 1) 1~15일 '휴' 표시행 전체 초기화 (헤더 텍스트 F3~U3에 있을 수도 있으니 F열부터만)
    # 단, F3~K3 같은 열은 '순 서'/'공종'/'성 명' 등 헤더가 아닌 날짜 위 자리임.
    # 안전하게: F~U 범위 (col 6~21)만 처리.
    for col in range(DAY_COL_START, 22):  # F(6) ~ U(21)
        # 1~15일 휴 행 초기화 (기본값 None으로)
        safe_set(ws, row_hol_1_15, col, None)
        # 16~31일 휴 행 초기화
        safe_set(ws, row_hol_16_31, col, None)

    # 2) 1~15일
    for day in range(1, 16):
        col = DAY_COL_START + (day - 1)  # F=6 for day=1
        if day > days_in_month:
            safe_set(ws, row_num_1_15, col, None)
            safe_set(ws, row_hol_1_15, col, None)
            continue
        safe_set(ws, row_num_1_15, col, day)
        wd = calendar.weekday(year, month, day)
        is_holiday = day in holiday_days
        num_cell = ws.cell(row=row_num_1_15, column=col)
        hol_cell = ws.cell(row=row_hol_1_15, column=col)
        if wd == 6 or is_holiday:  # 일요일 or 공휴일: 빨강 + "휴"
            _apply_font_color(num_cell, COLOR_SUNDAY)
            safe_set(ws, row_hol_1_15, col, '휴')
            _apply_font_color(hol_cell, COLOR_SUNDAY)
        elif wd == 5:  # 토요일: 파랑 (휴 표시 없음)
            _apply_font_color(num_cell, COLOR_SATURDAY)
        else:
            _apply_font_color(num_cell, '000000')

    # 3) 16~31일
    for day in range(16, 32):
        col = DAY_COL_START + (day - 16)  # F=6 for day=16
        if day > days_in_month:
            safe_set(ws, row_num_16_31, col, None)
            safe_set(ws, row_hol_16_31, col, None)
            continue
        safe_set(ws, row_num_16_31, col, day)
        wd = calendar.weekday(year, month, day)
        is_holiday = day in holiday_days
        num_cell = ws.cell(row=row_num_16_31, column=col)
        hol_cell = ws.cell(row=row_hol_16_31, column=col)
        if wd == 6 or is_holiday:
            _apply_font_color(num_cell, COLOR_SUNDAY)
            safe_set(ws, row_hol_16_31, col, '휴')
            _apply_font_color(hol_cell, COLOR_SUNDAY)
        elif wd == 5:
            _apply_font_color(num_cell, COLOR_SATURDAY)
        else:
            _apply_font_color(num_cell, '000000')


# v8: 한국 공휴일 조회 (holidays 라이브러리)
_holidays_cache = {}

def _get_holidays_for_month(year, month):
    """해당 년/월의 공휴일(일) 집합을 반환.
    대체공휴일, 선거일, 임시공휴일 모두 포함.
    holidays 라이브러리가 없으면 빈 set 반환 (일요일만 '휴' 표시).
    """
    key = (year, month)
    if key in _holidays_cache:
        return _holidays_cache[key]
    try:
        import holidays as _h
        kr = _h.KR(years=[year])
        result = set()
        for d in kr:
            if d.year == year and d.month == month:
                result.add(d.day)
        _holidays_cache[key] = result
        return result
    except Exception:
        _holidays_cache[key] = set()
        return set()


def _apply_font_color(cell, rgb_hex):
    if isinstance(cell, MergedCell):
        return
    old = cell.font
    cell.font = Font(
        name=old.name, size=old.size, bold=old.bold, italic=old.italic,
        vertAlign=old.vertAlign, underline=old.underline, strike=old.strike,
        color=Color(rgb='FF' + rgb_hex),
    )


# ==============================================================
# 섹션 생성 (가변 길이)
# ==============================================================

def build_section(ws, dst_start_row, header_block, worker_block, sum_block,
                  workers, year, month, category_label,
                  proto_first_worker_row, proto_section_start):
    """
    섹션을 페이지 분할 방식으로 생성.

    workers: [(gongjong, name, {'days': {...}, 'unit_price': ..., 'total': ...}), ...]

    레이아웃:
      페이지 1: 헤더(6행) + 근로자 21명(42행) [마지막 페이지가 아니면 합계 없음]
      페이지 2: 헤더(6행) + 근로자 21명 또는 남은 인원 + 빈 슬롯들 + 합계(2행, 마지막 페이지만)

    근로자가 M명일 때 필요한 페이지 수:
      pages = ceil(M / WORKERS_PER_PAGE)
      모든 페이지는 헤더 + 21명 슬롯 크기 고정 (48행)
      마지막 페이지는 그 아래에 합계 2행 추가

    합계는 "첫 페이지 근로자 첫 행"부터 "마지막 페이지 근로자 마지막 실제 행"까지.
    단, 빈 슬롯이 있는 페이지에서는 슬롯 범위 전체를 합계 범위로 잡아도 무방 (빈 행은 0).
    이 구현은 모든 페이지의 전체 슬롯 범위를 합계 범위로 사용 (계산상 동일).
    """
    M = len(workers)
    if M == 0:
        return dst_start_row
    # 섹션 전체 행 이동량 (절대 참조 헤더 이동용)
    section_delta = dst_start_row - proto_section_start

    # v10: 페이지 수 및 각 페이지의 근로자 수 계산
    # - 마지막 페이지는 최대 20명 + 합계 2행 = 48행 (PDF 한 페이지에 맞음)
    # - 중간 페이지들은 21명씩
    # - M=21 이하는 모두 한 페이지로 처리 (M=21일 때는 50행이 되어 PDF 예외이지만
    #   현실적으로 M=21을 1페이지에 두지 않으면 마지막 페이지가 0명이 되어 어색)
    import math
    if M <= 21:
        num_pages = 1
        workers_per_page_list = [M]  # 첫 페이지(=마지막 페이지)에 전부
    else:
        # 마지막 페이지 근로자 수 = M - front_pages * 21 (0 초과 20 이하 목표)
        front_pages = math.ceil((M - LAST_PAGE_WORKERS_MAX) / WORKERS_PER_PAGE)
        last_workers = M - front_pages * WORKERS_PER_PAGE
        # last_workers <= 0이면 (M이 21의 배수+20 형태) 마지막 페이지가 0명이 되는 경우
        # → front_pages를 1 줄이고, 마지막 페이지에 21명을 담되 50행 허용 (예외)
        if last_workers <= 0:
            front_pages -= 1
            last_workers = M - front_pages * WORKERS_PER_PAGE  # 21이 됨
        num_pages = front_pages + 1
        workers_per_page_list = [WORKERS_PER_PAGE] * front_pages + [last_workers]

    # 각 페이지의 첫 행: 중간 페이지는 PAGE_ROWS(48), 마지막 페이지도 48행 크기의 "슬롯 영역"을 가짐
    # (마지막 페이지 = 헤더 6 + 슬롯 40 + 합계 2 = 48, 또는 예외 시 50행)
    page_start_rows = []
    cur = dst_start_row
    for p in range(num_pages):
        page_start_rows.append(cur)
        # 페이지 높이: 중간 페이지 = 48, 마지막 페이지 = 헤더 6 + 슬롯 2*N + (있으면)2(합계)
        # 하지만 페이지 전체를 PAGE_ROWS로 일관되게 배치하면 단순해짐.
        # 마지막 페이지 슬롯 수: min(workers_per_page_list[p], LAST_PAGE_WORKERS_MAX) → 20 슬롯 고정
        #   예외(21명): 21 슬롯 사용
        if p < num_pages - 1:
            cur += PAGE_ROWS  # 중간 페이지: 48행
        else:
            # 마지막 페이지: 헤더 + 슬롯 영역 + 합계(2행)
            last_slots = max(LAST_PAGE_WORKERS_MAX, workers_per_page_list[p])  # 20 또는 21(예외)
            cur += HEADER_ROWS + last_slots * WORKER_ROW_SPAN + 2

    # 페이지별 헤더 + 근로자 슬롯 생성
    global_worker_idx = 0
    for p in range(num_pages):
        page_start = page_start_rows[p]
        # 각 페이지는 견본 섹션의 "헤더 위치"와 대응하므로
        # absolute_delta = page_start - proto_section_start
        page_delta = page_start - proto_section_start

        # 1) 헤더
        paste_block(ws, header_block, page_start)

        # 타이틀 C열 연월 치환
        title_cell = ws.cell(row=page_start, column=COL_ORDER)
        if not isinstance(title_cell, MergedCell) and isinstance(title_cell.value, str):
            title_cell.value = re.sub(r'20\d{2}년\s*\d{1,2}월',
                                        f'{year}년 {month:02d}월',
                                        title_cell.value)
        # AE 기준 문구
        ae_cell = ws.cell(row=page_start, column=31)
        if not isinstance(ae_cell, MergedCell) and isinstance(ae_cell.value, str):
            ae_cell.value = f'{year}년 {month}월 기준'

        # 요일 색상
        apply_weekend_colors(ws, page_start, year, month)

        # 2) 이 페이지의 근로자 슬롯
        # 중간 페이지: 21개 슬롯 (21명 모두 기입, 실근로자 없으면 빈 슬롯)
        # 마지막 페이지: 20개 슬롯 (또는 예외 시 21) - 나머지 근로자 기입 후 나머지 슬롯은 빈
        is_last_page = (p == num_pages - 1)
        if is_last_page:
            page_slots = max(LAST_PAGE_WORKERS_MAX, workers_per_page_list[p])
        else:
            page_slots = WORKERS_PER_PAGE

        worker_first_row_on_page = page_start + HEADER_ROWS
        for slot_idx in range(page_slots):
            row_top = worker_first_row_on_page + slot_idx * WORKER_ROW_SPAN
            row_bot = row_top + 1

            # 근로자 슬롯 블록 paste (절대 참조는 이 페이지 헤더로 이동)
            paste_block(ws, worker_block, row_top,
                        formula_base_row=proto_first_worker_row,
                        shift_absolute=True,
                        absolute_delta=page_delta)

            # 견본의 일자 데이터 제거 (F~U, 수식이 아닌 값만)
            for col in range(DAY_COL_START, 22):
                for rr in (row_top, row_bot):
                    cell = ws.cell(row=rr, column=col)
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith('=')):
                        cell.value = None
            # AB/AC/AD의 기본값(270000 등)도 제거
            for col in (COL_UNIT_PRICE, COL_TOTAL, 30):  # AB, AC, AD
                for rr in (row_top, row_bot):
                    cell = ws.cell(row=rr, column=col)
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith('=')):
                        cell.value = None
            # 견본의 C(순서) D(공종) E(성명) 값 제거 (빈 슬롯/실데이터 모두에 대해)
            # 실데이터 근로자가 있으면 뒤에서 다시 기입됨
            for col in (COL_ORDER, COL_TYPE, COL_NAME):
                for rr in (row_top, row_bot):
                    cell = ws.cell(row=rr, column=col)
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith('=')):
                        cell.value = None

            # 실제 근로자가 있으면 데이터 기입
            # (이 페이지 slot_idx < workers_per_page_list[p] 인 경우)
            if slot_idx < workers_per_page_list[p] and global_worker_idx < M:
                gongjong, name, info = workers[global_worker_idx]
                day_dict = info.get('days', {}) if isinstance(info, dict) else info
                unit_price = info.get('unit_price') if isinstance(info, dict) else None
                total = info.get('total') if isinstance(info, dict) else None

                safe_set(ws, row_top, COL_ORDER, global_worker_idx + 1)
                safe_set(ws, row_top, COL_TYPE, gongjong)
                safe_set(ws, row_top, COL_NAME, name)
                # 일자별 공수
                for day, manday in day_dict.items():
                    if 1 <= day <= 15:
                        col = DAY_COL_START + (day - 1)
                        safe_set(ws, row_top, col, manday)
                    elif 16 <= day <= 31:
                        col = DAY_COL_START + (day - 16)
                        safe_set(ws, row_bot, col, manday)
                # AB 단가, AC 총액 (v7 신규)
                if unit_price is not None:
                    safe_set(ws, row_top, COL_UNIT_PRICE, unit_price)
                if total is not None:
                    safe_set(ws, row_top, COL_TOTAL, total)
                global_worker_idx += 1

    # 3) 합계 (마지막 페이지 슬롯 끝 바로 다음, 2행)
    last_page_start = page_start_rows[-1]
    last_page_slots = max(LAST_PAGE_WORKERS_MAX, workers_per_page_list[-1])
    sum_first_row = last_page_start + HEADER_ROWS + last_page_slots * WORKER_ROW_SPAN

    paste_block(ws, sum_block, sum_first_row)

    # 합계 수식 범위 재조정:
    # 첫 페이지 근로자 첫 행 ~ 마지막 페이지 근로자 마지막 슬롯 행
    worker_first_abs = dst_start_row + HEADER_ROWS
    worker_last_abs = sum_first_row - 1

    range_re = re.compile(r'(\$?[A-Z]{1,3}\$?)(\d+)(:)(\$?[A-Z]{1,3}\$?)(\d+)')
    for offset in range(sum_block['num_rows']):
        r = sum_first_row + offset
        for col in range(1, 35):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith('='):
                def repl(m, _end=worker_last_abs, _start=worker_first_abs):
                    col1, row1, sep, col2, row2 = m.group(1), int(m.group(2)), m.group(3), m.group(4), int(m.group(5))
                    if row2 == _end:
                        row1 = _start
                    return f'{col1}{row1}{sep}{col2}{row2}'
                cell.value = range_re.sub(repl, cell.value)

    # 라벨 + "계" + "N명"
    safe_set(ws, sum_first_row, COL_LABEL, f'{year}년{month:02d}월({category_label})')
    safe_set(ws, sum_first_row, COL_ORDER, '계')
    safe_set(ws, sum_first_row, COL_NAME, f'{M}명')

    # 섹션 전체 행 수 = 합계 끝 행까지
    total_rows = (sum_first_row + 2) - dst_start_row
    # v11: 페이지 시작 행 리스트도 함께 반환 (페이지 나눔 설정용)
    # page_start_rows는 이 섹션 내부의 각 페이지 시작 행 (첫 페이지 포함)
    return dst_start_row + total_rows, page_start_rows


# ==============================================================
# External Link 제거 (속도 최적화의 핵심)
# ==============================================================

def strip_external_links(xlsx_path):
    """xlsx 파일에서 external link 관련 XML + definedNames 전부 제거.
    openpyxl은 external link 때문에 로드·저장이 수십 초 느려짐.
    또한 external link를 참조하는 definedNames는 Excel에서 복구 경고 원인.
    """
    tmp_path = xlsx_path + '.stripped.tmp'

    removed = 0
    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                # external link 관련 파일 제외
                if 'externalLink' in item:
                    removed += 1
                    continue
                data = zin.read(item)

                # workbook.xml 수정 (externalReferences + definedNames 제거)
                if item == 'xl/workbook.xml':
                    text = data.decode('utf-8')
                    text = re.sub(r'<externalReferences>.*?</externalReferences>',
                                  '', text, flags=re.DOTALL)
                    # definedNames 전체 제거 (대부분 external link 참조)
                    text = re.sub(r'<definedNames>.*?</definedNames>',
                                  '', text, flags=re.DOTALL)
                    data = text.encode('utf-8')

                elif item == 'xl/_rels/workbook.xml.rels':
                    text = data.decode('utf-8')
                    text = re.sub(r'<Relationship[^/]+externalLink[^/]+/>',
                                  '', text)
                    data = text.encode('utf-8')

                elif item == '[Content_Types].xml':
                    text = data.decode('utf-8')
                    text = re.sub(r'<Override[^/]+externalLink[^/]+/>',
                                  '', text)
                    data = text.encode('utf-8')

                zout.writestr(item, data)

    os.replace(tmp_path, xlsx_path)
    return removed


# ==============================================================
# 메인
# ==============================================================

def main():
    t_start = time.time()
    print("=" * 70)
    print("돌관공사비 노무비 출력일보 자동 작성 (v11)")
    print("=" * 70)

    if not os.path.exists(TEMPLATE_PATH):
        print(f"❌ 템플릿 파일 없음: {TEMPLATE_PATH}")
        return
    if not os.path.exists(SOURCE_DIR):
        print(f"❌ 소스 폴더 없음: {SOURCE_DIR}")
        return

    out_dir = os.path.dirname(OUTPUT_PATH)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    t0 = time.time()
    shutil.copy2(TEMPLATE_PATH, OUTPUT_PATH)
    print(f"✓ 템플릿 복사: {OUTPUT_PATH} ({time.time()-t0:.1f}초)")

    # External link 제거 (속도 최적화)
    t0 = time.time()
    removed = strip_external_links(OUTPUT_PATH)
    print(f"✓ External link 제거: {removed}개 ({time.time()-t0:.1f}초)")

    t0 = time.time()
    wb = load_workbook(OUTPUT_PATH)
    print(f"✓ 워크북 로드: {time.time()-t0:.1f}초")

    if TEMPLATE_SHEET not in wb.sheetnames:
        print(f"❌ '{TEMPLATE_SHEET}' 시트 없음")
        return
    ws = wb[TEMPLATE_SHEET]

    # 견본 섹션 위치
    proto_section_start = None
    for r in range(1, min(ws.max_row + 1, 200)):
        v = ws.cell(row=r, column=COL_ORDER).value
        if isinstance(v, str) and any(kw in v for kw in SECTION_TITLE_KEYWORDS):
            proto_section_start = r
            break
    if proto_section_start is None:
        print(f"❌ 견본 섹션 없음")
        return
    print(f"✓ 견본 섹션 위치: row {proto_section_start}")

    # 세 블록 스냅샷
    max_col = ws.max_column
    header_block = snapshot_cells(ws, proto_section_start, HEADER_ROWS, max_col)
    proto_first_worker_row = proto_section_start + HEADER_ROWS
    worker_block = snapshot_cells(ws, proto_first_worker_row, WORKER_ROW_SPAN, max_col)
    proto_sum_start = proto_section_start + PROTO_SECTION_SIZE - 2
    sum_block = snapshot_cells(ws, proto_sum_start, 2, max_col)

    print(f"  헤더 블록: {len(header_block['cells'])}셀")
    print(f"  근로자 블록: {len(worker_block['cells'])}셀")
    print(f"  합계 블록: {len(sum_block['cells'])}셀")

    # v10: 시트 삭제 전 열 너비/행 높이 기본값 백업
    col_dims_snapshot = snapshot_column_dimensions(ws)
    row_default_snapshot = snapshot_row_default(ws)
    # v11: 인쇄 설정 백업
    print_settings_snapshot = snapshot_print_settings(ws)
    print(f"  ✓ 열 너비 {len(col_dims_snapshot)}개 + 인쇄 설정 백업")

    # 시트 재생성
    del wb[TEMPLATE_SHEET]
    ws = wb.create_sheet(TEMPLATE_SHEET)
    # v10: 백업한 열 너비 복원
    restore_column_dimensions(ws, col_dims_snapshot)
    restore_row_default(ws, row_default_snapshot)
    # v11: 인쇄 설정 복원
    restore_print_settings(ws, print_settings_snapshot)
    print("✓ 기존 시트 삭제 후 재생성 (열 너비 + 인쇄 설정 복원)")

    # 소스 파일
    source_files = []
    for ext in ('*.xlsx', '*.xlsm'):
        source_files.extend(glob.glob(os.path.join(SOURCE_DIR, ext)))
    source_files = [f for f in source_files if not os.path.basename(f).startswith('~$')]
    source_files.sort()
    if not source_files:
        print(f"❌ 소스 폴더에 엑셀 파일 없음")
        return
    print(f"✓ 소스 파일 {len(source_files)}개 발견")

    # 소스 파일 읽기 (시간 측정)
    t_src = time.time()
    monthly_data = {}
    for idx_f, fp in enumerate(source_files, 1):
        fn = os.path.basename(fp)
        y, m = extract_year_month_from_filename(fn)
        if y is None:
            print(f"  ⚠ 연/월 인식 실패: {fn}")
            continue
        t_file = time.time()
        try:
            data = read_source_file(fp)
        except Exception as e:
            print(f"    ❌ 읽기 실패: {e}")
            continue
        monthly_data[(y, m)] = data
        print(f"  [{idx_f}/{len(source_files)}] {y}년 {m:02d}월 ({time.time()-t_file:.1f}초)")
    print(f"✓ 소스 파일 읽기 완료: {time.time()-t_src:.1f}초")

    if not monthly_data:
        print("❌ 읽은 데이터 없음")
        return

    # 섹션 생성 (가변 길이)
    t_sec = time.time()
    months_sorted = sorted(monthly_data.keys())
    cur_row = 1
    section_log = []
    all_page_starts = []  # v11: 모든 섹션의 페이지 시작 행 (나중에 row_breaks 설정용)
    print(f"\n[섹션 생성 시작] 총 {len(months_sorted)}개월 처리")

    for idx_m, (y, m) in enumerate(months_sorted, start=1):
        t_month = time.time()
        month_data = monthly_data[(y, m)]
        cat_counts = {}
        for cat in CATEGORIES:
            tot = sum(len(w) for w in month_data.get(cat, {}).values())
            if tot > 0:
                cat_counts[cat] = tot
        summary = ', '.join(f"{c}={n}명" for c, n in cat_counts.items())

        for cat in CATEGORIES:
            cat_data = month_data.get(cat, {})
            flat = []
            for gj in sorted(cat_data.keys()):
                for worker in sorted(cat_data[gj].keys()):
                    flat.append((gj, worker, cat_data[gj][worker]))
            if not flat:
                continue

            next_row, page_starts = build_section(
                ws, cur_row,
                header_block, worker_block, sum_block,
                flat, y, m, cat,
                proto_first_worker_row=proto_first_worker_row,
                proto_section_start=proto_section_start,
            )
            section_log.append((cur_row, y, m, cat, len(flat), next_row - cur_row))
            all_page_starts.extend(page_starts)
            cur_row = next_row

        print(f"  [{idx_m}/{len(months_sorted)}] {y}년 {m:02d}월  → {summary or '데이터 없음'} ({time.time()-t_month:.1f}초, 누적 {cur_row}행)")

    print(f"\n✓ 총 {len(section_log)}개 섹션 생성 완료: {time.time()-t_sec:.1f}초")
    for (r, y, m, cat, n, length) in section_log[:15]:
        print(f"   row {r:>6}: {y}년 {m:02d}월 [{cat}] {n}명 ({length}행)")
    if len(section_log) > 15:
        print(f"   ... (총 {len(section_log)}개)")

    # v11: 인쇄 영역 설정 + 페이지 나눔 설정
    # all_page_starts: 모든 섹션의 각 페이지 시작 행
    # 마지막 행 = 최종 cur_row - 1 (cur_row는 "다음 섹션이 시작될 행")
    last_row_total = cur_row - 1
    # page_boundary_rows: 각 페이지 시작 행 중 첫 페이지(1)를 제외
    boundary_rows = sorted(set(r for r in all_page_starts if r > 1))
    set_print_area_and_breaks(ws, last_row_total, boundary_rows)
    print(f"✓ 인쇄 영역 설정: C1:AF{last_row_total}, 페이지 나눔 {len(boundary_rows)}개")

    # 저장 - 통합 파일
    t_save = time.time()
    print("\n통합 파일 저장 중...")
    wb.save(OUTPUT_PATH)
    print(f"✓ 통합 파일 저장 완료: {time.time()-t_save:.1f}초 ({os.path.getsize(OUTPUT_PATH)/1024/1024:.1f} MB)")

    # 연도별 분할 파일 생성 (v8 신규)
    # 각 연도별로 '노무비 출력일보' 시트만 포함하는 별도 파일 저장
    t_split = time.time()
    print("\n연도별 파일 분할 중...")
    years_data = {}  # {year: [(y,m,cat,flat), ...]}
    for (y, m) in months_sorted:
        month_data = monthly_data[(y, m)]
        for cat in CATEGORIES:
            cat_data = month_data.get(cat, {})
            flat = []
            for gj in sorted(cat_data.keys()):
                for worker in sorted(cat_data[gj].keys()):
                    flat.append((gj, worker, cat_data[gj][worker]))
            if not flat:
                continue
            years_data.setdefault(y, []).append((y, m, cat, flat))

    out_base = os.path.splitext(OUTPUT_PATH)[0]  # 확장자 제외 경로

    for year in sorted(years_data.keys()):
        year_path = f"{out_base}_{year}.xlsx"
        t_year = time.time()

        # 원본 템플릿에서 시작 (통합 파일이 아닌 것이 핵심)
        shutil.copy2(TEMPLATE_PATH, year_path)
        # external link + definedNames 제거
        strip_external_links(year_path)

        wb_y = load_workbook(year_path)
        # 다른 시트들 모두 삭제 (노무비 출력일보만 남기기 위해)
        # 주의: 삭제된 시트를 참조하는 수식은 #REF!가 되지만,
        # 이 파일은 '노무비 출력일보' 시트만 필요하므로 무관.
        for sheet_name in list(wb_y.sheetnames):
            if sheet_name != TEMPLATE_SHEET:
                del wb_y[sheet_name]
        # 노무비 출력일보 시트도 새로 생성
        if TEMPLATE_SHEET in wb_y.sheetnames:
            del wb_y[TEMPLATE_SHEET]
        ws_y = wb_y.create_sheet(TEMPLATE_SHEET)
        # v10: 열 너비 복원
        restore_column_dimensions(ws_y, col_dims_snapshot)
        restore_row_default(ws_y, row_default_snapshot)
        # v11: 인쇄 설정 복원
        restore_print_settings(ws_y, print_settings_snapshot)

        # 해당 연도 섹션만 paste
        cur_row_y = 1
        year_sec_count = 0
        year_page_starts = []  # v11
        for (y, m, cat, flat) in years_data[year]:
            cur_row_y, page_starts = build_section(
                ws_y, cur_row_y,
                header_block, worker_block, sum_block,
                flat, y, m, cat,
                proto_first_worker_row=proto_first_worker_row,
                proto_section_start=proto_section_start,
            )
            year_sec_count += 1
            year_page_starts.extend(page_starts)

        # v11: 연도별 파일도 인쇄 영역 + row_breaks
        last_row_y = cur_row_y - 1
        boundary_rows_y = sorted(set(r for r in year_page_starts if r > 1))
        set_print_area_and_breaks(ws_y, last_row_y, boundary_rows_y)

        wb_y.save(year_path)
        size_mb = os.path.getsize(year_path) / 1024 / 1024
        print(f"  ✓ {year}년 ({year_sec_count}개 섹션, {last_row_y}행, 나눔 {len(boundary_rows_y)}개): {year_path} [{size_mb:.2f} MB, {time.time()-t_year:.1f}초]")

    print(f"✓ 연도별 분할 완료: {time.time()-t_split:.1f}초")

    print(f"\n{'='*70}")
    print(f"✓ 완료!")
    print(f"  통합 파일: {OUTPUT_PATH}")
    print(f"  연도별 파일: {out_base}_YYYY.xlsx (노무비 출력일보 시트만)")
    print(f"  [총 소요 시간] {time.time()-t_start:.1f}초")
    print(f"{'='*70}")
    print("\n※ Excel에서 Ctrl+Alt+F9로 전체 재계산 실행하세요.")


if __name__ == '__main__':
    main()
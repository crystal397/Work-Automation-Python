"""
사진대지 엑셀 자동 삽입 스크립트
=================================

사용법:
    python insert_photos.py <엑셀파일> <사진폴더> [출력파일]

예시:
    python insert_photos.py 사진대지.xlsx ./photos/
    python insert_photos.py 사진대지.xlsx ./photos/ 결과물.xlsx

동작:
    - 엑셀 파일에서 큰 병합 셀 영역(사진 슬롯)을 자동 탐지
    - 사진 폴더의 파일들을 파일명 순서대로 정렬
    - 각 슬롯에 비율 유지 + 중앙 정렬로 자동 삽입
    - 사진이 슬롯 경계(병합 셀 범위)를 절대 벗어나지 않음
    - 스마트폰 사진의 EXIF 회전 정보 자동 반영

필요 라이브러리:
    pip install openpyxl pillow
"""

import io
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage, ImageOps

# --- 설정값 (필요시 조정) ---
IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.gif', '.bmp'}
MIN_SLOT_ROWS = 5       # 사진 슬롯으로 인식할 최소 병합 행 수
MIN_SLOT_COLS = 4       # 사진 슬롯으로 인식할 최소 병합 열 수
SLOT_PADDING_PX = 4     # 슬롯 테두리와 사진 사이 여백 (픽셀)
JPEG_QUALITY = 85       # 삽입 이미지 JPEG 품질 (파일 크기 vs 품질)


def excel_width_to_pixels(width):
    """엑셀 컬럼 너비 → 픽셀 (Calibri 11pt 근사)"""
    if width is None:
        width = 8.43
    if width < 1:
        return int(width * 12)
    return int(width * 7 + 5)


def excel_height_to_pixels(height):
    """엑셀 행 높이(pt) → 픽셀 (96 DPI)"""
    if height is None:
        height = 15
    return int(height * 96 / 72)


def get_column_width(ws, col_idx):
    """
    주어진 열(1-index)의 실제 너비를 반환.
    엑셀 XML에서 <col min=X max=Y width=W>로 범위 지정된 경우
    openpyxl의 ws.column_dimensions[letter]는 X 이외 열에 대해 None을 반환함.
    → 모든 column_dimensions entry를 순회하며 min~max 범위 안에 있는지 확인
    """
    for cd in ws.column_dimensions.values():
        if cd.min is not None and cd.max is not None:
            if cd.min <= col_idx <= cd.max and cd.width is not None:
                return cd.width
    # 시트 기본 너비
    if ws.sheet_format and ws.sheet_format.defaultColWidth:
        return ws.sheet_format.defaultColWidth
    return 8.43  # 엑셀 기본값


def get_row_height(ws, row_idx):
    """행 높이를 안전하게 조회"""
    rd = ws.row_dimensions.get(row_idx)
    if rd and rd.height is not None:
        return rd.height
    if ws.sheet_format and ws.sheet_format.defaultRowHeight:
        return ws.sheet_format.defaultRowHeight
    return 15  # 엑셀 기본값


def find_photo_slots(ws):
    """큰 병합 영역을 사진 슬롯으로 탐지, 위→아래 순으로 정렬"""
    slots = []
    for mr in ws.merged_cells.ranges:
        rows = mr.max_row - mr.min_row + 1
        cols = mr.max_col - mr.min_col + 1
        if rows >= MIN_SLOT_ROWS and cols >= MIN_SLOT_COLS:
            slots.append(mr)
    slots.sort(key=lambda r: (r.min_row, r.min_col))
    return slots


def get_slot_pixel_size(ws, slot):
    """병합 영역의 총 가로·세로 픽셀 크기"""
    width_px = sum(
        excel_width_to_pixels(get_column_width(ws, c))
        for c in range(slot.min_col, slot.max_col + 1)
    )
    height_px = sum(
        excel_height_to_pixels(get_row_height(ws, r))
        for r in range(slot.min_row, slot.max_row + 1)
    )
    return width_px, height_px


def pixel_offset_to_cell(ws, start_col, start_row, offset_px_x, offset_px_y):
    """
    슬롯 좌상단(start_col, start_row)에서 (offset_px_x, offset_px_y)만큼
    떨어진 지점이 어느 셀의 어느 오프셋인지 역산.
    반환: (col_0index, col_offset_px, row_0index, row_offset_px)
    """
    # 가로 방향
    acc = 0
    end_col_0 = start_col - 1
    end_col_off = offset_px_x
    c = start_col
    while True:
        w = excel_width_to_pixels(get_column_width(ws, c))
        if acc + w >= offset_px_x:
            end_col_0 = c - 1
            end_col_off = offset_px_x - acc
            break
        acc += w
        c += 1

    # 세로 방향
    acc = 0
    end_row_0 = start_row - 1
    end_row_off = offset_px_y
    r = start_row
    while True:
        h = excel_height_to_pixels(get_row_height(ws, r))
        if acc + h >= offset_px_y:
            end_row_0 = r - 1
            end_row_off = offset_px_y - acc
            break
        acc += h
        r += 1

    return end_col_0, end_col_off, end_row_0, end_row_off


def prepare_image(img_path, slot_w_px, slot_h_px):
    """
    이미지를 슬롯 크기에 맞게 비율 유지 리사이즈한 뒤,
    슬롯과 정확히 같은 크기의 흰 배경 캔버스 중앙에 붙여서 반환.
    이렇게 하면 엑셀이 어떻게 앵커를 해석하든 사진이 항상 중앙에 위치함.
    """
    pil_img = PILImage.open(img_path)
    pil_img = ImageOps.exif_transpose(pil_img)  # 스마트폰 회전 보정

    if pil_img.mode not in ('RGB', 'L'):
        pil_img = pil_img.convert('RGB')

    # 비율 유지 리사이즈 (패딩 고려)
    avail_w = max(slot_w_px - SLOT_PADDING_PX * 2, 1)
    avail_h = max(slot_h_px - SLOT_PADDING_PX * 2, 1)

    ratio = min(avail_w / pil_img.width, avail_h / pil_img.height)
    new_w = max(int(pil_img.width * ratio), 1)
    new_h = max(int(pil_img.height * ratio), 1)
    pil_img = pil_img.resize((new_w, new_h), PILImage.LANCZOS)

    # 슬롯 크기의 흰 캔버스를 만들고 사진을 중앙에 붙이기
    canvas = PILImage.new('RGB', (slot_w_px, slot_h_px), (255, 255, 255))
    paste_x = (slot_w_px - new_w) // 2
    paste_y = (slot_h_px - new_h) // 2
    canvas.paste(pil_img, (paste_x, paste_y))

    buf = io.BytesIO()
    canvas.save(buf, format='JPEG', quality=JPEG_QUALITY, optimize=True)
    buf.seek(0)
    return buf, slot_w_px, slot_h_px


def insert_one_image(ws, slot, img_path):
    """
    한 슬롯에 이미지 삽입.
    이미지 자체가 슬롯 크기로 만들어져 있고 사진이 그 중앙에 있으므로,
    앵커는 슬롯 전체 범위(좌상단 셀 ~ 우하단 바로 다음 셀)에 그대로 붙이면 됨.
    """
    slot_w, slot_h = get_slot_pixel_size(ws, slot)
    img_buf, img_w, img_h = prepare_image(img_path, slot_w, slot_h)

    # 앵커: 슬롯 좌상단 셀의 시작 ~ 슬롯 우하단 셀의 끝(=다음 셀의 시작)
    from_marker = AnchorMarker(
        col=slot.min_col - 1, colOff=0,
        row=slot.min_row - 1, rowOff=0,
    )
    to_marker = AnchorMarker(
        col=slot.max_col, colOff=0,
        row=slot.max_row, rowOff=0,
    )

    xl_img = XLImage(img_buf)
    xl_img.width = img_w
    xl_img.height = img_h
    xl_img.anchor = TwoCellAnchor(editAs='oneCell', _from=from_marker, to=to_marker)
    ws.add_image(xl_img)


def insert_photos(excel_path, photo_dir, output_path=None):
    excel_path = Path(excel_path)
    photo_dir = Path(photo_dir)

    if output_path is None:
        output_path = excel_path.parent / f"{excel_path.stem}_완성{excel_path.suffix}"
    else:
        output_path = Path(output_path)

    # 사진 파일 수집 (파일명 오름차순)
    photos = sorted(
        [p for p in photo_dir.iterdir()
         if p.is_file() and p.suffix.lower() in IMAGE_EXTS],
        key=lambda p: p.name.lower()
    )
    if not photos:
        print(f"⚠️  '{photo_dir}' 에 이미지 파일이 없습니다.")
        return

    print(f"📸 발견한 사진: {len(photos)}장")

    # 엑셀 열기 & 슬롯 탐지
    wb = load_workbook(excel_path)
    ws = wb.active

    slots = find_photo_slots(ws)
    print(f"📋 발견한 사진 슬롯: {len(slots)}개")

    # 삽입
    insert_count = min(len(photos), len(slots))
    for i in range(insert_count):
        insert_one_image(ws, slots[i], photos[i])
        print(f"  ✅ [{i+1:3d}/{insert_count}] {photos[i].name} → {slots[i]}")

    wb.save(output_path)
    print(f"\n💾 저장 완료: {output_path}")

    if len(photos) > len(slots):
        print(f"⚠️  사진 {len(photos) - len(slots)}장은 슬롯 부족으로 삽입되지 않았습니다.")
    elif len(photos) < len(slots):
        print(f"ℹ️  슬롯 {len(slots) - len(photos)}개가 비어있습니다.")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    excel_file = sys.argv[1]
    photo_folder = sys.argv[2]
    output_file = sys.argv[3] if len(sys.argv) > 3 else None

    insert_photos(excel_file, photo_folder, output_file)

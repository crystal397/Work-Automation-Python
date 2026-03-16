"""
공기연장 간접비 보고서 자동화 스크립트
엑셀 데이터를 읽어 한글(HWP) 템플릿에 자동으로 입력합니다.
"""

import os
import warnings
from datetime import datetime

import openpyxl
import win32com.client as win32
import tkinter as tk
from tkinter import filedialog

# 엑셀 Print area 경고 무시
warnings.filterwarnings("ignore", category=UserWarning)


# ─────────────────────────────────────────────
# 상수 정의
# ─────────────────────────────────────────────
class SheetName:
    SUMMARY       = "전체총괄표"
    INDIRECT_WAGE = "1-1. 간접노무비 집계표"
    SEVERANCE     = "1-3. 퇴직금"


class FieldName:
    START_DAY  = "start_day"
    END_DAY    = "end_day"
    DURATION   = "duration"
    RATE_SAN   = "rate_san"
    RATE_GO    = "rate_go"
    RATE_IL    = "rate_il"

    TABLE_1_START = "table_start"
    TABLE_2_START = "second_table_start"
    TABLE_3_START = "third_table_start"
    TABLE_3_TOTAL_SALARY    = "third_total_salary"
    TABLE_3_TOTAL_SEVERANCE = "third_total_severance"
    TABLE_3_TOTAL_SUM       = "third_total_sum"
    TABLE_4_START = "fourth_table_start"
    TABLE_5_START = "fifth_table_start"
    TABLE_6_START = "sixth_table_start"


# 첫 번째 표: 전체총괄표에서 읽어올 셀 목록 (순서 중요)
TABLE_1_CELLS = [
    "C8",  "C26", "C21", "C10", "C11", "C12", "C13", "C14",
    "C15", "C16", "C17", "C18", "C19", "C20", "C25", "C23",
    "C24", "C27", "C28", "C29", "C30", "C31", "C32",
]

TABLE_4_CELLS = [
    "C10", "C11", "C12", "C13", "C14", "C15",
    "C16", "C17", "C18", "C19", "C20", "C21",
]
TABLE_5_CELLS = ["C23", "C24", "C25"]
TABLE_6_CELLS = ["C28", "C28"]       # 의도적 중복 (원본 유지)

STOP_KEYWORD = "간접노무비 합계"
MAX_SCAN_ROWS = 100


# ─────────────────────────────────────────────
# 헬퍼 함수
# ─────────────────────────────────────────────
def move_to_field_safe(hwp, field_name: str) -> bool:
    """누름틀로 안전하게 이동. 여러 오버로드를 순서대로 시도."""
    arg_sets = [
        (field_name, True, True, False),
        (field_name, True, True),
        (field_name, True),
        (field_name,),
    ]
    for args in arg_sets:
        try:
            if hwp.MoveToField(*args):
                return True
        except Exception:
            continue
    return False


def insert_text(hwp, text: str) -> None:
    """현재 커서 위치에 텍스트 삽입."""
    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
    hwp.HParameterSet.HInsertText.Text = text
    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)


def format_date(val) -> str:
    if isinstance(val, datetime):
        return val.strftime("%Y. %m. %d.")
    return str(val) if val else ""


def format_percent(val) -> str:
    try:
        return f"{float(val) * 100:.2f}"
    except (ValueError, TypeError):
        return "0.00"


def format_val(val, zero_placeholder: str = "-") -> str:
    """숫자는 천 단위 콤마, 없거나 0이면 placeholder 반환."""
    if val is None or val == "":
        return zero_placeholder
    if isinstance(val, (int, float)):
        return zero_placeholder if val == 0 else format(int(round(val)), ",")
    return str(val).replace('\r', ' ').replace('\n', ' ').strip()


def format_val_blank(val) -> str:
    """숫자는 천 단위 콤마, 없거나 0이면 빈 문자열 반환."""
    return format_val(val, zero_placeholder="")


# ─────────────────────────────────────────────
# 표 공통 입력 유틸
# ─────────────────────────────────────────────
def _write_table_rows(hwp, field_name: str, rows: list[list]) -> None:
    """
    rows: [[col0, col1, ...], ...]
    첫 번째 칸은 누름틀에 직접 쓰고, 나머지는 커서를 이동하며 삽입.
    """
    if not rows:
        return

    col_count = len(rows[0])

    for r, row_values in enumerate(rows):
        if r > 0:
            hwp.Run("TableLowerCell")
            for _ in range(col_count - 1):
                hwp.Run("TableLeftCell")

        for c, cell_value in enumerate(row_values):
            text = format_val_blank(cell_value)

            if r == 0 and c == 0:
                hwp.PutFieldText(field_name, text)
                move_to_field_safe(hwp, field_name)
                hwp.Run("Cancel")
            else:
                insert_text(hwp, text)

            if c < col_count - 1:
                hwp.Run("TableRightCell")


def _write_table_column(hwp, field_name: str, values: list) -> None:
    """
    단일 열(세로 방향)로 값을 내려가며 입력.
    첫 번째 값은 누름틀에, 나머지는 아래로 이동하며 삽입.
    """
    if not values:
        return

    hwp.PutFieldText(field_name, format_val_blank(values[0]))
    move_to_field_safe(hwp, field_name)

    for val in values[1:]:
        hwp.Run("Cancel")
        hwp.Run("TableLowerCell")
        insert_text(hwp, format_val_blank(val))


# ─────────────────────────────────────────────
# 표별 데이터 수집 함수
# ─────────────────────────────────────────────
def _collect_indirect_wage_rows(ws) -> list[list]:
    """1-1. 간접노무비 집계표 시트에서 행 데이터를 수집."""
    rows = []
    for row in range(5, MAX_SCAN_ROWS + 5):
        a = ws[f"A{row}"].value
        if a and STOP_KEYWORD in str(a):
            break
        rows.append([
            ws[f"A{row}"].value,
            ws[f"B{row}"].value,
            ws[f"C{row}"].value,
            ws[f"D{row}"].value,
            ws[f"E{row}"].value,
            ws[f"G{row}"].value,
        ])
    return rows


def _collect_severance_rows(ws) -> tuple[list[list], dict]:
    """1-3. 퇴직금 시트에서 행 데이터와 합계를 수집."""
    rows = []
    totals = {"salary": 0, "severance": 0, "sum": 0}

    for row in range(3, MAX_SCAN_ROWS + 3):
        a = ws[f"A{row}"].value
        if a is not None and STOP_KEYWORD in str(a):
            break
        if not isinstance(a, int):
            continue

        salary    = ws[f"F{row}"].value or 0
        severance = ws[f"G{row}"].value or 0
        total_sum = ws[f"H{row}"].value or 0

        rows.append([
            ws[f"A{row}"].value,
            ws[f"B{row}"].value,
            ws[f"C{row}"].value,
            ws[f"D{row}"].value,
            salary, severance, total_sum,
        ])
        totals["salary"]    += salary
        totals["severance"] += severance
        totals["sum"]       += total_sum

    return rows, totals


# ─────────────────────────────────────────────
# 표별 입력 함수
# ─────────────────────────────────────────────
def insert_second_table(hwp, wb) -> None:
    """두 번째 표: 간접노무비 집계표"""
    print("두 번째 표 데이터 읽는 중...")
    ws = wb[SheetName.INDIRECT_WAGE]
    rows = _collect_indirect_wage_rows(ws)

    if not rows:
        print("  ⚠ 두 번째 표에 입력할 데이터가 없습니다.")
        return

    print(f"  → {len(rows)}줄 입력")
    hwp.Run("MoveDocBegin")

    if not move_to_field_safe(hwp, FieldName.TABLE_2_START):
        print(f"  ✗ '{FieldName.TABLE_2_START}' 누름틀을 찾을 수 없습니다.")
        return

    _write_table_rows(hwp, FieldName.TABLE_2_START, rows)


def insert_third_table(hwp, wb) -> None:
    """세 번째 표: 퇴직금"""
    print("세 번째 표 데이터 읽는 중...")
    ws = wb[SheetName.SEVERANCE]
    rows, totals = _collect_severance_rows(ws)

    if not rows:
        print("  ⚠ 세 번째 표에 입력할 데이터가 없습니다.")
        return

    print(f"  → {len(rows)}줄 입력")
    hwp.Run("Cancel")
    hwp.Run("MoveDocBegin")

    if not move_to_field_safe(hwp, FieldName.TABLE_3_START):
        print(f"  ✗ '{FieldName.TABLE_3_START}' 누름틀을 찾을 수 없습니다.")
        return

    _write_table_rows(hwp, FieldName.TABLE_3_START, rows)

    hwp.PutFieldText(FieldName.TABLE_3_TOTAL_SALARY,    format_val_blank(totals["salary"]))
    hwp.PutFieldText(FieldName.TABLE_3_TOTAL_SEVERANCE, format_val_blank(totals["severance"]))
    hwp.PutFieldText(FieldName.TABLE_3_TOTAL_SUM,       format_val_blank(totals["sum"]))


def _insert_single_column_table(hwp, wb, field_name: str, cells: list[str],
                                 table_label: str) -> None:
    """단일 열 표 입력 공통 함수 (4·5·6번 표 공용)."""
    print(f"{table_label} 데이터 읽는 중...")
    ws = wb[SheetName.SUMMARY]
    values = [ws[cell].value for cell in cells]

    if not move_to_field_safe(hwp, field_name):
        print(f"  ✗ '{field_name}' 누름틀을 찾을 수 없습니다.")
        return

    _write_table_column(hwp, field_name, values)


def insert_fourth_table(hwp, wb) -> None:
    _insert_single_column_table(hwp, wb, FieldName.TABLE_4_START, TABLE_4_CELLS, "네 번째 표")


def insert_fifth_table(hwp, wb) -> None:
    _insert_single_column_table(hwp, wb, FieldName.TABLE_5_START, TABLE_5_CELLS, "다섯 번째 표")


def insert_sixth_table(hwp, wb) -> None:
    _insert_single_column_table(hwp, wb, FieldName.TABLE_6_START, TABLE_6_CELLS, "여섯 번째 표")


# ─────────────────────────────────────────────
# 파일 저장
# ─────────────────────────────────────────────
def save_outputs(hwp, output_dir: str) -> None:
    now = datetime.now().strftime("%Y%m%d")
    base = os.path.join(output_dir, f"공기연장보고서_{now}")

    for ext, fmt in ((base + ".hwp", "HWPX"), (base + ".pdf", "PDF")):
        try:
            hwp.SaveAs(ext, fmt, "")
            print(f"  ✅ 저장 완료: {ext}")
        except Exception as e:
            print(f"  ✗ {fmt} 저장 실패: {e}")


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
def select_files() -> tuple[str, str]:
    """tkinter 파일 다이얼로그로 엑셀·HWP 파일 경로를 반환."""
    root = tk.Tk()
    root.withdraw()

    excel_path = filedialog.askopenfilename(
        title="변환할 엑셀 파일을 선택하세요",
        filetypes=[("Excel Files", "*.xlsx;*.xls")],
    )
    if not excel_path:
        raise FileNotFoundError("엑셀 파일이 선택되지 않았습니다.")

    hwp_path = filedialog.askopenfilename(
        title="한글 템플릿 파일을 선택하세요",
        filetypes=[("HWP Files", "*.hwpx")],
    )
    if not hwp_path:
        raise FileNotFoundError("한글 템플릿 파일이 선택되지 않았습니다.")

    return excel_path, hwp_path


def run_automation() -> None:
    # 1. 파일 선택
    try:
        excel_path, hwp_template_path = select_files()
    except FileNotFoundError as e:
        print(e)
        return

    # 2. 엑셀 데이터 로드
    print("\n엑셀 데이터를 읽는 중...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[SheetName.SUMMARY]

    # 본문 텍스트용 데이터
    start_day = format_date(ws["H4"].value)
    end_day   = format_date(ws["I4"].value)
    duration  = str(ws["J4"].value)
    rate_san  = format_percent(ws["G23"].value)
    rate_go   = format_percent(ws["G24"].value)
    rate_il   = format_percent(ws["G28"].value)

    # 첫 번째 표 데이터 (단일 열)
    table1_values = [ws[cell].value for cell in TABLE_1_CELLS]

    # 3. 한글 실행 및 템플릿 열기
    print("한글 프로그램 실행 중...")
    try:
        hwp = win32.Dispatch("HWPFrame.HwpObject")
    except Exception as e:
        print(f"한글 실행 실패: {e}")
        return

    hwp.XHwpWindows.Item(0).Visible = True

    try:
        hwp.Open(hwp_template_path, "HWPX", "")
    except Exception as e:
        print(f"파일 열기 실패: {e}")
        return

    # 4. 본문 누름틀 입력
    print("\n데이터 입력 시작")
    hwp.PutFieldText(FieldName.START_DAY, start_day)
    hwp.PutFieldText(FieldName.END_DAY,   end_day)
    hwp.PutFieldText(FieldName.DURATION,  duration)

    # 5. 첫 번째 표 입력
    print("첫 번째 표 데이터 입력 중...")
    if move_to_field_safe(hwp, FieldName.TABLE_1_START):
        _write_table_column(hwp, FieldName.TABLE_1_START, table1_values)
    else:
        print(f"  ✗ '{FieldName.TABLE_1_START}' 누름틀을 찾을 수 없습니다.")

    # 6. 나머지 표 입력
    insert_second_table(hwp, wb)
    insert_third_table(hwp, wb)
    insert_fourth_table(hwp, wb)

    # 7. 비율 필드 입력 (표 제어 모드 탈출 후)
    hwp.Run("Cancel")
    hwp.Run("MoveDocBegin")
    for field, value in (
        (FieldName.RATE_SAN, rate_san),
        (FieldName.RATE_GO,  rate_go),
        (FieldName.RATE_IL,  rate_il),
    ):
        hwp.MoveToField(field, True, True, False)
        hwp.PutFieldText(field, value)

    # 8. 다섯·여섯 번째 표 입력
    insert_fifth_table(hwp, wb)
    insert_sixth_table(hwp, wb)

    # 9. 저장
    print("\n파일 저장 중...")
    save_outputs(hwp, os.getcwd())


if __name__ == "__main__":
    run_automation()

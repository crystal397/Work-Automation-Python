"""
excel_reader.py
엑셀 파일에서 데이터를 읽어오는 모듈
"""
import openpyxl
from utils import format_date, format_percent


# 전체총괄표 셀 매핑 (표 순서대로 정의)
MAIN_TABLE_CELLS = [
    "C8",  "C26", "C21", "C10", "C11", "C12", "C13",
    "C14", "C15", "C16", "C17", "C18", "C19", "C20",
    "C25", "C23", "C24", "C27", "C28", "C29",
    "C30", "C31", "C32",
]
FOURTH_TABLE_CELLS = [
    "C10", "C11", "C12", "C13", "C14", "C15",
    "C16", "C17", "C18", "C19", "C20", "C21",
]
FIFTH_TABLE_CELLS = ["C23", "C24", "C25"]
SIXTH_TABLE_CELLS = ["C28", "C28"]


def load_workbook(excel_path: str):
    """엑셀 파일 로드 (수식 결과값 사용)"""
    return openpyxl.load_workbook(excel_path, data_only=True)


def read_header_info(wb) -> dict:
    """전체총괄표에서 헤더 정보(기간, 비율) 읽기"""
    ws = wb["전체총괄표"]
    return {
        "start_day": format_date(ws["H4"].value),
        "end_day":   format_date(ws["I4"].value),
        "duration":  str(ws["J4"].value),
        "rate_san":  format_percent(ws["G23"].value),
        "rate_go":   format_percent(ws["G24"].value),
        "rate_il":   format_percent(ws["G28"].value),
    }


def read_main_table_values(wb) -> list:
    """전체총괄표에서 첫 번째 표에 들어갈 셀 값 목록 반환"""
    ws = wb["전체총괄표"]
    return [ws[cell].value for cell in MAIN_TABLE_CELLS]


def read_indirect_labor_table(wb) -> list[list]:
    """
    '1-1. 간접노무비 집계표' 시트에서 두 번째 표 데이터 읽기
    반환: [[A, B, C, D, E, G], ...] 형태의 리스트
    """
    ws = wb["1-1. 간접노무비 집계표"]
    rows = []
    row_num = 5

    while True:
        a_val = ws[f"A{row_num}"].value
        if a_val and "간접노무비 합계" in str(a_val):
            break
        rows.append([
            ws[f"A{row_num}"].value,
            ws[f"B{row_num}"].value,
            ws[f"C{row_num}"].value,
            ws[f"D{row_num}"].value,
            ws[f"E{row_num}"].value,
            ws[f"G{row_num}"].value,
        ])
        row_num += 1

    return rows


def read_severance_table(wb) -> tuple[list[list], dict]:
    """
    '1-3. 퇴직금' 시트에서 세 번째 표 데이터 읽기
    반환: (행 데이터 리스트, 합계 딕셔너리)
    """
    ws = wb["1-3. 퇴직금"]
    rows = []
    totals = {"salary": 0, "severance": 0, "sum": 0}
    row_num = 3

    while row_num < 100:
        a_val = ws[f"A{row_num}"].value
        if a_val is not None and "간접노무비 합계" in str(a_val):
            break
        if isinstance(a_val, int):
            row_data = [
                ws[f"A{row_num}"].value,
                ws[f"B{row_num}"].value,
                ws[f"C{row_num}"].value,
                ws[f"D{row_num}"].value,
                ws[f"F{row_num}"].value,
                ws[f"G{row_num}"].value,
                ws[f"H{row_num}"].value,
            ]
            rows.append(row_data)
            totals["salary"]    += ws[f"F{row_num}"].value or 0
            totals["severance"] += ws[f"G{row_num}"].value or 0
            totals["sum"]       += ws[f"H{row_num}"].value or 0
        row_num += 1

    return rows, totals


def read_single_column_values(wb, cells: list) -> list:
    """전체총괄표에서 단일 열 셀 목록의 값 반환"""
    ws = wb["전체총괄표"]
    return [ws[cell].value for cell in cells]


def read_fourth_table_values(wb) -> list:
    return read_single_column_values(wb, FOURTH_TABLE_CELLS)


def read_fifth_table_values(wb) -> list:
    return read_single_column_values(wb, FIFTH_TABLE_CELLS)


def read_sixth_table_values(wb) -> list:
    return read_single_column_values(wb, SIXTH_TABLE_CELLS)

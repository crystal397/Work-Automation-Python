"""산출내역서(Excel) 자동 추출.

구조가 회사마다 다르므로 두 가지 모드:
- AUTO: 시트 셀들을 스캔하면서 "일반관리비율", "이윤율", "산재보험료", "고용보험료" 라벨 우측/하단 셀의 퍼센트 값을 찾는다.
- MANUAL: 호출 시 mapping={"general_admin_percent": ("시트명","셀주소"), ...} 명시.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from contract_meta.audit import make_source
from contract_meta.models import Source

FieldLocator = tuple[str, str]   # (sheet_name, cell_address)


_LABEL_PATTERNS = {
    "general_admin_percent": re.compile(r"일반관리비"),
    "profit_percent": re.compile(r"이\s*윤"),
    "industrial_accident_insurance_percent": re.compile(r"산재"),
    "employment_insurance_percent": re.compile(r"고용"),
}


@dataclass
class ExtractedRate:
    field: str
    value: float
    source: Source


def extract_rates_auto(xlsx_path: str | Path) -> tuple[list[ExtractedRate], list[str]]:
    """라벨 셀을 찾아 같은 행/다음 열의 퍼센트 값을 추출 (AUTO 모드)."""
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    file_name = Path(xlsx_path).name
    rates: list[ExtractedRate] = []
    warnings: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                for field, pat in _LABEL_PATTERNS.items():
                    if not pat.search(cell.value):
                        continue
                    if any(r.field == field for r in rates):
                        continue
                    pct_cell = _find_percent_in_row(row, cell.column)
                    if pct_cell is None:
                        continue
                    src = make_source(
                        file=file_name,
                        method="xlsx",
                        sheet=sheet_name,
                        cell=pct_cell.coordinate,
                        field_label=cell.value.strip(),
                        raw_text=str(pct_cell.value),
                    )
                    rates.append(ExtractedRate(field=field, value=_to_percent(pct_cell.value), source=src))

    found = {r.field for r in rates}
    for f in _LABEL_PATTERNS:
        if f not in found:
            warnings.append(f"xlsx auto-scan: {f} 미발견")

    return rates, warnings


def extract_rates_mapping(xlsx_path: str | Path, mapping: dict[str, FieldLocator]) -> list[ExtractedRate]:
    """명시적으로 (sheet, cell)을 지정해 추출 (MANUAL 모드)."""
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    file_name = Path(xlsx_path).name
    rates: list[ExtractedRate] = []
    for field, (sheet_name, cell_addr) in mapping.items():
        ws = wb[sheet_name]
        v = ws[cell_addr].value
        src = make_source(
            file=file_name,
            method="xlsx",
            sheet=sheet_name,
            cell=cell_addr,
            raw_text=str(v),
        )
        rates.append(ExtractedRate(field=field, value=_to_percent(v), source=src))
    return rates


def _find_percent_in_row(row: Iterable[Cell], start_col: int) -> Cell | None:
    for cell in row:
        if cell.column <= start_col:
            continue
        v = cell.value
        if isinstance(v, (int, float)) and 0 < v < 1:
            return cell
        if isinstance(v, (int, float)) and 0 < v < 100:
            return cell
        if isinstance(v, str) and "%" in v:
            return cell
    return None


def _to_percent(v) -> float:
    if isinstance(v, str):
        v = v.replace("%", "").strip()
        return float(v)
    if isinstance(v, (int, float)):
        return float(v) * 100 if v < 1 else float(v)
    raise ValueError(f"퍼센트 값 해석 불가: {v!r}")

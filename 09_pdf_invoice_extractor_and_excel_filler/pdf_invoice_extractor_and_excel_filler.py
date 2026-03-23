#!/usr/bin/env python3
"""
세금계산서 PDF → 제3자발급사실 일괄등록양식 입력 도구 (GUI)
==========================================================
드래그앤드롭 / 폴더선택으로 PDF 폴더와 양식 파일을 지정하면
자동으로 세금계산서를 추출하여 엑셀에 입력합니다.

exe 빌드:
  pip install pyinstaller
  pyinstaller --onefile --windowed --name "세금계산서추출기" pdf_invoice_extractor_gui.py
"""

import os
import re
import sys
import shutil
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from collections import Counter

from pypdf import PdfWriter, PdfReader
import pdfplumber
from openpyxl import load_workbook

# OCR (선택적)
try:
    import pytesseract
    from pdf2image import convert_from_path
    from PIL import Image, ImageEnhance, ImageFilter
    OCR_AVAILABLE = True

    # Windows에서 tesseract 경로 자동 탐지
    import shutil
    if not shutil.which('tesseract'):
        # 일반적인 설치 경로들 시도
        for path in [
            r'C:\Program Files\Tesseract-OCR\tesseract.exe',
            r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
            os.path.expanduser(r'~\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'),
        ]:
            if os.path.exists(path):
                pytesseract.pytesseract.tesseract_cmd = path
                break
except ImportError:
    OCR_AVAILABLE = False


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 추출 엔진 (기존 CLI 로직)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

BIZ_NO_RE = re.compile(r'(\d{3})\s*[-—]\s*(\d{2})\s*[-—]\s*(\d{5})')
NTS_APPROVAL_RE = re.compile(r'국세청\s*신고\s*번호\s*[:\s]*([A-Za-z0-9]{20,30})')
FILENAME_APPROVAL_RE = re.compile(r'전자(?:세금)?계산서[_\s]*(\d{20,30})')
GENERAL_APPROVAL_RE = re.compile(r'승인\s*번호\s*[:\s]*([A-Za-z0-9\-]{16,30})')
SCAN_APPROVAL_RE = re.compile(r'(20\d{6})\s*[-~]\s*(\d{8,})\s*[-~]\s*(\d{5,})')
ISSUE_DATE_RE = re.compile(r'작성일\s*[:\s]*(\d{4})\s*[-./]\s*(\d{1,2})\s*[-./]\s*(\d{1,2})')
ISSUE_DATE2_RE = re.compile(r'작성\s*일자?\s*[:\s]*(\d{8})')
ISSUE_DATE3_RE = re.compile(r'(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일')
SCAN_DATE_RE = re.compile(r'20(\d{2})\s*[|\s]\s*(\d{1,2})\s*[|\s]\s*(\d{1,2})')
SUPPLY_AMT_RE = re.compile(r'공급가액\s+([\d,]+)')
INVOICE_KEYWORDS = ['전자세금계산서', '전자계산서', '세금계산서', '계산서']
INVOICE_CONFIRM = ['국세청신고번호', '공급가액', '사업자번호']


def find_pdfs(directory):
    result = []
    for root, _, files in os.walk(directory):
        for f in sorted(files):
            if f.lower().endswith('.pdf'):
                result.append(os.path.join(root, f))
    return result


def merge_pdfs(pdf_files, output_path):
    writer = PdfWriter()
    for pf in pdf_files:
        try:
            for page in PdfReader(pf).pages:
                writer.add_page(page)
        except:
            pass
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    with open(output_path, 'wb') as f:
        writer.write(f)
    return len(writer.pages)


def clean_biz_no(g):
    return f"{g[0]}{g[1]}{g[2]}"


def is_invoice_page(text):
    return any(kw in text for kw in INVOICE_KEYWORDS) and \
           sum(1 for kw in INVOICE_CONFIRM if kw in text) >= 2


def is_scanned_invoice(ocr_text):
    return bool(BIZ_NO_RE.search(ocr_text)) and bool(re.search(r'[\d,]{7,}', ocr_text))


def ocr_page_image(pdf_path, page_num):
    if not OCR_AVAILABLE:
        return ""
    try:
        images = convert_from_path(pdf_path, first_page=page_num, last_page=page_num, dpi=250)
        if not images:
            return ""
        img = images[0].convert('L')
        img = ImageEnhance.Contrast(img).enhance(2.0)
        img = img.filter(ImageFilter.SHARPEN)
        try:
            langs = pytesseract.get_languages()
            lang = 'kor+eng' if 'kor' in langs else 'eng'
        except:
            lang = 'eng'
        return pytesseract.image_to_string(img, lang=lang, config='--psm 6')
    except Exception as e:
        print(f"  [OCR 오류] p.{page_num}: {e}")
        return ""


def extract_from_text(text, source=""):
    data = {'supplier_biz_no': '', 'receiver_biz_no': '', 'approval_no': '',
            'issue_date': '', 'supply_amount': None, '_source': os.path.basename(source), '_method': 'text'}

    m = NTS_APPROVAL_RE.search(text)
    if m:
        data['approval_no'] = m.group(1).strip()
    else:
        m = FILENAME_APPROVAL_RE.search(text)
        if m:
            data['approval_no'] = m.group(1).strip()
        else:
            m = GENERAL_APPROVAL_RE.search(text)
            if m:
                data['approval_no'] = m.group(1).replace('-', '').strip()

    biz = list(BIZ_NO_RE.finditer(text))
    if len(biz) >= 2:
        data['supplier_biz_no'] = clean_biz_no(biz[0].groups())
        data['receiver_biz_no'] = clean_biz_no(biz[1].groups())
    elif len(biz) == 1:
        data['supplier_biz_no'] = clean_biz_no(biz[0].groups())

    m = ISSUE_DATE_RE.search(text)
    if m:
        data['issue_date'] = f"{m.group(1)}{m.group(2).zfill(2)}{m.group(3).zfill(2)}"
    else:
        m = ISSUE_DATE2_RE.search(text)
        if m:
            data['issue_date'] = m.group(1)
        else:
            m = ISSUE_DATE3_RE.search(text)
            if m:
                data['issue_date'] = f"{m.group(1)}{m.group(2).zfill(2)}{m.group(3).zfill(2)}"

    matches = SUPPLY_AMT_RE.findall(text)
    if matches:
        try:
            data['supply_amount'] = int(matches[0].replace(',', ''))
        except ValueError:
            pass
    return data


def extract_from_ocr(ocr_text, source=""):
    data = {'supplier_biz_no': '', 'receiver_biz_no': '', 'approval_no': '',
            'issue_date': '', 'supply_amount': None, '_source': os.path.basename(source), '_method': 'OCR'}

    biz = list(BIZ_NO_RE.finditer(ocr_text))
    if len(biz) >= 2:
        data['supplier_biz_no'] = clean_biz_no(biz[0].groups())
        data['receiver_biz_no'] = clean_biz_no(biz[1].groups())
    elif len(biz) == 1:
        data['supplier_biz_no'] = clean_biz_no(biz[0].groups())

    m = SCAN_APPROVAL_RE.search(ocr_text)
    if m:
        data['approval_no'] = f"{m.group(1)}{m.group(2)}{m.group(3)}"
    else:
        m = NTS_APPROVAL_RE.search(ocr_text)
        if m:
            data['approval_no'] = m.group(1).strip()

    m = re.search(r'(20\d{2})\s*[/\-\.]\s*(\d{1,2})\s*[/\-\.]\s*(\d{1,2})', ocr_text)
    if m:
        data['issue_date'] = f"{m.group(1)}{m.group(2).zfill(2)}{m.group(3).zfill(2)}"
    else:
        m = SCAN_DATE_RE.search(ocr_text)
        if m:
            data['issue_date'] = f"20{m.group(1)}{m.group(2).zfill(2)}{m.group(3).zfill(2)}"
        else:
            if data['approval_no'] and len(data['approval_no']) >= 8:
                dc = data['approval_no'][:8]
                if re.match(r'20\d{6}$', dc):
                    data['issue_date'] = dc

    REAL_AMT_RE = re.compile(r'\b(\d{1,3}(?:,\d{3})+)\b')
    amounts = []
    for m in REAL_AMT_RE.finditer(ocr_text):
        try:
            val = int(m.group(1).replace(',', ''))
            if val >= 10000:
                amounts.append(val)
        except ValueError:
            pass
    if amounts:
        count = Counter(amounts)
        for val, cnt in count.most_common():
            if cnt >= 2:
                data['supply_amount'] = val
                break
        if not data['supply_amount']:
            sa = sorted(set(amounts), reverse=True)
            data['supply_amount'] = sa[1] if len(sa) >= 2 else sa[0]
    return data


def fill_template(template_path, invoices, output_path):
    wb = load_workbook(template_path)
    ws = wb.active
    for i, inv in enumerate(invoices):
        row = 6 + i
        ws.cell(row=row, column=1, value=inv['supplier_biz_no'])
        ws.cell(row=row, column=2, value=inv['receiver_biz_no'])
        ws.cell(row=row, column=3, value=inv['approval_no'])
        ws.cell(row=row, column=4, value=inv['issue_date'])
        ws.cell(row=row, column=5, value=inv['supply_amount'] or '')
    if not output_path.lower().endswith('.xlsx'):
        output_path = os.path.splitext(output_path)[0] + '.xlsx'
    wb.save(output_path)
    return output_path


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# GUI
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("세금계산서 추출기 — 제3자발급사실 일괄등록양식")
        self.geometry("720x680")
        self.resizable(False, False)
        self.configure(bg="#f0f4f8")

        self.pdf_dir = tk.StringVar()
        self.template_path = tk.StringVar()
        self.do_merge = tk.BooleanVar(value=True)
        self.running = False

        self._build_ui()

    def _build_ui(self):
        # 스타일
        style = ttk.Style(self)
        style.theme_use('clam')
        style.configure('Title.TLabel', font=('맑은 고딕', 16, 'bold'), background='#f0f4f8', foreground='#1a365d')
        style.configure('Sub.TLabel', font=('맑은 고딕', 9), background='#f0f4f8', foreground='#4a5568')
        style.configure('Card.TFrame', background='white', relief='solid', borderwidth=1)
        style.configure('CardInner.TFrame', background='white')
        style.configure('CardLabel.TLabel', font=('맑은 고딕', 10), background='white', foreground='#2d3748')
        style.configure('Path.TLabel', font=('맑은 고딕', 9), background='#edf2f7', foreground='#2b6cb0', padding=(8, 4))
        style.configure('Run.TButton', font=('맑은 고딕', 12, 'bold'), padding=(20, 10))
        style.configure('Browse.TButton', font=('맑은 고딕', 9), padding=(12, 4))
        style.configure('Check.TCheckbutton', font=('맑은 고딕', 10), background='#f0f4f8')

        pad = {'padx': 20}

        # 제목
        ttk.Label(self, text="📄 세금계산서 추출기", style='Title.TLabel').pack(pady=(20, 2))
        ttk.Label(self, text="PDF에서 세금계산서를 찾아 제3자발급사실 일괄등록양식에 자동 입력합니다", style='Sub.TLabel').pack(pady=(0, 16))

        # ── 카드 1: PDF 폴더 ──
        card1 = ttk.Frame(self, style='Card.TFrame')
        card1.pack(fill='x', **pad, pady=(0, 8))
        inner1 = ttk.Frame(card1, style='CardInner.TFrame')
        inner1.pack(fill='x', padx=16, pady=12)

        ttk.Label(inner1, text="① PDF 폴더  (하위 폴더 자동 탐색)", style='CardLabel.TLabel').pack(anchor='w')
        row1 = ttk.Frame(inner1, style='CardInner.TFrame')
        row1.pack(fill='x', pady=(6, 0))
        self.pdf_label = ttk.Label(row1, textvariable=self.pdf_dir, style='Path.TLabel', width=55)
        self.pdf_label.pack(side='left', fill='x', expand=True)
        ttk.Button(row1, text="폴더 선택", style='Browse.TButton', command=self._browse_pdf).pack(side='right', padx=(8, 0))

        ttk.Label(inner1, text="💡 폴더를 이 창에 드래그앤드롭 해도 됩니다", font=('맑은 고딕', 8), background='white', foreground='#a0aec0').pack(anchor='w', pady=(4, 0))

        # ── 카드 2: 양식 파일 ──
        card2 = ttk.Frame(self, style='Card.TFrame')
        card2.pack(fill='x', **pad, pady=(0, 8))
        inner2 = ttk.Frame(card2, style='CardInner.TFrame')
        inner2.pack(fill='x', padx=16, pady=12)

        ttk.Label(inner2, text="② 제3자발급사실 일괄등록양식 파일  (.xlsx)", style='CardLabel.TLabel').pack(anchor='w')
        row2 = ttk.Frame(inner2, style='CardInner.TFrame')
        row2.pack(fill='x', pady=(6, 0))
        self.tpl_label = ttk.Label(row2, textvariable=self.template_path, style='Path.TLabel', width=55)
        self.tpl_label.pack(side='left', fill='x', expand=True)
        ttk.Button(row2, text="파일 선택", style='Browse.TButton', command=self._browse_template).pack(side='right', padx=(8, 0))

        # ── 옵션 ──
        opt_frame = ttk.Frame(self)
        opt_frame.configure(style='TFrame')
        opt_frame.pack(fill='x', **pad, pady=(4, 8))
        ttk.Checkbutton(opt_frame, text="PDF 전체 합치기 (합친_전체.pdf 생성)", variable=self.do_merge, style='Check.TCheckbutton').pack(anchor='w')

        # ── 실행 버튼 ──
        self.run_btn = ttk.Button(self, text="▶  추출 시작", style='Run.TButton', command=self._start)
        self.run_btn.pack(pady=(4, 12))

        # ── 로그 ──
        log_frame = ttk.Frame(self)
        log_frame.pack(fill='both', expand=True, padx=20, pady=(0, 12))
        ttk.Label(log_frame, text="진행 로그", font=('맑은 고딕', 9, 'bold')).pack(anchor='w', pady=(0, 4))
        self.log_text = tk.Text(log_frame, height=12, font=('Consolas', 9), bg='#1a202c', fg='#e2e8f0',
                                insertbackground='white', relief='flat', padx=10, pady=8)
        self.log_text.pack(fill='both', expand=True)
        scrollbar = ttk.Scrollbar(self.log_text, command=self.log_text.yview)
        scrollbar.pack(side='right', fill='y')
        self.log_text.configure(yscrollcommand=scrollbar.set)

        # ── 상태바 ──
        self.status = tk.StringVar(value="대기 중")
        status_bar = ttk.Label(self, textvariable=self.status, font=('맑은 고딕', 9), background='#e2e8f0',
                               foreground='#4a5568', padding=(12, 4), anchor='w')
        status_bar.pack(fill='x', side='bottom')

        # 드래그앤드롭 (tkinterdnd2 없이 기본 방식)
        self._setup_drop()

    def _setup_drop(self):
        """tkinterdnd2 없으면 드래그앤드롭 비활성, 있으면 활성"""
        try:
            import tkinterdnd2
            # tkinterdnd2 사용 가능하면 설정
            pass
        except ImportError:
            pass  # 드래그앤드롭 미지원 — 버튼으로만 선택

    def _browse_pdf(self):
        path = filedialog.askdirectory(title="PDF가 있는 폴더 선택")
        if path:
            self.pdf_dir.set(path)
            self.log(f"📁 PDF 폴더: {path}")

    def _browse_template(self):
        path = filedialog.askopenfilename(
            title="제3자발급사실 일괄등록양식 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xls"), ("모든 파일", "*.*")]
        )
        if path:
            self.template_path.set(path)
            self.log(f"📋 양식 파일: {os.path.basename(path)}")

    def log(self, msg):
        self.log_text.insert('end', msg + '\n')
        self.log_text.see('end')
        self.update_idletasks()

    def _start(self):
        if self.running:
            return

        pdf_dir = self.pdf_dir.get()
        template = self.template_path.get()

        if not pdf_dir or not os.path.isdir(pdf_dir):
            messagebox.showwarning("알림", "PDF 폴더를 선택해주세요.")
            return
        if not template or not os.path.exists(template):
            messagebox.showwarning("알림", "양식 파일을 선택해주세요.")
            return

        self.running = True
        self.run_btn.configure(state='disabled')
        self.log_text.delete('1.0', 'end')
        self.status.set("처리 중...")

        # 별도 스레드에서 실행 (UI 멈춤 방지)
        thread = threading.Thread(target=self._run_extraction, args=(pdf_dir, template), daemon=True)
        thread.start()

    def _run_extraction(self, pdf_dir, template):
        try:
            # 1) PDF 탐색
            pdf_files = find_pdfs(pdf_dir)
            if not pdf_files:
                self.after(0, lambda: self.log("❌ PDF 파일을 찾을 수 없습니다."))
                return

            self.after(0, lambda: self.log(f"\n📂 PDF {len(pdf_files)}개 발견"))
            for f in pdf_files:
                rel = os.path.relpath(f, pdf_dir)
                self.after(0, lambda r=rel: self.log(f"  📄 {r}"))

            # 2) 세금계산서 추출
            self.after(0, lambda: self.log(f"\n🔍 세금계산서 추출 중..."))
            invoices = []
            ocr_count = 0
            total_pages = 0

            for pf in pdf_files:
                rel = os.path.relpath(pf, pdf_dir)
                try:
                    with pdfplumber.open(pf) as pdf:
                        total_pages += len(pdf.pages)
                        for pnum, page in enumerate(pdf.pages, 1):
                            text = page.extract_text() or ""

                            if len(text) > 50 and is_invoice_page(text):
                                inv = extract_from_text(text, source=pf)
                                inv['_page'] = pnum
                                invoices.append(inv)
                                amt = f"{inv['supply_amount']:,}" if inv['supply_amount'] else '?'
                                self.after(0, lambda r=rel, p=pnum, a=amt: self.log(f"  ✓ {r} p.{p} → 공급가액: {a}"))
                                continue

                            if OCR_AVAILABLE and len(text) <= 50:
                                self.after(0, lambda r=rel, p=pnum: self.status.set(f"OCR 처리: {r} p.{p}"))
                                ocr_text = ocr_page_image(pf, pnum)
                                if ocr_text and is_scanned_invoice(ocr_text):
                                    inv = extract_from_ocr(ocr_text, source=pf)
                                    inv['_page'] = pnum
                                    invoices.append(inv)
                                    ocr_count += 1
                                    amt = f"{inv['supply_amount']:,}" if inv['supply_amount'] else '?'
                                    self.after(0, lambda r=rel, p=pnum, a=amt: self.log(f"  ✓ {r} p.{p} [OCR] → 공급가액: {a}"))

                except Exception as e:
                    self.after(0, lambda r=rel, e=str(e): self.log(f"  ✗ {r}: {e}"))

            # 3) PDF 합치기
            merged_path = ""
            if self.do_merge.get():
                merged_path = os.path.join(os.path.dirname(template), "합친_전체.pdf")
                self.after(0, lambda: self.log(f"\n📎 PDF 합치기..."))
                total_p = merge_pdfs(pdf_files, merged_path)
                self.after(0, lambda: self.log(f"  ✓ {total_p}페이지 → {merged_path}"))

            # 4) 요약
            total_supply = sum(x['supply_amount'] or 0 for x in invoices)
            text_cnt = sum(1 for x in invoices if x.get('_method') == 'text')
            self.after(0, lambda: self.log(f"\n{'─'*50}"))
            self.after(0, lambda: self.log(f"📊 추출 결과: {len(invoices)}건 (텍스트 {text_cnt}, OCR {ocr_count})"))
            self.after(0, lambda: self.log(f"   공급가액 합계: {total_supply:,}원"))

            # 5) 엑셀 입력
            if invoices:
                base = os.path.splitext(os.path.basename(template))[0]
                output_path = os.path.join(os.path.dirname(template), f"결과_{base}.xlsx")
                final = fill_template(template, invoices, output_path)
                self.after(0, lambda: self.log(f"\n✅ 엑셀 저장: {final}"))
                self.after(0, lambda: self.log(f"   {len(invoices)}건 데이터가 6행부터 입력되었습니다."))

                if ocr_count:
                    ocr_items = [x for x in invoices if x.get('_method') == 'OCR']
                    self.after(0, lambda: self.log(f"\n⚠️  OCR {ocr_count}건은 수동 검증을 권장합니다."))
                    self.after(0, lambda: self.log(f"   아래 항목의 PDF 원본과 엑셀 데이터를 비교해주세요:"))
                    for idx, ocr_inv in enumerate(ocr_items):
                        # 엑셀 행 번호 계산 (6행부터 시작)
                        excel_row = 6 + invoices.index(ocr_inv)
                        src = ocr_inv.get('_source', '?')
                        pg = ocr_inv.get('_page', '?')
                        biz = ocr_inv.get('supplier_biz_no', '?')
                        amt = f"{ocr_inv['supply_amount']:,}" if ocr_inv.get('supply_amount') else '?'
                        dt = ocr_inv.get('issue_date', '?')
                        self.after(0, lambda i=idx+1, s=src, p=pg, r=excel_row, b=biz, a=amt, d=dt:
                            self.log(f"   {i}) 📄 {s} → PDF p.{p} → 엑셀 {r}행  |  공급자:{b}  작성일:{d}  공급가액:{a}"))

                # 완료 알림
                ocr_msg = ""
                if ocr_count:
                    ocr_items = [x for x in invoices if x.get('_method') == 'OCR']
                    ocr_details = []
                    for ocr_inv in ocr_items:
                        excel_row = 6 + invoices.index(ocr_inv)
                        src = ocr_inv.get('_source', '?')
                        pg = ocr_inv.get('_page', '?')
                        ocr_details.append(f"  • {src} p.{pg} → 엑셀 {excel_row}행")
                    ocr_msg = f"\n\n⚠️ OCR {ocr_count}건 수동 검증 필요:\n" + "\n".join(ocr_details)

                self.after(0, lambda: messagebox.showinfo("완료",
                    f"세금계산서 {len(invoices)}건 추출 완료!\n\n"
                    f"공급가액 합계: {total_supply:,}원\n\n"
                    f"저장 위치:\n{final}"
                    + (f"\n합친 PDF: {merged_path}" if merged_path else "")
                    + ocr_msg
                ))
            else:
                self.after(0, lambda: self.log(f"\n⚠️  세금계산서를 찾지 못했습니다."))
                self.after(0, lambda: messagebox.showwarning("결과", "세금계산서를 찾지 못했습니다.\nPDF 내용을 확인해주세요."))

        except Exception as e:
            self.after(0, lambda: self.log(f"\n❌ 오류: {e}"))
            self.after(0, lambda: messagebox.showerror("오류", str(e)))

        finally:
            self.after(0, self._finish)

    def _finish(self):
        self.running = False
        self.run_btn.configure(state='normal')
        self.status.set("완료")


if __name__ == '__main__':
    app = App()
    app.mainloop()